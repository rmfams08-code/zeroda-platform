# zeroda_reflex/utils/excel_export.py
# ══════════════════════════════════════════════════════════════
#  Excel 다운로드 유틸리티 (Phase 3)
#  - openpyxl로 Excel 파일을 메모리에서 생성하여 bytes로 반환
#  - 각 역할별 state에서 호출하여 rx.download()로 전달
# ══════════════════════════════════════════════════════════════
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 미설치 — Excel 다운로드 기능 비활성화")


# ══════════════════════════════════════════
#  공통 스타일 정의
# ══════════════════════════════════════════

# 헤더 스타일 — 진한 파란 배경 + 흰 글자
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF") if HAS_OPENPYXL else None
HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid") if HAS_OPENPYXL else None
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True) if HAS_OPENPYXL else None

# 데이터 셀 스타일
DATA_FONT = Font(name="맑은 고딕", size=10) if HAS_OPENPYXL else None
DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center") if HAS_OPENPYXL else None
DATA_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center") if HAS_OPENPYXL else None
DATA_ALIGN_CENTER = Alignment(horizontal="center", vertical="center") if HAS_OPENPYXL else None

# 테두리
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if HAS_OPENPYXL else None

# 제목 스타일 — 큰 제목 행
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1A73E8") if HAS_OPENPYXL else None

# 소계/합계 행 스타일
TOTAL_FONT = Font(name="맑은 고딕", size=10, bold=True) if HAS_OPENPYXL else None
TOTAL_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid") if HAS_OPENPYXL else None


def _apply_header_style(ws, row: int, col_count: int):
    """헤더 행에 스타일 적용"""
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_style(ws, row: int, col_count: int, aligns: list[str] | None = None):
    """데이터 행에 스타일 적용

    Args:
        aligns: 각 컬럼별 정렬 ("L"=왼쪽, "R"=오른쪽, "C"=가운데)
    """
    align_map = {"L": DATA_ALIGN_LEFT, "R": DATA_ALIGN_RIGHT, "C": DATA_ALIGN_CENTER}
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if aligns and c <= len(aligns):
            cell.alignment = align_map.get(aligns[c - 1], DATA_ALIGN_LEFT)
        else:
            cell.alignment = DATA_ALIGN_LEFT


def _auto_width(ws, col_count: int, min_width: int = 10, max_width: int = 40):
    """컬럼 너비 자동 조절 — 한글 폭 고려"""
    for c in range(1, col_count + 1):
        letter = get_column_letter(c)
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            cell = row[0]
            if cell.value:
                # 한글은 약 2칸, 영문/숫자는 1칸으로 계산
                val = str(cell.value)
                char_len = sum(2 if ord(ch) > 127 else 1 for ch in val)
                max_len = max(max_len, char_len + 2)
        ws.column_dimensions[letter].width = min(max_len, max_width)


def _create_workbook_with_title(title: str, subtitle: str = "") -> tuple:
    """제목이 포함된 워크북 생성

    Returns:
        (workbook, worksheet, 데이터 시작 행 번호)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel 시트명 31자 제한

    # 제목 행
    ws.cell(row=1, column=1, value=f"ZERODA — {title}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # 부제목/생성일시
    if not subtitle:
        subtitle = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.cell(row=2, column=1, value=subtitle)
    ws.cell(row=2, column=1).font = Font(name="맑은 고딕", size=9, color="666666")

    return wb, ws, 4  # 4행부터 데이터 시작


def build_excel(
    title: str,
    headers: list[str],
    rows: list[list],
    aligns: list[str] | None = None,
    subtitle: str = "",
    totals: list | None = None,
) -> bytes:
    """범용 Excel 생성 함수 — bytes로 반환

    모든 Excel 다운로드의 핵심 함수입니다.

    Args:
        title: 시트 제목 (예: "수거데이터")
        headers: 컬럼 헤더 리스트 (예: ["날짜", "학교명", "품목", "수거량(kg)"])
        rows: 데이터 행 리스트 (예: [["2026-04-01", "가나초등", "음식물", 12.5], ...])
        aligns: 컬럼별 정렬 ("L","R","C") 리스트. None이면 전부 왼쪽
        subtitle: 부제목 (비워두면 생성일시 자동)
        totals: 합계 행 리스트 (예: ["합계", "", "", 125.0])

    Returns:
        Excel 파일의 bytes 데이터

    사용 예 (state 메서드 내):
        data = build_excel("수거데이터", headers, rows, aligns=["C","L","C","R"])
        return rx.download(data=data, filename="수거데이터_2026-04.xlsx")
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 미설치 — Excel 생성 불가")
        return b""

    wb, ws, start_row = _create_workbook_with_title(title, subtitle)
    col_count = len(headers)

    # 헤더 작성
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    _apply_header_style(ws, start_row, col_count)

    # 데이터 행 작성
    for r_idx, row_data in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        _apply_data_style(ws, r_idx, col_count, aligns)

    # 합계 행 (있으면)
    if totals:
        total_row = start_row + 1 + len(rows)
        for c_idx, val in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=c_idx, value=val)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            if aligns and c_idx <= len(aligns):
                align_map = {"L": DATA_ALIGN_LEFT, "R": DATA_ALIGN_RIGHT, "C": DATA_ALIGN_CENTER}
                cell.alignment = align_map.get(aligns[c_idx - 1], DATA_ALIGN_LEFT)

    # 컬럼 너비 자동 조절
    _auto_width(ws, col_count)

    # bytes로 변환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════
#  역할별 Excel 생성 함수 (7종)
# ══════════════════════════════════════════

def export_collection_data(data: list[dict], year: str, month: str) -> bytes:
    """3-1. 수거데이터 Excel (본사/업체 공용)

    Args:
        data: 수거 기록 딕셔너리 리스트
        year, month: 필터링된 연/월
    """
    headers = ["수거일자", "업체", "학교명", "품목", "수거량(kg)", "기사", "상태"]
    aligns = ["C", "L", "L", "C", "R", "L", "C"]
    rows = []
    total_weight = 0.0
    for r in data:
        w = float(r.get("weight", 0) or 0)
        total_weight += w
        rows.append([
            r.get("collect_date", ""),
            r.get("vendor", ""),
            r.get("school_name", ""),
            r.get("item_type", ""),
            round(w, 1),
            r.get("driver", ""),
            r.get("status", ""),
        ])
    totals = ["합계", "", "", "", round(total_weight, 1), f"{len(rows)}건", ""]
    return build_excel(
        title="수거데이터",
        headers=headers,
        rows=rows,
        aligns=aligns,
        subtitle=f"{year}년 {month}월 수거 데이터",
        totals=totals,
    )


def export_settlement(
    detail_data: list[dict],
    expenses: list[dict],
    summary: dict,
    year: str,
    month: str,
    vendor: str = "",
) -> bytes:
    """3-2. 월말정산 Excel — 2시트 (수입내역 + 지출내역)

    Args:
        detail_data: get_settlement_detail() 결과
            [{name, cust_type, item_type, weight, unit_price, supply, vat, total, is_fixed_fee}, ...]
        expenses: 지출 내역
            [{id, item, amount, pay_date, memo}, ...]
        summary: 정산 요약
            {total_revenue, total_expense, net_profit}
        year, month: 정산 기간
        vendor: 업체명
    """
    if not HAS_OPENPYXL:
        logger.error("openpyxl 미설치 — Excel 생성 불가")
        return b""

    from openpyxl.worksheet.filters import AutoFilter

    # ── 구분별 배경색 정의 ──
    cust_colors = {
        "학교":              "DAEEF3",
        "관공서":            "E2EFDA",
        "기업":              "FCE4D6",
        "일반업장":          "F2DCDB",
        "기타":              "F2F2F2",
        "기타1(면세사업장)":  "F2F2F2",
        "기타2(부가세포함)":  "F2F2F2",
    }
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    blue_price_font = Font(name="맑은 고딕", size=10, color="0000FF")
    red_font = Font(name="맑은 고딕", size=10, color="FF0000")
    bold_font_12 = Font(name="맑은 고딕", size=12, bold=True)
    double_border = Border(
        top=Side(style="double"), bottom=Side(style="double"),
        left=Side(style="thin"), right=Side(style="thin"),
    )
    mm = str(month).zfill(2)

    wb = Workbook()

    # ════════════════════════════════════════
    #  시트1: 수입내역
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "수입내역"

    # 제목
    ws1.merge_cells("A1:I1")
    title_cell = ws1.cell(row=1, column=1, value=f"( {mm} )月 월말정산서 — 수입내역")
    title_cell.font = Font(name="맑은 고딕", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 부제목
    ws1.merge_cells("A2:I2")
    sub_cell = ws1.cell(row=2, column=1, value=f"{year}년 {mm}월 | {vendor}")
    sub_cell.font = Font(name="맑은 고딕", size=9, color="666666")

    # 헤더 (Row 3)
    rev_headers = ["No", "구분", "거래처명", "품목", "누적수거량(kg)", "단가(원)", "공급가(원)", "부가세(원)", "정산금액(원)"]
    rev_widths = [6, 12, 16, 12, 14, 12, 14, 12, 14]
    for c, h in enumerate(rev_headers, 1):
        ws1.cell(row=3, column=c, value=h)
    _apply_header_style(ws1, 3, 9)

    # 열 너비
    for c, w in enumerate(rev_widths, 1):
        ws1.column_dimensions[get_column_letter(c)].width = w

    # 데이터 행 — 거래처별 그룹핑
    # 같은 cust_type 끼리 묶고, 같은 거래처의 다품목은 B/C 병합
    row_idx = 4
    no = 1

    # cust_type별 그룹
    from itertools import groupby
    type_subtotal_rows: list[int] = []  # 소계 행 번호 (총합계 SUM용)

    for ct, ct_group in groupby(detail_data, key=lambda x: x["cust_type"]):
        ct_list = list(ct_group)
        ct_fill = PatternFill(
            start_color=cust_colors.get(ct, "FFFFFF"),
            end_color=cust_colors.get(ct, "FFFFFF"),
            fill_type="solid",
        )
        ct_start_row = row_idx

        # 거래처별 그룹
        for name, name_group in groupby(ct_list, key=lambda x: x["name"]):
            items = list(name_group)
            name_start = row_idx

            for item in items:
                ws1.cell(row=row_idx, column=1, value=no)
                ws1.cell(row=row_idx, column=2, value=ct)
                ws1.cell(row=row_idx, column=3, value=name)
                ws1.cell(row=row_idx, column=4, value=item["item_type"])
                ws1.cell(row=row_idx, column=5, value=round(item["weight"], 1))
                ws1.cell(row=row_idx, column=5).number_format = "#,##0.0"

                if item.get("is_fixed_fee"):
                    ws1.cell(row=row_idx, column=6, value=0)
                    ws1.cell(row=row_idx, column=7, value=int(item["supply"]))
                else:
                    ws1.cell(row=row_idx, column=6, value=int(item["unit_price"]))
                    # 공급가 수식: =E*F
                    ws1.cell(row=row_idx, column=7, value=f"=E{row_idx}*F{row_idx}")

                ws1.cell(row=row_idx, column=6).font = blue_price_font
                ws1.cell(row=row_idx, column=6).number_format = "#,##0"
                ws1.cell(row=row_idx, column=7).number_format = "#,##0"

                # 부가세 수식
                tax_free_types = ("학교", "기타1(면세사업장)", "기타")
                if ct in tax_free_types:
                    ws1.cell(row=row_idx, column=8, value=0)
                else:
                    ws1.cell(row=row_idx, column=8, value=f"=G{row_idx}*0.1")
                ws1.cell(row=row_idx, column=8).number_format = "#,##0"

                # 정산금액 수식: =G+H
                ws1.cell(row=row_idx, column=9, value=f"=G{row_idx}+H{row_idx}")
                ws1.cell(row=row_idx, column=9).number_format = "#,##0"

                # 배경색 + 테두리
                for c in range(1, 10):
                    cell = ws1.cell(row=row_idx, column=c)
                    cell.fill = ct_fill
                    cell.border = THIN_BORDER
                    if c != 6:  # col6 already has blue_price_font
                        cell.font = DATA_FONT
                    if c <= 4:
                        cell.alignment = DATA_ALIGN_CENTER
                    else:
                        cell.alignment = DATA_ALIGN_RIGHT

                no += 1
                row_idx += 1

            # 같은 거래처 다품목이면 B(구분), C(거래처명) 셀 병합
            if len(items) > 1:
                ws1.merge_cells(
                    start_row=name_start, start_column=2,
                    end_row=row_idx - 1, end_column=2,
                )
                ws1.merge_cells(
                    start_row=name_start, start_column=3,
                    end_row=row_idx - 1, end_column=3,
                )

        # 구분 소계 행
        ct_end_row = row_idx - 1
        ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        ws1.cell(row=row_idx, column=1, value=f"{ct} 소계")
        ws1.cell(row=row_idx, column=1).alignment = DATA_ALIGN_CENTER

        # E열: 수거량 합
        ws1.cell(row=row_idx, column=5, value=f"=SUM(E{ct_start_row}:E{ct_end_row})")
        ws1.cell(row=row_idx, column=5).number_format = "#,##0.0"
        ws1.cell(row=row_idx, column=6, value="")
        # G열: 공급가 합
        ws1.cell(row=row_idx, column=7, value=f"=SUM(G{ct_start_row}:G{ct_end_row})")
        ws1.cell(row=row_idx, column=7).number_format = "#,##0"
        # H열: 부가세 합
        ws1.cell(row=row_idx, column=8, value=f"=SUM(H{ct_start_row}:H{ct_end_row})")
        ws1.cell(row=row_idx, column=8).number_format = "#,##0"
        # I열: 정산금액 합
        ws1.cell(row=row_idx, column=9, value=f"=SUM(I{ct_start_row}:I{ct_end_row})")
        ws1.cell(row=row_idx, column=9).number_format = "#,##0"

        for c in range(1, 10):
            cell = ws1.cell(row=row_idx, column=c)
            cell.font = TOTAL_FONT
            cell.fill = ct_fill
            cell.border = THIN_BORDER
            if c >= 5:
                cell.alignment = DATA_ALIGN_RIGHT

        type_subtotal_rows.append(row_idx)
        row_idx += 1

    # 총 합계 행
    total_row_rev = row_idx
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
    ws1.cell(row=row_idx, column=1, value="총 합계")
    ws1.cell(row=row_idx, column=1).alignment = DATA_ALIGN_CENTER

    if type_subtotal_rows:
        e_refs = "+".join(f"E{r}" for r in type_subtotal_rows)
        g_refs = "+".join(f"G{r}" for r in type_subtotal_rows)
        h_refs = "+".join(f"H{r}" for r in type_subtotal_rows)
        i_refs = "+".join(f"I{r}" for r in type_subtotal_rows)
        ws1.cell(row=row_idx, column=5, value=f"={e_refs}")
        ws1.cell(row=row_idx, column=7, value=f"={g_refs}")
        ws1.cell(row=row_idx, column=8, value=f"={h_refs}")
        ws1.cell(row=row_idx, column=9, value=f"={i_refs}")
    else:
        for c in [5, 7, 8, 9]:
            ws1.cell(row=row_idx, column=c, value=0)

    ws1.cell(row=row_idx, column=5).number_format = "#,##0.0"
    ws1.cell(row=row_idx, column=6, value="")
    for c in [7, 8, 9]:
        ws1.cell(row=row_idx, column=c).number_format = "#,##0"

    for c in range(1, 10):
        cell = ws1.cell(row=row_idx, column=c)
        cell.font = TOTAL_FONT
        cell.fill = yellow_fill
        cell.border = THIN_BORDER
        if c >= 5:
            cell.alignment = DATA_ALIGN_RIGHT

    # AutoFilter
    if row_idx > 3:
        ws1.auto_filter.ref = f"A3:I{row_idx}"

    # ════════════════════════════════════════
    #  시트2: 지출내역
    # ════════════════════════════════════════
    ws2 = wb.create_sheet(title="지출내역")

    # 제목
    ws2.merge_cells("A1:E1")
    t2 = ws2.cell(row=1, column=1, value=f"( {mm} )月 월말정산서 — 지출내역")
    t2.font = Font(name="맑은 고딕", size=14, bold=True)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    # 부제목
    ws2.merge_cells("A2:E2")
    s2 = ws2.cell(row=2, column=1, value=f"{year}년 {mm}월 | {vendor}")
    s2.font = Font(name="맑은 고딕", size=9, color="666666")

    # 헤더 (Row 3)
    exp_headers = ["No", "지출항목", "금액(원)", "결제일", "비고"]
    exp_widths = [6, 18, 16, 12, 18]
    for c, h in enumerate(exp_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    _apply_header_style(ws2, 3, 5)

    for c, w in enumerate(exp_widths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # 데이터 행
    exp_row = 4
    if expenses:
        for i, exp in enumerate(expenses, 1):
            ws2.cell(row=exp_row, column=1, value=i)
            ws2.cell(row=exp_row, column=2, value=str(exp.get("item", "")))
            amt = abs(float(exp.get("amount", 0) or 0))
            ws2.cell(row=exp_row, column=3, value=int(amt))
            ws2.cell(row=exp_row, column=3).font = red_font
            ws2.cell(row=exp_row, column=3).number_format = "#,##0"
            ws2.cell(row=exp_row, column=4, value=str(exp.get("pay_date", "")))
            ws2.cell(row=exp_row, column=5, value=str(exp.get("memo", "")))

            for c in range(1, 6):
                cell = ws2.cell(row=exp_row, column=c)
                cell.border = THIN_BORDER
                cell.font = cell.font if c == 3 else DATA_FONT
                if c == 3:
                    cell.alignment = DATA_ALIGN_RIGHT
                elif c in (1, 4):
                    cell.alignment = DATA_ALIGN_CENTER
                else:
                    cell.alignment = DATA_ALIGN_LEFT
            exp_row += 1
    else:
        # 빈 템플릿 10행
        for i in range(10):
            for c in range(1, 6):
                cell = ws2.cell(row=exp_row, column=c)
                cell.border = THIN_BORDER
            exp_row += 1

    last_exp_data = exp_row - 1

    # 지출 합계 행
    exp_total_row = exp_row
    ws2.merge_cells(start_row=exp_row, start_column=1, end_row=exp_row, end_column=2)
    ws2.cell(row=exp_row, column=1, value="지출 합계")
    ws2.cell(row=exp_row, column=1).alignment = DATA_ALIGN_CENTER
    ws2.cell(row=exp_row, column=3, value=f"=SUM(C4:C{last_exp_data})")
    ws2.cell(row=exp_row, column=3).number_format = "#,##0"

    for c in range(1, 6):
        cell = ws2.cell(row=exp_row, column=c)
        cell.font = TOTAL_FONT
        cell.fill = yellow_fill
        cell.border = THIN_BORDER

    # 순수익 행 (합계 +2줄)
    profit_row = exp_row + 2
    ws2.merge_cells(start_row=profit_row, start_column=1, end_row=profit_row, end_column=2)
    ws2.cell(row=profit_row, column=1, value="순수익 (매출 - 지출)")
    ws2.cell(row=profit_row, column=1).alignment = DATA_ALIGN_CENTER
    ws2.cell(row=profit_row, column=1).font = bold_font_12
    # 시트간 참조: 수입내역 총합계 I열 - 지출합계 C열
    ws2.cell(row=profit_row, column=3,
             value=f"=수입내역!I{total_row_rev}-C{exp_total_row}")
    ws2.cell(row=profit_row, column=3).number_format = "#,##0"
    ws2.cell(row=profit_row, column=3).font = bold_font_12

    for c in range(1, 6):
        cell = ws2.cell(row=profit_row, column=c)
        cell.border = double_border

    # ── 바이트로 반환 ──
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_carbon_data(data: dict, ranking: list[dict], year: str) -> bytes:
    """3-3. 탄소감축 Excel (본사/교육청 공용)

    Args:
        data: 탄소감축 요약 (food_kg, recycle_kg, carbon_reduced 등)
        ranking: 학교별 탄소감축 순위
        year: 연도
    """
    # 요약 시트
    headers = ["항목", "값", "단위"]
    aligns = ["L", "R", "L"]
    summary_rows = [
        ["음식물 수거량", float(data.get("food_kg", 0) or 0), "kg"],
        ["재활용 수거량", float(data.get("recycle_kg", 0) or 0), "kg"],
        ["일반 수거량", float(data.get("general_kg", 0) or 0), "kg"],
        ["총 수거량", float(data.get("total_kg", 0) or 0), "kg"],
        ["탄소 감축량", float(data.get("carbon_reduced", 0) or 0), "kg CO₂"],
        ["나무 환산", int(float(data.get("tree_equivalent", 0) or 0)), "그루"],
    ]

    if not HAS_OPENPYXL:
        return b""

    wb, ws, start = _create_workbook_with_title("탄소감축 현황", f"{year}년 탄소감축 분석")
    col_count = 3

    # 요약 테이블
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    _apply_header_style(ws, start, col_count)
    for r_idx, row_data in enumerate(summary_rows, start + 1):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
        _apply_data_style(ws, r_idx, col_count, aligns)

    # 학교별 순위 테이블
    rank_start = start + len(summary_rows) + 3
    ws.cell(row=rank_start - 1, column=1, value="학교별 탄소감축 순위")
    ws.cell(row=rank_start - 1, column=1).font = Font(name="맑은 고딕", size=12, bold=True)
    rank_headers = ["순위", "학교명", "수거량(kg)", "탄소감축(kg CO₂)"]
    for c, h in enumerate(rank_headers, 1):
        ws.cell(row=rank_start, column=c, value=h)
    _apply_header_style(ws, rank_start, 4)

    for idx, r in enumerate(ranking, 1):
        row_num = rank_start + idx
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=r.get("school_name", ""))
        ws.cell(row=row_num, column=3, value=float(r.get("total_weight", 0) or 0))
        ws.cell(row=row_num, column=4, value=float(r.get("carbon", 0) or 0))
        _apply_data_style(ws, row_num, 4, ["C", "L", "R", "R"])

    _auto_width(ws, 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_safety_data(
    edu_data: list[dict],
    check_data: list[dict],
    accident_data: list[dict],
    vendor: str = "",
) -> bytes:
    """3-4. 안전관리 Excel (본사/업체 공용)

    3개 시트: 안전교육, 차량점검, 사고이력
    """
    if not HAS_OPENPYXL:
        return b""

    wb = Workbook()

    # ── 시트1: 안전교육 ──
    ws1 = wb.active
    ws1.title = "안전교육"
    edu_headers = ["교육일", "교육유형", "참석자", "이수시간", "결과"]
    for c, h in enumerate(edu_headers, 1):
        ws1.cell(row=1, column=c, value=h)
    _apply_header_style(ws1, 1, len(edu_headers))
    for r_idx, r in enumerate(edu_data, 2):
        ws1.cell(row=r_idx, column=1, value=r.get("edu_date", ""))
        ws1.cell(row=r_idx, column=2, value=r.get("edu_type", ""))
        ws1.cell(row=r_idx, column=3, value=r.get("attendee", ""))
        ws1.cell(row=r_idx, column=4, value=r.get("hours", ""))
        ws1.cell(row=r_idx, column=5, value=r.get("result", ""))
        _apply_data_style(ws1, r_idx, len(edu_headers), ["C", "L", "L", "C", "C"])
    _auto_width(ws1, len(edu_headers))

    # ── 시트2: 차량점검 ──
    ws2 = wb.create_sheet("차량점검")
    chk_headers = ["점검일", "기사", "차량번호", "결과", "비고"]
    for c, h in enumerate(chk_headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _apply_header_style(ws2, 1, len(chk_headers))
    for r_idx, r in enumerate(check_data, 2):
        ws2.cell(row=r_idx, column=1, value=r.get("check_date", ""))
        ws2.cell(row=r_idx, column=2, value=r.get("driver", ""))
        ws2.cell(row=r_idx, column=3, value=r.get("vehicle_no", ""))
        ws2.cell(row=r_idx, column=4, value=r.get("result", ""))
        ws2.cell(row=r_idx, column=5, value=r.get("memo", ""))
        _apply_data_style(ws2, r_idx, len(chk_headers), ["C", "L", "L", "C", "L"])
    _auto_width(ws2, len(chk_headers))

    # ── 시트3: 사고이력 ──
    ws3 = wb.create_sheet("사고이력")
    acc_headers = ["발생일", "유형", "심각도", "기사", "장소", "내용"]
    for c, h in enumerate(acc_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    _apply_header_style(ws3, 1, len(acc_headers))
    for r_idx, r in enumerate(accident_data, 2):
        ws3.cell(row=r_idx, column=1, value=r.get("accident_date", ""))
        ws3.cell(row=r_idx, column=2, value=r.get("accident_type", ""))
        ws3.cell(row=r_idx, column=3, value=r.get("severity", ""))
        ws3.cell(row=r_idx, column=4, value=r.get("driver", ""))
        ws3.cell(row=r_idx, column=5, value=r.get("location", ""))
        ws3.cell(row=r_idx, column=6, value=r.get("description", ""))
        _apply_data_style(ws3, r_idx, len(acc_headers), ["C", "C", "C", "L", "L", "L"])
    _auto_width(ws3, len(acc_headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_school_collections(data: list[dict], school: str, year: str, month: str) -> bytes:
    """3-5. 수거내역 Excel (학교용)"""
    headers = ["수거일자", "품목", "수거량(kg)", "기사", "상태"]
    aligns = ["C", "C", "R", "L", "C"]
    rows = []
    total_weight = 0.0
    for r in data:
        w = float(r.get("weight", 0) or 0)
        total_weight += w
        rows.append([
            r.get("collect_date", ""),
            r.get("item_type", ""),
            round(w, 1),
            r.get("driver", ""),
            r.get("status", ""),
        ])
    totals = ["합계", "", round(total_weight, 1), f"{len(rows)}건", ""]
    return build_excel(
        title=f"{school} 수거내역",
        headers=headers,
        rows=rows,
        aligns=aligns,
        subtitle=f"{year}년 {month}월 | {school}",
        totals=totals,
    )


def export_customers(data: list[dict], vendor: str = "") -> bytes:
    """3-6. 거래처 목록 Excel (업체용)"""
    headers = ["거래처명", "유형", "대표자", "사업자번호", "연락처", "이메일",
               "음식물 단가", "재활용 단가", "일반 단가"]
    aligns = ["L", "C", "L", "C", "C", "L", "R", "R", "R"]
    rows = []
    for r in data:
        rows.append([
            r.get("name", ""),
            r.get("cust_type", ""),
            r.get("ceo", ""),
            r.get("biz_no", ""),
            r.get("phone", ""),
            r.get("email", ""),
            int(float(r.get("price_food", 0) or 0)),
            int(float(r.get("price_recycle", 0) or 0)),
            int(float(r.get("price_general", 0) or 0)),
        ])
    return build_excel(
        title="거래처 목록",
        headers=headers,
        rows=rows,
        aligns=aligns,
        subtitle=f"업체: {vendor}" if vendor else "",
    )


def export_meal_data(data: list[dict], site: str, year: str, month: str) -> bytes:
    """3-7. 식단 데이터 Excel (급식용)"""
    headers = ["날짜", "식단명", "메뉴", "칼로리(kcal)", "인원수", "알레르기", "잔반등급"]
    aligns = ["C", "L", "L", "R", "R", "L", "C"]
    rows = []
    for r in data:
        rows.append([
            r.get("meal_date", ""),
            r.get("meal_name", ""),
            r.get("menu", ""),
            int(float(r.get("calories", 0) or 0)),
            int(float(r.get("servings", 0) or 0)),
            r.get("allergens", ""),
            r.get("waste_grade", "-"),
        ])
    return build_excel(
        title="식단 데이터",
        headers=headers,
        rows=rows,
        aligns=aligns,
        subtitle=f"{year}년 {month}월 | {site}",
    )
