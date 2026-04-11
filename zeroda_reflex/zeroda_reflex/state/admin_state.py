# zeroda_reflex/state/admin_state.py
# 본사관리자(HQ Admin) 상태 관리
import reflex as rx
import logging
from datetime import datetime
from zeroda_reflex.state.auth_state import AuthState

logger = logging.getLogger(__name__)
from zeroda_reflex.utils.database import (
    db_get, db_insert, db_upsert,
    get_all_collections, get_all_vendors_list,
    get_all_users, update_user_approval, update_user_active,
    reset_user_password, update_user_fields, delete_user,
    create_user, validate_password,
    get_user_by_id, upsert_customer_neis_codes, is_school_in_customer_info,
    # 섹션B: 수거데이터
    get_pending_collections, confirm_all_pending, reject_collection_by_id,
    get_filtered_collections, get_all_schools_list,
    get_hq_processing_confirms, confirm_processing_item, reject_processing_item,
    get_processing_vendors, get_processing_drivers,
    # 섹션B 확장: 업로드 + 편집 (P1 복원)
    bulk_insert_collections, update_collection_row, delete_collection,
    # 섹션C: 외주업체관리
    get_all_vendor_info, save_hq_vendor_info,
    get_school_master_all, update_school_alias,
    hq_add_violation, hq_get_violations,
    hq_get_safety_scores, hq_calculate_safety_score,
    # 섹션D: 수거일정 + NEIS
    get_hq_schedules, hq_save_schedule, hq_delete_schedule,
    get_neis_schools_by_vendor, save_neis_meal_schedule,
    get_meal_schedules, approve_meal_schedules, cancel_meal_schedules,
    check_schedule_duplicate,
    # 섹션E: 정산+탄소
    get_settlement_data, get_hq_settlement_summary,
    get_customers_by_vendor, save_customer, delete_customer,
    get_carbon_data,
    # 섹션F: 안전관리+폐기물분석
    get_hq_safety_education, get_hq_accident_reports,
    get_hq_safety_checklist, get_hq_daily_checks, update_accident_status,
    get_waste_analytics,
    # 섹션G: 현장사진 (P3)
    get_photo_records_all,
    # 계정관리: 업체 드롭다운
    get_active_vendor_names,
    get_vendor_info,
)

# ── 안전관리 서브탭 (P1 복원) ──
SAFETY_SUB_TABS = ["안전교육", "점검결과", "사고보고", "일일점검", "기사모니터링"]

# ── 본사관리자 메뉴 목록 ──
HQ_TABS = [
    "대시보드", "수거데이터", "외주업체관리", "거래처관리", "수거일정",
    "정산관리", "안전관리", "탄소감축", "폐기물분석", "현장사진", "계정관리",
    "문서서비스",
]

# ── 문서서비스 서브탭 (본사관리자 — 2026-04-10 신규) ──
# 본사관리자는 "양식관리" 접근 가능 (업로드/삭제/태그미리보기)
DOC_SERVICE_HQ_TABS = ["문서발급", "양식관리", "발급내역"]

# 문서 카테고리 (hwpx 양식 업로드 시 분류)
DOC_CATEGORIES = ["2자계약서", "3자계약서", "견적서", "처리확인서", "기타"]

# ── 양식별 동적 입력 필드용 선택지 (2026-04-10 추가) ──
WASTE_TYPE_OPTIONS = ["음식물류폐기물", "일반폐기물", "재활용폐기물", "음식물+일반", "음식물+재활용"]
TREATMENT_OPTIONS = ["소각", "매립", "퇴비화", "사료화", "바이오가스", "기타"]
FREQUENCY_OPTIONS = ["매일", "주6회", "주5회", "주3회", "주2회", "주1회", "격주", "월1회"]

# ── 현장사진 유형 옵션 (P3) ──
PHOTO_TYPE_OPTIONS = ["전체", "before", "after", "issue", "etc"]
PHOTO_TYPE_MAP = {
    "before": "수거전",
    "after": "수거후",
    "issue": "이슈",
    "etc": "기타",
}

# ── 거래처 유형 옵션 (등록폼용 — '전체' 제외) ──
CUST_FORM_TYPE_OPTIONS = [
    "학교", "기업", "관공서", "일반업장",
    "기타", "기타1(면세사업장)", "기타2(부가세포함)",
]

# ── 폐기물분석 서브탭 ──
ANALYTICS_SUB_TABS = ["종합분석", "이상치탐지", "기상분석"]

# ── 정산관리: 거래처 유형 옵션 (P1 복원) ──
CUST_TYPE_OPTIONS = [
    "전체", "학교", "기업", "관공서", "일반업장",
    "기타", "기타1(면세사업장)", "기타2(부가세포함)",
]


def _get_tax_type(cust_type: str) -> str:
    """거래처 유형 → 세금 분류

    tax_free      : 면세 (학교, 기타1) — 공급가액만
    fixed_fee     : 고정비 (기타) — 월정액, VAT 없음
    fixed_fee_vat : 고정비 + VAT 10% (기타2)
    vat_10        : 부가세 10% (기업/관공서/일반업장)
    """
    if cust_type in ("학교", "기타1(면세사업장)"):
        return "tax_free"
    if cust_type == "기타":
        return "fixed_fee"
    if cust_type == "기타2(부가세포함)":
        return "fixed_fee_vat"
    return "vat_10"


def _correct_settlement_prices(rows: list[dict], vendor: str) -> list[dict]:
    """거래처별 등록 단가로 settle_rows 보정.

    customer_info.price_food/price_recycle/price_general 값이 0보다 크면
    해당 학교·품목에 한해 unit_price/amount 를 재계산한다.
    또한 각 row 에 cust_type, fixed_fee, tax_type 을 채워 넣는다.
    """
    if not rows or not vendor:
        for r in rows:
            r["cust_type"] = ""
            r["fixed_fee"] = "0"
            r["tax_type"] = "vat_10"
        return rows

    customers = get_customers_by_vendor(vendor) or []
    info_map: dict[str, dict] = {}
    for c in customers:
        name = c.get("name", "")
        if not name:
            continue
        try:
            pf = float(c.get("price_food", 0) or 0)
        except (ValueError, TypeError):
            pf = 0.0
        try:
            pr = float(c.get("price_recycle", 0) or 0)
        except (ValueError, TypeError):
            pr = 0.0
        try:
            pg = float(c.get("price_general", 0) or 0)
        except (ValueError, TypeError):
            pg = 0.0
        try:
            ff = float(c.get("fixed_fee", 0) or 0)
        except (ValueError, TypeError):
            ff = 0.0
        info_map[name] = {
            "cust_type": c.get("cust_type", "학교"),
            "fixed_fee": ff,
            "음식물": pf,
            "재활용": pr,
            "일반": pg,
        }

    for r in rows:
        school = r.get("school_name", "")
        item = r.get("item_type", "")
        info = info_map.get(school, {})
        ctype = info.get("cust_type", "")
        r["cust_type"] = ctype
        r["fixed_fee"] = str(int(info.get("fixed_fee", 0)))
        r["tax_type"] = _get_tax_type(ctype)

        # 단가 보정 (등록 단가가 있을 때만)
        item_key = "음식물"
        if "재활용" in item:
            item_key = "재활용"
        elif "일반" in item:
            item_key = "일반"
        unit = info.get(item_key, 0) or 0
        if unit > 0:
            try:
                w = float(r.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            r["unit_price"] = str(int(unit))
            r["amount"] = str(int(round(w * unit)))
    return rows


class AdminState(AuthState):
    """본사관리자 전체 상태"""

    # ══════════════════════════════
    #  공통
    # ══════════════════════════════
    active_tab: str = "대시보드"
    selected_year: str = str(datetime.now().year)
    selected_month: str = str(datetime.now().month)

    # ══════════════════════════════
    #  문서서비스 (2026-04-10 Phase 3 리팩토링)
    # ══════════════════════════════
    doc_sub_tab: str = "문서발급"
    # 양식관리
    doc_templates: list[dict] = []
    doc_upload_name: str = ""
    doc_upload_category: str = "2자계약서"
    doc_upload_msg: str = ""
    doc_upload_ok: bool = False
    # 문서발급 — 양식/거래처 선택
    doc_selected_template_id: int = 0
    doc_selected_customer_id: int = 0
    doc_selected_category: str = ""
    doc_customer_query: str = ""
    doc_customer_candidates: list[dict] = []          # 필터된 결과 (UI 표시용)
    doc_customer_all: list[dict] = []                 # 전체 거래처 (원본, 필터 소스)
    doc_customer_vendor_filter: str = ""              # 업체 필터 ("" = 전체)
    doc_customer_vendor_options: list[str] = []       # 드롭다운 선택지
    doc_selected_customer_info: dict = {}
    doc_vendor_info: dict = {}
    # 중간처리업체 정보 (Phase 5 — mid_processor 테이블)
    doc_processor_info: dict = {}
    # 신규거래처 직접입력 모드
    doc_new_customer_mode: bool = False
    doc_new_cust_name: str = ""
    doc_new_cust_bizno: str = ""
    doc_new_cust_rep: str = ""
    doc_new_cust_address: str = ""
    doc_new_cust_phone: str = ""
    doc_new_cust_type: str = ""
    # ── 통합 폼 데이터 (Phase 3: 개별 변수 50개 → dict 1개) ──
    doc_form_data: dict = {}
    # 미리보기 카드
    doc_preview_visible: bool = False
    doc_preview_data: dict = {}
    doc_preview_summary: str = ""
    # 발급 상태
    doc_issue_msg: str = ""
    doc_issue_ok: bool = False
    doc_last_issued_pdf: str = ""
    # 발급내역
    doc_issue_log: list[dict] = []
    doc_log_filter_from: str = ""
    doc_log_filter_to: str = ""

    # ══════════════════════════════
    #  대시보드
    # ══════════════════════════════
    dash_vendor_summary: list[dict] = []
    dash_top5_schools: list[dict] = []
    dash_total_weight: float = 0.0
    dash_total_count: int = 0
    dash_vendor_count: int = 0
    dash_school_count: int = 0
    # P3: 대시보드 탄소 KPI
    dash_carbon_reduced: str = "0"
    dash_tree_equivalent: str = "0"

    # ══════════════════════════════
    #  학사일정 동기화
    # ══════════════════════════════
    sched_sync_msg: str = ""
    sched_sync_running: bool = False

    # ══════════════════════════════
    #  계정관리
    # ══════════════════════════════
    all_users: list[dict] = []
    acct_filter_role: str = "전체"
    acct_filter_status: str = "전체"
    acct_msg: str = ""
    acct_ok: bool = False

    # 생성 다이얼로그
    acct_create_open: bool = False
    acct_new_id: str = ""
    acct_new_name: str = ""
    acct_new_pw: str = ""
    acct_new_role: str = "driver"
    acct_new_vendor: str = ""
    acct_new_schools: str = ""
    acct_new_edu: str = ""
    acct_vendor_options: list[str] = []
    # 수정 다이얼로그
    acct_edit_open: bool = False
    acct_edit_target: str = ""
    acct_edit_name: str = ""
    acct_edit_role: str = ""
    acct_edit_vendor: str = ""
    acct_edit_schools: str = ""
    acct_edit_edu: str = ""
    acct_edit_new_pw: str = ""
    acct_edit_neis_edu: str = ""
    acct_edit_neis_school: str = ""
    # 삭제 확인
    acct_delete_open: bool = False
    acct_delete_target: str = ""

    # ══════════════════════════════
    #  수거데이터 (섹션B)
    # ══════════════════════════════
    data_sub_tab: str = "전송대기"

    # 전송 대기
    pending_rows: list[dict] = []
    data_msg: str = ""
    data_ok: bool = False

    # 전체 수거 내역 / 시뮬레이션 필터
    data_vendor_filter: str = "전체"
    data_month_filter: str = "전체"
    data_school_filter: str = "전체"
    data_collection_rows: list[dict] = []
    data_vendor_options: list[str] = []
    data_school_options: list[str] = []

    # ── 데이터 업로드 (P1 복원) ──
    data_upload_progress: int = 0
    data_upload_success: int = 0
    data_upload_fail: int = 0
    data_upload_vendor: str = ""        # 업로드 시 vendor 강제 지정 (옵션)

    # ── 행 편집 (P1 복원) ──
    edit_row_id: str = ""
    edit_row_weight: str = "0"
    edit_row_unit_price: str = "0"
    edit_row_item_type: str = ""
    edit_row_memo: str = ""
    edit_row_open: bool = False

    # 처리확인
    proc_vendor_filter: str = "전체"
    proc_driver_filter: str = "전체"
    proc_status_filter: str = "전체"
    proc_rows: list[dict] = []
    proc_vendor_options: list[str] = []
    proc_driver_options: list[str] = []
    proc_total_count: int = 0
    proc_pending_count: int = 0
    proc_confirmed_count: int = 0
    proc_total_weight: float = 0.0
    # 오늘 비교 (P1 복원 — 섹션 5)
    proc_today_coll_weight: float = 0.0
    proc_today_proc_weight: float = 0.0
    proc_today_diff: float = 0.0
    proc_show_today_compare: bool = False

    # ══════════════════════════════
    #  외주업체관리 (섹션C)
    # ══════════════════════════════
    vendor_sub_tab: str = "외주업체목록"
    vendor_list: list[dict] = []
    vendor_msg: str = ""
    vendor_ok: bool = False

    # #7 담당거래처 드릴다운 — 외주업체목록에서 업체 클릭 시 해당 업체 거래처만 표시
    selected_vendor_for_clients: str = ""
    selected_vendor_clients: list[dict] = []

    # ── 직인 관리 (멀티테넌트) ──
    stamp_vendor_select: str = ""
    stamp_upload_status: str = ""
    stamp_upload_loading: bool = False
    stamp_current_path: str = ""

    # 업체 등록/수정 폼
    vf_vendor: str = ""
    vf_biz_name: str = ""
    vf_rep: str = ""
    vf_biz_no: str = ""
    vf_address: str = ""
    vf_contact: str = ""
    vf_email: str = ""
    vf_vehicle_no: str = ""
    vf_account: str = ""  # 계약서/견적서 입금계좌 (은행명+계좌번호 한 줄)
    # P2: 학교 배정 (콤마 구분 텍스트)
    vf_schools_text: str = ""

    # ══════════════════════════════
    #  P2: 거래처관리 (섹션 신규)
    # ══════════════════════════════
    cust_vendor_filter: str = ""          # 업체 선택
    cust_type_filter: str = "전체"        # 거래처 유형 필터
    cust_rows: list[dict] = []            # 거래처 목록
    cust_msg: str = ""
    cust_ok: bool = True

    # 거래처 편집 폼
    cf_name: str = ""                     # 거래처명
    cf_cust_type: str = "학교"            # 유형
    cf_price_food: str = "0"              # 음식물 단가
    cf_price_recycle: str = "0"           # 재활용 단가
    cf_price_general: str = "0"           # 일반 단가
    cf_address: str = ""                  # 주소
    cf_phone: str = ""                    # 연락처
    cf_biz_no: str = ""                   # 사업자번호
    cf_ceo: str = ""                      # 대표자명
    cf_fixed_fee: str = "0"               # 월정액
    cf_email: str = ""                    # 이메일

    # 학교 별칭
    school_master_list: list[dict] = []
    alias_school_sel: str = ""
    alias_input: str = ""

    # 안전관리 평가
    eval_vendor_sel: str = ""
    eval_year: str = str(datetime.now().year)
    eval_month: str = str(datetime.now().month)
    eval_result: dict = {}
    safety_scores_list: list[dict] = []
    violations_list: list[dict] = []
    viol_vendor_filter: str = "전체"

    # 위반 기록 입력
    viol_vendor: str = ""
    viol_driver: str = ""
    viol_date: str = ""
    viol_type: str = "과속"
    viol_location: str = ""
    viol_fine: str = "0"
    viol_memo: str = ""

    # ══════════════════════════════
    #  수거일정 (섹션D)
    # ══════════════════════════════
    sched_sub_tab: str = "일정조회"
    sched_vendor_filter: str = "전체"
    sched_month_filter: str = ""
    sched_rows: list[dict] = []
    sched_msg: str = ""
    sched_ok: bool = False

    # 일정 등록 폼
    sf_vendor: str = ""
    sf_month_key: str = ""
    sf_weekdays: str = ""
    sf_schools: str = ""
    sf_items: str = "음식물"
    sf_driver: str = ""

    # ── 오늘현황 (P1 복원) ──
    today_collection_rows: list[dict] = []
    today_total_weight: str = "0.0"
    today_total_count: int = 0
    today_vendor_filter: str = "전체"

    # ── 급식승인 워크플로우 (P1 복원) ──
    meal_approval_vendor: str = "전체"
    meal_approval_month: str = datetime.now().strftime("%Y-%m")
    meal_approval_status: str = "draft"          # draft/approved/cancelled
    meal_draft_rows: list[dict] = []
    meal_pending_count: int = 0
    meal_approved_count: int = 0
    meal_cancelled_count: int = 0
    meal_approval_driver: str = ""
    meal_approval_offset: str = "1"               # 급식일+1일 = 수거일
    meal_selected_ids: list[str] = []
    meal_cancel_note: str = ""

    # NEIS
    neis_vendor: str = ""
    neis_month: str = ""
    neis_school_list: list[dict] = []
    neis_school_sel: str = ""
    neis_result_msg: str = ""
    neis_meal_dates: list[str] = []
    neis_collect_offset: str = "0"
    neis_item_type: str = "음식물"
    neis_driver: str = ""

    # ══════════════════════════════
    #  정산관리 + 탄소감축 (섹션E)
    # ══════════════════════════════
    settle_year: str = str(datetime.now().year)
    settle_month: str = str(datetime.now().month)
    settle_vendor: str = "전체"
    settle_cust_type: str = "전체"          # 거래처 유형 필터 (P1)
    settle_tax_label: str = "공급가액"      # KPI 라벨 (면세/공급가액/월정액)
    settle_summary: dict = {}
    settle_rows: list[dict] = []
    settlement_list: list[dict] = []        # PDF/이메일용 (school 단위)

    # ── 미수금 관리 (P1 복원) ──
    overdue_amount: str = "0"
    overdue_months: str = ""
    overdue_memo: str = ""

    # ── [이메일 발송] (Phase 6) ──
    email_to: str = ""
    email_msg: str = ""
    email_ok: bool = False
    email_sending: bool = False

    # P3: 수신자(거래처) 정보 + 이메일 템플릿
    settle_rcv_rep: str = ""
    settle_rcv_biz_no: str = ""
    settle_rcv_phone: str = ""
    settle_rcv_address: str = ""
    settle_rcv_biz_type: str = ""
    settle_rcv_biz_item: str = ""
    settle_email_subject: str = ""
    settle_email_body: str = ""

    # ── SMS 발송 (Phase 8) ──
    sms_to: str = ""
    sms_msg: str = ""
    sms_ok: bool = False
    sms_sending: bool = False

    carbon_year: str = str(datetime.now().year)
    carbon_month: str = "전체"
    carbon_data: dict = {}
    carbon_school_ranking: list[dict] = []
    # P3: 품목별 기여도 + 월별 추이
    carbon_by_item: list[dict] = []
    carbon_monthly_trend: list[dict] = []

    # ══════════════════════════════
    #  안전관리 + 폐기물분석 (섹션F)
    # ══════════════════════════════
    safety_vendor_filter: str = "전체"
    safety_sub_tab: str = "안전교육"          # P1 복원: 5개 서브탭
    safety_edu_rows: list[dict] = []
    safety_accident_rows: list[dict] = []

    # 점검결과
    safety_checklist_rows: list[dict] = []
    safety_checklist_fail_count: int = 0  # P2: 불합격 점검 건수

    # 일일점검
    daily_check_rows: list[dict] = []
    daily_check_month: str = datetime.now().strftime("%Y-%m")
    daily_check_ok_count: int = 0
    daily_check_fail_count: int = 0
    daily_check_ok_rate: str = "0%"

    # 기사 모니터링
    driver_monitor_rows: list[dict] = []
    monitor_normal_count: int = 0
    monitor_caution_count: int = 0
    monitor_warning_count: int = 0
    monitor_emergency_count: int = 0

    # 사고 상태 변경
    accident_status_id: str = ""
    accident_new_status: str = "처리중"
    safety_msg: str = ""
    safety_ok: bool = False

    analytics_year: str = str(datetime.now().year)
    analytics_month: str = "전체"
    analytics_data: dict = {}
    analytics_by_item: list[dict] = []
    analytics_by_school: list[dict] = []
    analytics_by_vendor: list[dict] = []
    analytics_by_month: list[dict] = []
    # P3: 종합분석 확장
    analytics_mom_change: str = "0%"
    analytics_by_weekday: list[dict] = []
    analytics_by_season: list[dict] = []
    analytics_by_driver: list[dict] = []

    # ── P2: 폐기물분석 서브탭 ──
    analytics_sub_tab: str = "종합분석"

    # ── P2: 이상치 탐지 ──
    anomaly_rows: list[dict] = []
    anomaly_threshold: str = "2.0"
    anomaly_mean: str = "0"
    anomaly_std: str = "0"
    anomaly_count: int = 0
    anomaly_msg: str = ""

    # ── P2: 기상 상관분석 ──
    weather_start_date: str = ""
    weather_end_date: str = ""
    weather_corr_temp: str = "0.000"
    weather_corr_rain: str = "0.000"
    weather_corr_humidity: str = "0.000"
    weather_corr_wind: str = "0.000"
    weather_rainy_avg: str = "0"
    weather_clear_avg: str = "0"
    weather_diff_pct: str = "0"
    weather_temp_bins: list[dict] = []
    weather_msg: str = ""
    weather_ok: bool = True

    # ── P3: 현장사진 ──
    photo_vendor_filter: str = ""
    photo_type_filter: str = "전체"
    photo_date_from: str = ""
    photo_date_to: str = ""
    photo_rows: list[dict] = []
    photo_msg: str = ""

    # ══════════════════════════════
    #  Computed Vars — 대시보드
    # ══════════════════════════════

    @rx.var
    def has_vendor_summary(self) -> bool:
        return len(self.dash_vendor_summary) > 0

    @rx.var
    def has_top5(self) -> bool:
        return len(self.dash_top5_schools) > 0

    @rx.var
    def year_month_str(self) -> str:
        return f"{self.selected_year}-{self.selected_month.zfill(2)}"

    # ══════════════════════════════
    #  Computed Vars — 계정관리
    # ══════════════════════════════

    @rx.var
    def filtered_users(self) -> list[dict]:
        result = self.all_users
        if self.acct_filter_role != "전체":
            result = [u for u in result if u.get("role") == self.acct_filter_role]
        if self.acct_filter_status == "승인대기":
            result = [u for u in result if u.get("approval_status") == "pending"]
        elif self.acct_filter_status == "승인완료":
            result = [u for u in result if u.get("approval_status") == "approved"]
        elif self.acct_filter_status == "반려":
            result = [u for u in result if u.get("approval_status") == "rejected"]
        elif self.acct_filter_status == "비활성":
            result = [u for u in result if str(u.get("is_active", "1")) == "0"]
        return result

    @rx.var
    def has_users(self) -> bool:
        return len(self.filtered_users) > 0

    @rx.var
    def pending_count(self) -> int:
        return sum(1 for u in self.all_users if u.get("approval_status") == "pending")

    @rx.var
    def acct_has_msg(self) -> bool:
        return self.acct_msg != ""

    @rx.var
    def vendor_user_counts(self) -> list[dict]:
        """업체별 사용자 수 집계"""
        counts: dict[str, dict] = {}
        for u in self.all_users:
            v = u.get("vendor") or "(업체없음)"
            if v not in counts:
                counts[v] = {"vendor": v, "total": 0, "active": 0, "pending": 0}
            counts[v]["total"] += 1
            if u.get("approval_status") == "pending":
                counts[v]["pending"] += 1
            elif u.get("is_active") == 1:
                counts[v]["active"] += 1
        return sorted(counts.values(), key=lambda x: x["vendor"])

    # ══════════════════════════════
    #  Computed Vars — 수거데이터
    # ══════════════════════════════

    @rx.var
    def has_pending(self) -> bool:
        return len(self.pending_rows) > 0

    @rx.var
    def pending_row_count(self) -> int:
        return len(self.pending_rows)

    @rx.var
    def data_has_msg(self) -> bool:
        return self.data_msg != ""

    @rx.var
    def has_collection_rows(self) -> bool:
        return len(self.data_collection_rows) > 0

    @rx.var
    def has_proc_rows(self) -> bool:
        return len(self.proc_rows) > 0

    # ══════════════════════════════
    #  Computed Vars — 외주업체
    # ══════════════════════════════

    @rx.var
    def has_vendors(self) -> bool:
        return len(self.vendor_list) > 0

    @rx.var
    def vendor_has_msg(self) -> bool:
        return self.vendor_msg != ""

    @rx.var
    def has_school_master(self) -> bool:
        return len(self.school_master_list) > 0

    @rx.var
    def school_name_options(self) -> list[str]:
        return [s["school_name"] for s in self.school_master_list if s.get("school_name")]

    @rx.var
    def has_eval_result(self) -> bool:
        return bool(self.eval_result)

    @rx.var
    def has_safety_scores(self) -> bool:
        return len(self.safety_scores_list) > 0

    @rx.var
    def has_violations(self) -> bool:
        return len(self.violations_list) > 0

    @rx.var
    def vendor_name_options(self) -> list[str]:
        return [v["vendor"] for v in self.vendor_list if v.get("vendor")]

    @rx.var
    def vendor_name_options_all(self) -> list[str]:
        """'전체' 포함 업체 목록 (UI selectbox용)"""
        return ["전체"] + [v["vendor"] for v in self.vendor_list if v.get("vendor")]

    @rx.var
    def data_vendor_options_all(self) -> list[str]:
        return ["전체"] + list(self.data_vendor_options)

    @rx.var
    def data_school_options_all(self) -> list[str]:
        return ["전체"] + list(self.data_school_options)

    @rx.var
    def proc_vendor_options_all(self) -> list[str]:
        return ["전체"] + list(self.proc_vendor_options)

    @rx.var
    def proc_driver_options_all(self) -> list[str]:
        return ["전체"] + list(self.proc_driver_options)

    # ══════════════════════════════
    #  Computed Vars — 수거일정
    # ══════════════════════════════

    @rx.var
    def has_sched_rows(self) -> bool:
        return len(self.sched_rows) > 0

    @rx.var
    def sched_has_msg(self) -> bool:
        return self.sched_msg != ""

    @rx.var
    def has_neis_schools(self) -> bool:
        return len(self.neis_school_list) > 0

    @rx.var
    def neis_school_name_options(self) -> list[str]:
        return [s["name"] for s in self.neis_school_list if s.get("name")]

    @rx.var
    def has_neis_meal_dates(self) -> bool:
        return len(self.neis_meal_dates) > 0

    @rx.var
    def neis_meal_count(self) -> int:
        return len(self.neis_meal_dates)

    # ══════════════════════════════
    #  Computed Vars — 정산/탄소
    # ══════════════════════════════

    @rx.var
    def has_settle_rows(self) -> bool:
        return len(self.settle_rows) > 0

    @rx.var
    def has_settle_summary(self) -> bool:
        return bool(self.settle_summary)

    @rx.var
    def has_carbon_data(self) -> bool:
        return bool(self.carbon_data)

    @rx.var
    def has_carbon_ranking(self) -> bool:
        return len(self.carbon_school_ranking) > 0

    # ══════════════════════════════
    #  Computed Vars — 안전/분석
    # ══════════════════════════════

    @rx.var
    def has_edu_rows(self) -> bool:
        return len(self.safety_edu_rows) > 0

    @rx.var
    def has_accident_rows(self) -> bool:
        return len(self.safety_accident_rows) > 0

    @rx.var
    def has_safety_checklist(self) -> bool:
        return len(self.safety_checklist_rows) > 0

    @rx.var
    def has_daily_check_rows(self) -> bool:
        return len(self.daily_check_rows) > 0

    @rx.var
    def has_driver_monitor(self) -> bool:
        return len(self.driver_monitor_rows) > 0

    @rx.var
    def safety_has_msg(self) -> bool:
        return self.safety_msg != ""

    @rx.var
    def has_cust_rows(self) -> bool:
        return len(self.cust_rows) > 0

    @rx.var
    def has_cust_msg(self) -> bool:
        return self.cust_msg != ""

    @rx.var
    def has_anomaly_rows(self) -> bool:
        return len(self.anomaly_rows) > 0

    @rx.var
    def has_weather_msg(self) -> bool:
        return self.weather_msg != ""

    @rx.var
    def has_weather_temp_bins(self) -> bool:
        return len(self.weather_temp_bins) > 0

    @rx.var
    def has_analytics(self) -> bool:
        return bool(self.analytics_data)

    @rx.var
    def has_by_item(self) -> bool:
        return len(self.analytics_by_item) > 0

    @rx.var
    def has_by_school(self) -> bool:
        return len(self.analytics_by_school) > 0

    @rx.var
    def doc_vendor_filter_options(self) -> list[str]:
        """업체 필터 드롭다운 선택지 (전체 포함)."""
        return ["전체"] + self.doc_customer_vendor_options

    # ══════════════════════════════
    #  이벤트 — 공통
    # ══════════════════════════════

    def on_admin_load(self):
        """페이지 로드 시"""
        if not self.is_authenticated or self.user_role != "admin":
            return rx.redirect("/")
        self.load_dashboard()

    def set_active_tab(self, tab: str):
        self.active_tab = tab
        if tab == "대시보드":
            self.load_dashboard()
        elif tab == "수거데이터":
            self.load_data_tab()
        elif tab == "외주업체관리":
            self.load_vendor_tab()
        elif tab == "수거일정":
            self.load_schedule_tab()
        elif tab == "정산관리":
            self.load_settlement()
        elif tab == "탄소감축":
            self.load_carbon()
        elif tab == "안전관리":
            self.load_safety()
        elif tab == "폐기물분석":
            self.load_analytics()
        elif tab == "거래처관리":
            self.load_customers()
        elif tab == "현장사진":
            self.load_photos()
        elif tab == "계정관리":
            self.acct_msg = ""
            self.acct_ok = False
            self.load_users()
        elif tab == "문서서비스":
            self.load_doc_service()

    # ══════════════════════════════
    #  이벤트 — 문서서비스 (2026-04-10 신규)
    # ══════════════════════════════

    def load_doc_service(self):
        """문서서비스 진입 시: 활성 양식 목록 + 발급이력 로드."""
        self.doc_upload_msg = ""
        self.doc_issue_msg = ""
        try:
            from ..utils.database import get_db
            conn = get_db()
            try:
                # PG + SQLite 모두 INTEGER 컬럼이므로 = 1 통일
                _active_cond = "is_active = 1"
                # 활성 양식 목록
                tpl_rows = conn.execute(
                    "SELECT id, template_name, category, file_path, tag_list, "
                    "created_by, created_at FROM document_templates "
                    f"WHERE {_active_cond} ORDER BY id DESC"
                ).fetchall()
                self.doc_templates = []
                for r in tpl_rows:
                    d = dict(r)
                    self.doc_templates.append({
                        "id": d.get("id"),
                        "template_name": d.get("template_name") or "",
                        "category": d.get("category") or "",
                        "file_path": d.get("file_path") or "",
                        "tag_list": d.get("tag_list") or "",
                        "created_by": d.get("created_by") or "",
                        "created_at": d.get("created_at") or "",
                    })
                # 최근 발급이력 (최근 100건)
                log_rows = conn.execute(
                    "SELECT id, vendor, template_name, customer_name, issued_by, "
                    "issued_at, file_path, issue_number FROM document_issue_log "
                    "ORDER BY id DESC LIMIT 100"
                ).fetchall()
                self.doc_issue_log = []
                for r in log_rows:
                    d = dict(r)
                    self.doc_issue_log.append({
                        "id": d.get("id"),
                        "vendor": d.get("vendor") or "",
                        "template_name": d.get("template_name") or "",
                        "customer_name": d.get("customer_name") or "",
                        "issued_by": d.get("issued_by") or "",
                        "issued_at": d.get("issued_at") or "",
                        "file_path": d.get("file_path") or "",
                        "issue_number": d.get("issue_number") or "",
                    })
                # ── 전체 거래처 로드 (U1: 거래처 필터용) ──
                cust_rows = conn.execute(
                    "SELECT name, biz_no, rep, addr, phone, "
                    "price_food, price_recycle, price_general, vendor "
                    "FROM customer_info ORDER BY name"
                ).fetchall()
                all_custs = []
                vendors_set = set()
                idx = 1
                for r in cust_rows:
                    d = dict(r)
                    v = str(d.get("vendor", "") or "")
                    all_custs.append({
                        "id": idx,
                        "customer_name": str(d.get("name", "") or ""),
                        "business_no": str(d.get("biz_no", "") or ""),
                        "representative": str(d.get("rep", "") or ""),
                        "address": str(d.get("addr", "") or ""),
                        "phone": str(d.get("phone", "") or ""),
                        "price_food": d.get("price_food") or 0,
                        "price_recycle": d.get("price_recycle") or 0,
                        "price_general": d.get("price_general") or 0,
                        "vendor": v,
                    })
                    if v:
                        vendors_set.add(v)
                    idx += 1
                self.doc_customer_all = all_custs
                self.doc_customer_candidates = all_custs
                self.doc_customer_vendor_options = sorted(vendors_set)
                self.doc_customer_vendor_filter = ""
                self.doc_customer_query = ""
            finally:
                conn.close()
        except Exception as e:
            logger.error(f"[load_doc_service] {e}", exc_info=True)
            self.doc_upload_msg = "문서서비스 로드에 실패했습니다."
            self.doc_upload_ok = False

    def set_doc_sub_tab(self, tab: str):
        """문서서비스 서브탭 전환."""
        self.doc_sub_tab = tab
        # 서브탭 바뀔 때마다 최신화
        self.load_doc_service()

    def set_doc_upload_category(self, cat: str):
        self.doc_upload_category = cat

    def set_doc_upload_name(self, name: str):
        self.doc_upload_name = name

    def set_doc_selected_template(self, tpl_id: int):
        self.doc_selected_template_id = int(tpl_id) if tpl_id else 0
        # 선택된 양식의 카테고리 자동 세팅
        for t in self.doc_templates:
            if t.get("id") == self.doc_selected_template_id:
                self.doc_selected_category = t.get("category", "")
                break
        # 양식 변경 시 폼 데이터 초기화 (Phase 3: dict 통합)
        # Phase 4: 카테고리별 필드 키를 기본값으로 미리 채움
        #   → UI에서 doc_form_data[key].to(str) 접근 시 undefined 방지
        from ..utils.form_field_config import FORM_FIELDS as _FF
        cat = self.doc_selected_category
        init_data: dict = {}
        if cat in _FF:
            for sec in _FF[cat].get("sections", []):
                for fld in sec.get("fields", []):
                    init_data[fld["key"]] = fld.get("default", "")
        self.doc_form_data = init_data
        self.doc_preview_visible = False
        self.doc_preview_data = {}
        # 중간처리업체 정보 자동 로드
        self._load_processor_info()

    def set_doc_selected_customer(self, cust_id):
        try:
            self.doc_selected_customer_id = int(cust_id) if cust_id else 0
        except (ValueError, TypeError):
            self.doc_selected_customer_id = 0

        # ── 선택된 거래처 상세정보 세팅 (후보 목록에서 찾기) ──
        self.doc_selected_customer_info = {}
        for c in self.doc_customer_candidates:
            if c.get("id") == self.doc_selected_customer_id:
                self.doc_selected_customer_info = {
                    "customer_name": str(c.get("customer_name") or ""),
                    "business_no": str(c.get("business_no") or ""),
                    "representative": str(c.get("representative") or ""),
                    "address": str(c.get("address") or ""),
                    "phone": str(c.get("phone") or ""),
                    "price_food": str(c.get("price_food") or "0"),
                    "price_recycle": str(c.get("price_recycle") or "0"),
                    "price_general": str(c.get("price_general") or "0"),
                }
                break

        # ── 외주업체(수집운반업체) 정보 로드 — vendor_info 테이블 ──
        self.doc_vendor_info = {}
        try:
            from ..utils.database import get_vendor_info
            vendor_name = ""
            # 로그인 사용자의 vendor 확인
            try:
                from ..utils.database import get_db
                conn = get_db()
                row = conn.execute(
                    "SELECT vendor FROM users WHERE user_id = ? LIMIT 1",
                    (self.user_id or "",),
                ).fetchone()
                if row:
                    rd = dict(row)
                    vendor_name = rd.get("vendor") or ""
                conn.close()
            except Exception:
                pass
            if not vendor_name:
                vendor_name = "하영자원"
            vi = get_vendor_info(vendor_name)
            self.doc_vendor_info = {
                "biz_name": str(vi.get("biz_name") or vendor_name),
                "rep": str(vi.get("rep") or ""),
                "biz_no": str(vi.get("biz_no") or ""),
                "address": str(vi.get("address") or ""),
                "contact": str(vi.get("contact") or ""),
                "account": str(vi.get("account") or ""),
                "license_no": str(vi.get("license_no") or ""),
            }
        except Exception as e:
            logger.warning("doc_vendor_info 로드 실패: %s", e)
            self.doc_vendor_info = {
                "biz_name": "하영자원",
                "rep": "", "biz_no": "", "address": "",
                "contact": "", "account": "", "license_no": "",
            }
        # 중간처리업체 로드 (Phase 5)
        self._load_processor_info()
        self.doc_preview_visible = False

    def clear_doc_customer_selection(self):
        """거래처 선택 초기화 (U2: '변경' 버튼용)."""
        self.doc_selected_customer_id = 0
        self.doc_selected_customer_info = {}
        self.doc_new_cust_name = ""
        self.doc_new_cust_bizno = ""
        self.doc_new_cust_rep = ""
        self.doc_new_cust_address = ""
        self.doc_new_cust_phone = ""
        self.doc_new_cust_type = ""
        self.doc_new_customer_mode = False
        self.doc_preview_visible = False
        self.doc_preview_data = {}

    def set_doc_customer_query(self, q: str):
        self.doc_customer_query = q

    def noop(self):
        """아무 동작 안 함 (토글 버튼 비활성 상태용)."""
        pass

    # ── 신규거래처 모드 전환 + setter (2026-04-10 추가) ──
    def toggle_new_customer_mode(self):
        """기존거래처 검색 ↔ 신규거래처 직접입력 모드 전환."""
        self.doc_new_customer_mode = not self.doc_new_customer_mode
        if self.doc_new_customer_mode:
            # 신규 모드로 전환 → 기존 선택 초기화
            self.doc_selected_customer_id = 0
            self.doc_selected_customer_info = {}
            self.search_doc_customers()    # 현재 필터 상태로 재적용
        else:
            # 기존 모드로 복귀 → 신규 입력 초기화
            self.doc_new_cust_name = ""
            self.doc_new_cust_bizno = ""
            self.doc_new_cust_rep = ""
            self.doc_new_cust_address = ""
            self.doc_new_cust_phone = ""
            self.doc_new_cust_type = ""

    def set_doc_new_cust_name(self, v: str):
        self.doc_new_cust_name = v

    def set_doc_new_cust_bizno(self, v: str):
        self.doc_new_cust_bizno = v

    def set_doc_new_cust_rep(self, v: str):
        self.doc_new_cust_rep = v

    def set_doc_new_cust_address(self, v: str):
        self.doc_new_cust_address = v

    def set_doc_new_cust_phone(self, v: str):
        self.doc_new_cust_phone = v

    def set_doc_new_cust_type(self, v: str):
        self.doc_new_cust_type = v

    # ── Phase 3: 통합 폼 필드 setter ──
    def set_doc_form_field(self, key_value: str):
        """범용 setter — 'key::value' 형식으로 전달받아 doc_form_data에 저장.

        Reflex on_change는 단일 str만 전달 가능하므로
        UI에서 lambda v: AdminState.set_doc_form_field(key + '::' + v) 형태로 호출.
        """
        if "::" in key_value:
            key, val = key_value.split("::", 1)
            new_data = dict(self.doc_form_data)
            new_data[key] = val
            self.doc_form_data = new_data

    def _load_processor_info(self):
        """중간처리업체(재활용업체) 정보를 mid_processor 테이블에서 로드."""
        self.doc_processor_info = {}
        try:
            from ..utils.database import get_db
            conn = get_db()
            try:
                row = conn.execute(
                    "SELECT name, biz_no, rep, address, phone, "
                    "license_no, biz_type, corp_no, fax "
                    "FROM mid_processor WHERE is_default = 1 LIMIT 1"
                ).fetchone()
                if row:
                    d = dict(row)
                    self.doc_processor_info = {
                        "name": str(d.get("name") or ""),
                        "biz_no": str(d.get("biz_no") or ""),
                        "rep": str(d.get("rep") or ""),
                        "address": str(d.get("address") or ""),
                        "phone": str(d.get("phone") or ""),
                        "license_no": str(d.get("license_no") or ""),
                        "biz_type": str(d.get("biz_type") or ""),
                        "corp_no": str(d.get("corp_no") or ""),
                        "fax": str(d.get("fax") or ""),
                    }
            finally:
                conn.close()
        except Exception as e:
            logger.warning("_load_processor_info 실패: %s", e)

    def build_fill_data(self) -> dict:
        """doc_form_data + DB 자동채움 데이터를 합쳐서 hwpx 채움용 dict 생성."""
        from ..utils.form_field_config import (
            CUSTOMER_TO_EMITTER, VENDOR_TO_TRANSPORTER, PROCESSOR_TO_FIELDS,
        )
        result = dict(self.doc_form_data)

        # 배출자 ← customer_info
        ci = self.doc_selected_customer_info
        if ci:
            for fill_key, db_col in CUSTOMER_TO_EMITTER.items():
                if fill_key not in result or not result[fill_key]:
                    result[fill_key] = str(ci.get(db_col) or ci.get("customer_name", "") if db_col == "name" else ci.get(db_col, ""))

        # 수집운반자 ← vendor_info
        vi = self.doc_vendor_info
        if vi:
            for fill_key, db_col in VENDOR_TO_TRANSPORTER.items():
                if fill_key not in result or not result[fill_key]:
                    if db_col == "음식물류폐기물":
                        result[fill_key] = "음식물류폐기물"
                    else:
                        result[fill_key] = str(vi.get(db_col, ""))

        # 중간처리업체 ← mid_processor
        pi = self.doc_processor_info
        if pi:
            for fill_key, db_col in PROCESSOR_TO_FIELDS.items():
                if fill_key not in result or not result[fill_key]:
                    result[fill_key] = str(pi.get(db_col, ""))

        # 대표자 뒤에 (인) 붙이기
        for rep_key in ("emitter_rep", "transporter_rep", "processor_rep"):
            v = result.get(rep_key, "")
            if v and "(인)" not in v:
                result[rep_key] = v + "(인)"

        return result

    async def issue_hwpx_document(self):
        """양식 선택 + 폼 데이터 → hwpx 셀 채움 → PDF 변환 → DB 기록."""
        self.doc_preview_visible = False
        import os as _os
        from datetime import datetime as _dt
        from ..utils.form_field_config import CATEGORY_CELL_MAP
        from ..utils.hwpx_engine import fill_by_cell_map, convert_to_pdf
        from ..utils.database import get_db

        if not self.doc_selected_template_id:
            self.doc_issue_msg = "양식을 먼저 선택하세요."
            self.doc_issue_ok = False
            return
        if not self.doc_selected_customer_id and not self.doc_new_customer_mode:
            self.doc_issue_msg = "거래처를 선택하세요."
            self.doc_issue_ok = False
            return

        # 양식 file_path 조회
        tpl_row = None
        for t in self.doc_templates:
            if t.get("id") == self.doc_selected_template_id:
                tpl_row = t
                break
        if not tpl_row or not tpl_row.get("file_path"):
            self.doc_issue_msg = "양식 파일 경로를 찾을 수 없습니다."
            self.doc_issue_ok = False
            return

        hwpx_src = tpl_row["file_path"]
        category = tpl_row.get("category", "")
        cell_map = CATEGORY_CELL_MAP.get(category, {})

        if not cell_map:
            self.doc_issue_msg = f"카테고리 '{category}'에 대한 셀 매핑이 없습니다."
            self.doc_issue_ok = False
            return

        # 자동채움 데이터 빌드
        fill_data = self.build_fill_data()

        # 출력 경로
        base_dir = _os.environ.get("ZERODA_UPLOAD_DIR", "/opt/zeroda-platform/uploads")
        out_dir = _os.path.join(base_dir, "issued")
        _os.makedirs(out_dir, exist_ok=True)

        stamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        cust_name = fill_data.get("emitter_name", "unknown")
        safe_cust = cust_name.replace("/", "_").replace(" ", "_")[:20]
        issue_number = f"{category}_{stamp}_{safe_cust}"

        hwpx_out = _os.path.join(out_dir, f"{issue_number}.hwpx")
        pdf_out = _os.path.join(out_dir, f"{issue_number}.pdf")

        try:
            # hwpx 셀 채움
            fill_by_cell_map(hwpx_src, cell_map, fill_data, hwpx_out)

            # PDF 변환
            pdf_ok = convert_to_pdf(hwpx_out, pdf_out)

            # 서버 절대경로 → 웹 URL 변환 (nginx /uploads/ alias 기준)
            _base_prefix = _os.environ.get(
                "ZERODA_UPLOAD_DIR", "/opt/zeroda-platform/uploads"
            )
            def _to_url(path: str) -> str:
                if path.startswith(_base_prefix):
                    return "/uploads/" + path[len(_base_prefix):].lstrip("/")
                return path.replace("/opt/zeroda-platform", "")

            # DB 기록
            conn = get_db()
            try:
                conn.execute(
                    "INSERT INTO document_issue_log "
                    "(vendor, template_id, template_name, customer_id, customer_name, "
                    "issued_by, issued_at, file_path, issue_number) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        self.user_vendor or "하영자원",
                        self.doc_selected_template_id,
                        tpl_row.get("template_name", ""),
                        self.doc_selected_customer_id,
                        cust_name,
                        self.user_id or "admin",
                        _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                        _to_url(pdf_out) if pdf_ok else _to_url(hwpx_out),
                        issue_number,
                    ),
                )
                conn.commit()
            finally:
                conn.close()

            if pdf_ok:
                self.doc_issue_msg = ""
                self.doc_last_issued_pdf = _to_url(pdf_out)
                yield rx.toast.info(f"발급 완료: {issue_number}")
            else:
                self.doc_issue_msg = ""
                self.doc_last_issued_pdf = _to_url(hwpx_out)
                yield rx.toast.warning(f"hwpx 저장 완료 (PDF 변환 실패): {issue_number}")
            self.doc_issue_ok = True
            self.load_doc_service()

        except Exception as e:
            logger.error(f"[issue_document] {e}", exc_info=True)
            self.doc_issue_msg = "문서 발급에 실패했습니다."
            self.doc_issue_ok = False

    def search_doc_customers(self):
        """거래처 필터링 (로컬 — doc_customer_all에서 텍스트+업체 필터)."""
        q = (self.doc_customer_query or "").strip().lower()
        vf = self.doc_customer_vendor_filter or ""

        result = []
        for c in self.doc_customer_all:
            if vf and c.get("vendor", "") != vf:
                continue
            if q:
                name_match = q in (c.get("customer_name", "") or "").lower()
                biz_match = q in (c.get("business_no", "") or "").lower()
                if not name_match and not biz_match:
                    continue
            result.append(c)

        self.doc_customer_candidates = result

    def set_doc_customer_vendor_filter(self, vendor: str):
        """업체 필터 드롭다운 변경 시 호출."""
        self.doc_customer_vendor_filter = vendor
        self.search_doc_customers()

    def set_doc_customer_query_and_filter(self, query: str):
        """거래처 검색어 변경 시 즉시 필터링."""
        self.doc_customer_query = query
        self.search_doc_customers()

    async def upload_document_template(self, files: list[rx.UploadFile]):
        """hwpx 양식 업로드 → 서버 저장 → 태그 추출 → DB 기록."""
        import json
        import os as _os
        from datetime import datetime as _dt
        from ..utils.database import get_db

        if not files:
            self.doc_upload_msg = "파일이 선택되지 않았습니다."
            self.doc_upload_ok = False
            return
        if not self.doc_upload_name.strip():
            self.doc_upload_msg = "양식 이름을 입력하세요."
            self.doc_upload_ok = False
            return

        f = files[0]
        filename = getattr(f, "name", None) or getattr(f, "filename", "template.hwpx")
        if not filename.lower().endswith(".hwpx"):
            self.doc_upload_msg = "hwpx 파일만 업로드 가능합니다."
            self.doc_upload_ok = False
            return

        # 저장 경로 — /opt/zeroda-platform/uploads/templates/
        base_dir = _os.environ.get(
            "ZERODA_UPLOAD_DIR", "/opt/zeroda-platform/uploads"
        )
        tpl_dir = _os.path.join(base_dir, "templates")
        _os.makedirs(tpl_dir, exist_ok=True)

        stamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        safe_name = f"{stamp}_{filename}"
        save_path = _os.path.join(tpl_dir, safe_name)

        try:
            data = await f.read()
            with open(save_path, "wb") as out:
                out.write(data)
        except Exception as e:
            logger.error(f"[upload_template_file] 파일 저장: {e}", exc_info=True)
            self.doc_upload_msg = "파일 저장에 실패했습니다."
            self.doc_upload_ok = False
            return

        # 태그 추출
        try:
            from ..utils.hwpx_engine import extract_tags
            tags = extract_tags(save_path)
        except Exception as e:
            tags = []
            logger.error(f"[upload_template_file] 태그 추출: {e}", exc_info=True)
            self.doc_upload_msg = "태그 추출에 실패했습니다(파일 저장은 완료)."

        # DB insert
        try:
            conn = get_db()
            dup_row = conn.execute(
                "SELECT id FROM document_templates "
                "WHERE template_name = ? AND category = ? AND is_active = 1 "
                "LIMIT 1",
                (self.doc_upload_name.strip(), self.doc_upload_category),
            ).fetchone()
            if dup_row:
                conn.close()
                self.doc_upload_msg = f"이미 같은 이름의 양식이 있습니다: '{self.doc_upload_name.strip()}' ({self.doc_upload_category})"
                self.doc_upload_ok = False
                return
            conn.execute(
                "INSERT INTO document_templates "
                "(template_name, category, file_path, file_type, tag_list, "
                " created_by, created_at, is_active) "
                "VALUES (?, ?, ?, 'hwpx', ?, ?, ?, 1)",
                (
                    self.doc_upload_name.strip(),
                    self.doc_upload_category,
                    save_path,
                    json.dumps(tags, ensure_ascii=False),
                    self.user_id or "admin",
                    _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"[upload_template_file] DB 등록: {e}", exc_info=True)
            self.doc_upload_msg = "DB 등록에 실패했습니다."
            self.doc_upload_ok = False
            return

        self.doc_upload_msg = ""
        self.doc_upload_ok = True
        self.doc_upload_name = ""
        self.load_doc_service()
        yield rx.toast.info(f"업로드 완료. 추출된 태그 {len(tags)}개")

    def deactivate_doc_template(self, tpl_id: int):
        """양식 비활성화 (삭제 아닌 숨김)."""
        try:
            from ..utils.database import get_db
            conn = get_db()
            conn.execute(
                "UPDATE document_templates SET is_active = 0 WHERE id = ?",
                (int(tpl_id),),
            )
            conn.commit()
            conn.close()
            self.load_doc_service()
            self.doc_upload_msg = ""
            self.doc_upload_ok = True
            yield rx.toast.info("양식을 비활성화했습니다.")
        except Exception as e:
            logger.error(f"[deactivate_template] {e}", exc_info=True)
            self.doc_upload_msg = "비활성화에 실패했습니다."
            self.doc_upload_ok = False

    def _build_issue_data(self) -> tuple[dict, dict, dict, str]:
        """발급에 필요한 template row, customer row, extra dict, template file_path 수집."""
        from ..utils.database import get_db
        conn = get_db()
        # 양식
        tpl_row = conn.execute(
            "SELECT id, template_name, category, file_path FROM document_templates "
            "WHERE id = ?",
            (self.doc_selected_template_id,),
        ).fetchone()
        template_row = {}
        if tpl_row:
            td = dict(tpl_row)
            template_row = {
                "id": td.get("id"),
                "template_name": td.get("template_name") or "",
                "category": td.get("category") or "",
                "file_path": td.get("file_path") or "",
            }
        # 거래처 — 신규직접입력 모드 또는 기존검색 모드
        customer_row = {}
        if self.doc_new_customer_mode and self.doc_new_cust_name.strip():
            # 신규거래처 직접입력 모드
            customer_row = {
                "id": 0,
                "customer_name": self.doc_new_cust_name.strip(),
                "business_no": self.doc_new_cust_bizno.strip(),
                "representative": self.doc_new_cust_rep.strip(),
                "address": self.doc_new_cust_address.strip(),
                "phone": self.doc_new_cust_phone.strip(),
                "cust_type": self.doc_new_cust_type.strip(),
                "price_food": 0,
                "price_recycle": 0,
                "price_general": 0,
            }
        elif self.doc_selected_customer_id:
            # 후보 목록에서 이미 매핑된 데이터 사용 (DB 컬럼명 불일치 방지)
            for cc in self.doc_customer_candidates:
                if cc.get("id") == self.doc_selected_customer_id:
                    customer_row = {
                        "id": cc.get("id"),
                        "customer_name": str(cc.get("customer_name") or ""),
                        "business_no": str(cc.get("business_no") or ""),
                        "representative": str(cc.get("representative") or ""),
                        "address": str(cc.get("address") or ""),
                        "phone": str(cc.get("phone") or ""),
                        "price_food": cc.get("price_food") or 0,
                        "price_recycle": cc.get("price_recycle") or 0,
                        "price_general": cc.get("price_general") or 0,
                    }
                    break
        conn.close()
        # Phase 4: doc_form_data 통합 dict에서 extra 구성
        fd = self.doc_form_data or {}
        extra = dict(fd)  # 전체 폼 데이터 복사
        return template_row, customer_row, extra, (template_row.get("file_path") or "")

    def _get_vendor_info(self) -> dict:
        """로그인한 본사(또는 외주업체) 회사정보 조회 — 수집운반업체 태그용."""
        try:
            from ..utils.database import get_db
            conn = get_db()
            row = conn.execute(
                "SELECT vendor FROM users WHERE user_id = ? LIMIT 1",
                (self.user_id or "",),
            ).fetchone()
            vendor_name = ""
            if row:
                rd = dict(row)
                vendor_name = rd.get("vendor") or ""
            if not vendor_name:
                vendor_name = "하영자원"
            conn.close()
            return {
                "company_name": vendor_name or "하영자원",
                "business_no": "",
                "representative": "정석완",
                "address": "경기 화성시 남양읍 남양성지로 219, 2층",
                "phone": "010-3114-4030",
                "license_no": "제20-35호",
            }
        except Exception:
            return {
                "company_name": "하영자원",
                "representative": "정석완",
                "address": "경기 화성시 남양읍 남양성지로 219, 2층",
                "phone": "010-3114-4030",
                "license_no": "제20-35호",
            }

    def preview_document(self):
        """미리보기 — build_fill_data로 채움 데이터를 구성하고 미리보기 카드 표시."""
        self.doc_preview_visible = False
        self.doc_preview_data = {}
        self.doc_preview_summary = ""

        if not self.doc_selected_template_id:
            self.doc_issue_msg = "양식을 먼저 선택하세요."
            self.doc_issue_ok = False
            return
        has_customer = (
            (self.doc_new_customer_mode and self.doc_new_cust_name.strip())
            or (not self.doc_new_customer_mode and self.doc_selected_customer_id)
        )
        if not has_customer:
            self.doc_issue_msg = "거래처를 선택하거나 신규거래처 정보를 입력하세요."
            self.doc_issue_ok = False
            return
        try:
            data = self.build_fill_data()
            filled = sum(1 for v in data.values() if v)
            total = len(data)
            cat = self.doc_selected_category or "양식"
            cust_name = data.get("emitter_name", "")

            preview = {}
            preview["emitter_name"] = str(data.get("emitter_name", "") or "")
            preview["emitter_bizno"] = str(data.get("emitter_bizno", "") or "")
            preview["emitter_address"] = str(data.get("emitter_address", "") or "")
            preview["emitter_phone"] = str(data.get("emitter_phone", "") or "")
            preview["emitter_rep"] = str(data.get("emitter_rep", "") or "")
            preview["transporter_name"] = str(data.get("transporter_name", "") or "")
            preview["transporter_license"] = str(data.get("transporter_license", "") or "")
            preview["transporter_phone"] = str(data.get("transporter_phone", "") or "")
            preview["transporter_rep"] = str(data.get("transporter_rep", "") or "")
            preview["waste_type_1"] = str(data.get("waste_type_1", "") or "")
            preview["quantity_1"] = str(data.get("quantity_1", "") or "")
            preview["method_1"] = str(data.get("method_1", "") or "")
            preview["total_amount"] = str(data.get("total_amount", "") or "")
            preview["category"] = cat

            self.doc_preview_data = preview
            self.doc_preview_visible = True
            self.doc_preview_summary = (
                cat + " / " + cust_name + " / 필드 "
                + str(filled) + "/" + str(total) + "개"
            )
            self.doc_issue_msg = ""
            self.doc_issue_ok = True
            yield rx.toast.info(
                "미리보기 생성 완료 — 아래 카드에서 확인하세요"
            )
        except Exception as e:
            logger.error(f"[preview_document] {e}", exc_info=True)
            self.doc_issue_msg = "미리보기에 실패했습니다."
            self.doc_issue_ok = False

    def issue_document(self):
        """실제 PDF 발급: fill_template → insert_stamp → convert_to_pdf → log insert."""
        import os as _os
        from datetime import datetime as _dt
        from ..utils.database import get_db
        from ..utils.hwpx_engine import fill_template, insert_stamp, convert_to_pdf
        from ..utils.document_mapper import build_template_data, generate_issue_number

        if not self.doc_selected_template_id:
            self.doc_issue_msg = "양식을 먼저 선택하세요."
            self.doc_issue_ok = False
            return
        # 거래처 체크: 신규모드면 상호명 필수, 기존모드면 ID 필수
        has_customer = (
            (self.doc_new_customer_mode and self.doc_new_cust_name.strip())
            or (not self.doc_new_customer_mode and self.doc_selected_customer_id)
        )
        if not has_customer:
            self.doc_issue_msg = "거래처를 선택하거나 신규거래처 정보를 입력하세요."
            self.doc_issue_ok = False
            return

        try:
            tpl, cust, extra, tpl_path = self._build_issue_data()
            if not tpl_path or not _os.path.exists(tpl_path):
                self.doc_issue_msg = "양식 파일이 서버에 없습니다."
                self.doc_issue_ok = False
                return
            vendor = self._get_vendor_info()

            # 발급번호 채번 (당일 순번 단순 COUNT+1)
            conn = get_db()
            today = _dt.now().strftime("%Y-%m-%d")
            cnt_row = conn.execute(
                "SELECT COUNT(*) AS cnt FROM document_issue_log WHERE issued_at LIKE ?",
                (f"{today}%",),
            ).fetchone()
            seq = 1
            if cnt_row:
                try:
                    rd = dict(cnt_row)
                    seq = int(rd.get("cnt") or 0) + 1
                except Exception:
                    # sqlite3.Row 인덱스 fallback
                    try:
                        seq = int(cnt_row[0] or 0) + 1
                    except Exception:
                        seq = 1
            issue_no = generate_issue_number(tpl.get("category", ""), seq)

            data = build_template_data(
                category=tpl.get("category", ""),
                customer_row=cust,
                vendor_row=vendor,
                issuer=self.user_id or "admin",
                extra=extra,
                issue_number=issue_no,
            )

            # 출력 경로
            base_dir = _os.environ.get(
                "ZERODA_UPLOAD_DIR", "/opt/zeroda-platform/uploads"
            )
            out_dir = _os.path.join(base_dir, "issued")
            _os.makedirs(out_dir, exist_ok=True)
            stamp_str = _dt.now().strftime("%Y%m%d_%H%M%S")
            filled_hwpx = _os.path.join(out_dir, f"{issue_no}_filled.hwpx")
            stamped_hwpx = _os.path.join(out_dir, f"{issue_no}_stamped.hwpx")
            output_pdf = _os.path.join(out_dir, f"{issue_no}.pdf")

            # 1) 태그 치환
            fill_template(tpl_path, data, filled_hwpx)

            # 2) 직인 삽입 — 발급자의 직인 이미지 경로 조회
            stamp_path = self._get_stamp_image_path()
            insert_stamp(filled_hwpx, stamp_path or "", stamped_hwpx)

            # 3) PDF 변환
            pdf_ok = convert_to_pdf(stamped_hwpx, output_pdf)

            # 4) 로그 insert
            conn.execute(
                "INSERT INTO document_issue_log "
                "(vendor, template_id, template_name, customer_id, customer_name, "
                " issued_by, issued_at, file_path, issue_number, output_format) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    vendor.get("company_name") or "",
                    tpl.get("id"),
                    tpl.get("template_name") or "",
                    cust.get("id"),
                    cust.get("customer_name") or "",
                    self.user_id or "admin",
                    _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                    output_pdf if pdf_ok else stamped_hwpx,
                    issue_no,
                    "pdf" if pdf_ok else "hwpx",
                ),
            )
            conn.commit()
            conn.close()

            # 서버 절대경로 → 웹 URL로 변환 (nginx /uploads/ alias 기준)
            _base_prefix = _os.environ.get(
                "ZERODA_UPLOAD_DIR", "/opt/zeroda-platform/uploads"
            )
            def _to_url(path: str) -> str:
                if path.startswith(_base_prefix):
                    return "/uploads/" + path[len(_base_prefix):].lstrip("/")
                return path.replace("/opt/zeroda-platform", "")

            if pdf_ok:
                self.doc_last_issued_pdf = _to_url(output_pdf)
                self.doc_issue_msg = f"발급 완료: {issue_no}"
                self.doc_issue_ok = True
            else:
                self.doc_last_issued_pdf = _to_url(stamped_hwpx)
                self.doc_issue_msg = (
                    f"PDF 변환 실패 — hwpx 파일로 저장됨: {issue_no}. "
                    "다운로드 버튼으로 hwpx 파일을 직접 받으실 수 있습니다."
                )
                self.doc_issue_ok = False

            self.load_doc_service()
        except Exception as e:
            logger.error(f"[issue_document] {e}", exc_info=True)
            self.doc_issue_msg = "문서 발급에 실패했습니다."
            self.doc_issue_ok = False

    def _get_stamp_image_path(self) -> str:
        """발급자의 외주업체 등록 직인 이미지 경로. 없으면 빈 문자열."""
        try:
            from ..utils.database import get_db
            conn = get_db()
            # 기존 외주업체관리자 쪽 직인 컬럼은 환경별로 다를 수 있어
            # 여러 후보 컬럼을 순회 조회 (없으면 skip)
            candidates = [
                ("company_info", "stamp_image_path"),
                ("company_info", "stamp_path"),
                ("vendor_info", "stamp_image"),
                ("users", "stamp_path"),
            ]
            for table, col in candidates:
                try:
                    r = conn.execute(
                        f"SELECT {col} AS v FROM {table} LIMIT 1"
                    ).fetchone()
                    if r:
                        rd = dict(r)
                        val = rd.get("v") or rd.get(col)
                        if val:
                            conn.close()
                            return str(val)
                except Exception:
                    continue
            conn.close()
        except Exception:
            pass
        return ""

    def set_selected_year(self, y: str):
        self.selected_year = y
        self.load_dashboard()

    def set_selected_month(self, m: str):
        self.selected_month = m
        self.load_dashboard()

    # ══════════════════════════════
    #  이벤트 — 대시보드
    # ══════════════════════════════

    def load_dashboard(self):
        """대시보드 데이터 로드"""
        ym = self.year_month_str
        rows = get_all_collections(year_month=ym)
        self.dash_total_count = len(rows)
        total_w = 0.0
        vendor_map: dict[str, float] = {}
        school_map: dict[str, float] = {}
        for r in rows:
            try:
                w = float(r.get("weight", 0))
            except (ValueError, TypeError):
                w = 0.0
            total_w += w
            v = r.get("vendor", "")
            s = r.get("school_name", "")
            if v:
                vendor_map[v] = vendor_map.get(v, 0.0) + w
            if s:
                school_map[s] = school_map.get(s, 0.0) + w

        self.dash_total_weight = round(total_w, 1)
        self.dash_vendor_count = len(vendor_map)
        self.dash_school_count = len(school_map)

        # 업체별 요약
        self.dash_vendor_summary = [
            {"vendor": k, "total_weight": round(v, 1)}
            for k, v in sorted(vendor_map.items(), key=lambda x: -x[1])
        ]

        # 학교별 TOP5
        sorted_schools = sorted(school_map.items(), key=lambda x: -x[1])[:5]
        self.dash_top5_schools = [
            {"school_name": k, "total_weight": round(v, 1)}
            for k, v in sorted_schools
        ]

        # P3: 탄소감축 KPI
        try:
            y_int = int(self.selected_year)
            m_int = int(self.selected_month)
            cdata = get_carbon_data(y_int, m_int) or {}
            self.dash_carbon_reduced = str(cdata.get("carbon_reduced", "0") or "0")
            self.dash_tree_equivalent = str(cdata.get("tree_equivalent", "0") or "0")
        except Exception:
            self.dash_carbon_reduced = "0"
            self.dash_tree_equivalent = "0"

    def refresh_dashboard(self):
        self.load_dashboard()

    # ══════════════════════════════
    #  이벤트 — 계정관리
    # ══════════════════════════════

    def load_users(self):
        users = get_all_users() or []
        for u in users:
            # pending_biz_no / pending_rep 기본값 보장 (컬럼 없는 구버전 DB 호환)
            if "pending_biz_no" not in u:
                u["pending_biz_no"] = ""
            if "pending_rep" not in u:
                u["pending_rep"] = ""
            if u["pending_biz_no"] is None:
                u["pending_biz_no"] = ""
            if u["pending_rep"] is None:
                u["pending_rep"] = ""
            u["_school_in_db"] = True  # 기본값 (school/meal_manager 외 역할)
            if u.get("role") in ("school", "meal_manager"):
                v = u.get("pending_vendor") or ""
                s = u.get("pending_school_name") or ""
                if v and s:
                    u["_school_in_db"] = is_school_in_customer_info(v, s)
                else:
                    u["_school_in_db"] = False
        self.all_users = users
        # acct_msg는 여기서 초기화하지 않음 — approve/reject/toggle에서 설정 후 유지

    def set_acct_filter_role(self, v: str):
        self.acct_filter_role = v

    def set_acct_filter_status(self, v: str):
        self.acct_filter_status = v

    def approve_user(self, user_id: str):
        """사용자 승인 — school/meal_manager이면 NEIS 코드를 customer_info에 반영"""
        # P2: 거래처 미등록 학교 사전 차단
        target = get_user_by_id(user_id)
        if target and target.get("role") in ("school", "meal_manager"):
            v = target.get("pending_vendor") or ""
            s = target.get("pending_school_name") or ""
            if v and s and not is_school_in_customer_info(v, s):
                self.acct_msg = f"거래처 미등록: '{s}'({v})을 먼저 거래처 관리에 등록하세요."
                self.acct_ok = False
                return
        # vendor_admin 승인 시 vendor_info 자동 생성
        if target and target.get("role") == "vendor_admin":
            vendor_name = target.get("vendor") or ""
            if vendor_name:
                existing = get_vendor_info(vendor_name)
                if not existing.get("biz_no"):
                    auto_data = {
                        "vendor": vendor_name,
                        "biz_name": target.get("pending_biz_no") and vendor_name or vendor_name,
                        "rep": target.get("pending_rep") or "",
                        "biz_no": target.get("pending_biz_no") or "",
                        "address": target.get("pending_address") or "",
                        "contact": target.get("pending_contact") or "",
                    }
                    save_hq_vendor_info(auto_data)
        ok = update_user_approval(user_id, "approved")
        self.all_users = get_all_users()
        if not ok:
            self.acct_msg = "승인 처리 실패"
            self.acct_ok = False
            return
        # school/meal_manager: 가입 시 저장된 NEIS 임시 코드 → customer_info 자동 반영
        user = get_user_by_id(user_id)
        neis_warning = ""
        if user and user.get("role") in ("school", "meal_manager"):
            pv = user.get("pending_vendor") or ""
            ps = user.get("pending_school_name") or ""
            ne = user.get("neis_edu_pending") or ""
            ns = user.get("neis_school_pending") or ""
            if pv and ps and ne and ns:
                neis_ok = upsert_customer_neis_codes(pv, ps, ne, ns)
                if not neis_ok:
                    neis_warning = (
                        f" (주의: 거래처 '{ps}'({pv})를 찾지 못해 NEIS 코드 미반영 — "
                        "거래처 등록 후 수동 입력 필요)"
                    )
        self.acct_msg = f"{user_id} 승인 완료{neis_warning}"
        self.acct_ok = True if not neis_warning else False

    def reject_user(self, user_id: str):
        """사용자 반려"""
        ok = update_user_approval(user_id, "rejected")
        self.all_users = get_all_users()
        if ok:
            self.acct_msg = f"{user_id} 반려 완료"
            self.acct_ok = True
        else:
            self.acct_msg = "반려 처리 실패"
            self.acct_ok = False

    def toggle_user_active(self, user_id: str):
        """사용자 활성/비활성 전환"""
        user = next((u for u in self.all_users if u.get("user_id") == user_id), None)
        if not user:
            return
        current = 1 if user.get("is_active") == 1 else 0
        new_val = 0 if current == 1 else 1
        ok = update_user_active(user_id, new_val)
        self.all_users = get_all_users()
        if ok:
            self.acct_msg = f"{user_id} {'활성화' if new_val else '비활성화'} 완료"
            self.acct_ok = True
        else:
            self.acct_msg = "처리 실패"
            self.acct_ok = False

    def reset_password(self, user_id: str):
        """비밀번호 초기화 — 임시비번 생성 (정책: 대소문자+숫자+특수문자 각 1자 이상)"""
        import secrets
        import string
        # 정책 준수: 대문자·소문자·숫자·특수문자 각 1자 보장 + 나머지 6자 랜덤
        specials = "!@#$%^&*"
        pool = string.ascii_letters + string.digits + specials
        required = [
            secrets.choice(string.ascii_uppercase),
            secrets.choice(string.ascii_lowercase),
            secrets.choice(string.digits),
            secrets.choice(specials),
        ]
        rest = [secrets.choice(pool) for _ in range(6)]
        chars = required + rest
        secrets.SystemRandom().shuffle(chars)
        temp_pw = ''.join(chars)
        ok = reset_user_password(user_id, temp_pw)
        if ok:
            # 관리자에게 임시 비밀번호를 표시하여 사용자에게 전달할 수 있게 함
            self.acct_msg = f"{user_id} 임시비밀번호: {temp_pw} (사용자에게 전달하세요)"
            self.acct_ok = True
        else:
            self.acct_msg = "초기화 실패"
            self.acct_ok = False

    # ── 계정 생성 ──
    def open_create_dialog(self):
        self.acct_new_id = ""
        self.acct_new_name = ""
        self.acct_new_pw = ""
        self.acct_new_role = "driver"
        self.acct_new_vendor = ""
        self.acct_new_schools = ""
        self.acct_new_edu = ""
        self.acct_msg = ""
        self.acct_vendor_options = get_active_vendor_names()
        self.acct_create_open = True

    def close_create_dialog(self):
        self.acct_create_open = False

    def set_acct_new_id(self, v: str): self.acct_new_id = v
    def set_acct_new_name(self, v: str): self.acct_new_name = v
    def set_acct_new_pw(self, v: str): self.acct_new_pw = v
    def set_acct_new_role(self, v: str):
        self.acct_new_role = v
        self.acct_new_vendor = ""
        self.acct_new_schools = ""
        self.acct_new_edu = ""
    def set_acct_new_vendor(self, v: str): self.acct_new_vendor = v
    def set_acct_new_schools(self, v: str): self.acct_new_schools = v
    def set_acct_new_edu(self, v: str): self.acct_new_edu = v

    def submit_create_user(self):
        if not self.acct_new_id or not self.acct_new_name or not self.acct_new_pw:
            self.acct_msg = "아이디, 이름, 비밀번호는 필수입니다."
            self.acct_ok = False
            return
        ok, msg = validate_password(self.acct_new_pw)
        if not ok:
            self.acct_msg = msg
            self.acct_ok = False
            return
        created, msg = create_user(
            user_id=self.acct_new_id.strip(),
            password=self.acct_new_pw,
            role=self.acct_new_role,
            name=self.acct_new_name.strip(),
            vendor=self.acct_new_vendor.strip(),
            schools=self.acct_new_schools.strip(),
            edu_office=self.acct_new_edu.strip(),
            approval_status="approved",
            is_active=1,
        )
        if not created:
            self.acct_msg = msg
            self.acct_ok = False
            return
        self.acct_msg = msg
        self.acct_ok = True
        self.acct_create_open = False
        self.load_users()

    # ── 계정 수정 ──
    def open_edit_dialog(self, user_id: str):
        user = next((u for u in self.all_users if u.get("user_id") == user_id), None)
        if not user:
            return
        self.acct_edit_target = user_id
        self.acct_edit_name = user.get("name", "")
        self.acct_edit_role = user.get("role", "")
        self.acct_edit_vendor = user.get("vendor", "")
        self.acct_edit_schools = user.get("schools", "")
        self.acct_edit_edu = user.get("edu_office", "")
        self.acct_edit_new_pw = ""
        self.acct_edit_neis_edu = ""
        self.acct_edit_neis_school = ""
        # school/meal_manager: 기존 NEIS 코드 로드
        role = user.get("role", "")
        if role in ("school", "meal_manager"):
            school_name = user.get("schools", "") or user.get("pending_school_name", "")
            if school_name:
                rows = db_get("customer_info", {"school_name": school_name})
                if rows:
                    self.acct_edit_neis_edu = rows[0].get("neis_edu_code", "") or ""
                    self.acct_edit_neis_school = rows[0].get("neis_school_code", "") or ""
        self.acct_msg = ""
        self.acct_edit_open = True

    def close_edit_dialog(self):
        self.acct_edit_open = False

    def set_acct_edit_name(self, v: str): self.acct_edit_name = v
    def set_acct_edit_role(self, v: str): self.acct_edit_role = v
    def set_acct_edit_vendor(self, v: str): self.acct_edit_vendor = v
    def set_acct_edit_schools(self, v: str): self.acct_edit_schools = v
    def set_acct_edit_edu(self, v: str): self.acct_edit_edu = v
    def set_acct_edit_new_pw(self, v: str): self.acct_edit_new_pw = v
    def set_acct_edit_neis_edu(self, v: str): self.acct_edit_neis_edu = v
    def set_acct_edit_neis_school(self, v: str): self.acct_edit_neis_school = v

    def submit_edit_user(self):
        if self.acct_edit_new_pw:
            ok, msg = validate_password(self.acct_edit_new_pw)
            if not ok:
                self.acct_msg = msg
                self.acct_ok = False
                return
        ok, msg = update_user_fields(
            user_id=self.acct_edit_target,
            name=self.acct_edit_name or None,
            role=self.acct_edit_role or None,
            vendor=self.acct_edit_vendor,
            schools=self.acct_edit_schools,
            edu_office=self.acct_edit_edu,
            new_password=self.acct_edit_new_pw if self.acct_edit_new_pw else None,
        )
        self.acct_msg = msg
        self.acct_ok = ok
        if ok:
            # school/meal_manager: NEIS 코드 저장
            if self.acct_edit_role in ("school", "meal_manager"):
                school_name = self.acct_edit_schools.strip()
                if school_name and (self.acct_edit_neis_edu.strip() or self.acct_edit_neis_school.strip()):
                    upsert_customer_neis_codes(
                        school_name=school_name,
                        neis_edu_code=self.acct_edit_neis_edu.strip() or None,
                        neis_school_code=self.acct_edit_neis_school.strip() or None,
                    )
            self.acct_edit_open = False
            self.load_users()

    # ── 계정 삭제 ──
    def open_delete_dialog(self, user_id: str):
        self.acct_delete_target = user_id
        self.acct_msg = ""
        self.acct_delete_open = True

    def close_delete_dialog(self):
        self.acct_delete_open = False

    def confirm_delete_user(self):
        ok, msg = delete_user(self.acct_delete_target)
        self.acct_msg = msg
        self.acct_ok = ok
        self.acct_delete_open = False
        self.load_users()

    # ══════════════════════════════
    #  이벤트 — 학사일정 동기화
    # ══════════════════════════════

    @rx.event(background=True)
    async def sync_all_school_schedules(self):
        """전체 학교 학사일정 NEIS 동기화 (백그라운드)."""
        async with self:
            if self.sched_sync_running:
                return
            self.sched_sync_running = True
            self.sched_sync_msg = "학사일정 동기화 중..."
        try:
            from zeroda_reflex.utils.neis_sync_service import sync_all_schools
            stats = sync_all_schools()
            async with self:
                self.sched_sync_msg = (
                    f"완료: {stats['success']}/{stats['total_schools']}교, "
                    f"일정 {stats['events']}건"
                )
        except Exception as e:
            logger.error(f"[admin] sync_all_school_schedules 실패: {e}", exc_info=True)
            async with self:
                self.sched_sync_msg = "일정 동기화에 실패했습니다."
        finally:
            async with self:
                self.sched_sync_running = False

    # ══════════════════════════════
    #  이벤트 — 수거데이터 (섹션B)
    # ══════════════════════════════

    def load_data_tab(self):
        """수거데이터 탭 진입 시 초기 로드"""
        self.data_msg = ""
        self.data_vendor_options = get_all_vendors_list()
        self.data_school_options = get_all_schools_list()
        self.proc_vendor_options = get_processing_vendors()
        self.proc_driver_options = get_processing_drivers()
        if self.data_sub_tab == "전송대기":
            self.load_pending()
        elif self.data_sub_tab in ("전체수거", "시뮬레이션"):
            self.load_collection_data()
        elif self.data_sub_tab == "처리확인":
            self.load_processing()

    def set_data_upload_progress_zero(self):
        self.data_upload_progress = 0
        self.data_upload_success = 0
        self.data_upload_fail = 0

    def set_data_sub_tab(self, tab: str):
        self.data_sub_tab = tab
        self.data_msg = ""
        if tab == "전송대기":
            self.load_pending()
        elif tab in ("전체수거", "시뮬레이션"):
            self.load_collection_data()
        elif tab == "처리확인":
            self.load_processing()
        elif tab == "데이터업로드":
            self.set_data_upload_progress_zero()

    # ── 전송 대기 ──

    def load_pending(self):
        self.pending_rows = get_pending_collections()

    def confirm_all_pending_data(self):
        """미확인 전송 데이터 전체 확인"""
        cnt = confirm_all_pending()
        if cnt > 0:
            self.data_msg = f"{cnt}건 확인 완료"
            self.data_ok = True
        else:
            self.data_msg = "처리할 데이터가 없습니다"
            self.data_ok = False
        self.load_pending()

    def reject_single_collection(self, row_id: str):
        """특정 수거 데이터 반려"""
        v = self.data_vendor_filter if self.data_vendor_filter != "전체" else ""
        ok = reject_collection_by_id(int(row_id), vendor=v)
        if ok:
            self.data_msg = f"ID {row_id} 반려 완료"
            self.data_ok = True
        else:
            self.data_msg = "반려 처리 실패"
            self.data_ok = False
        self.load_pending()

    # ── 전체 수거 내역 / 시뮬레이션 ──

    def set_data_vendor_filter(self, v: str):
        self.data_vendor_filter = v
        self.load_collection_data()

    def set_data_month_filter(self, m: str):
        self.data_month_filter = m
        self.load_collection_data()

    def set_data_school_filter(self, s: str):
        self.data_school_filter = s
        self.load_collection_data()

    def load_collection_data(self):
        """필터 적용하여 수거 데이터 로드"""
        tbl = "sim_collection" if self.data_sub_tab == "시뮬레이션" else "real_collection"
        vendor = self.data_vendor_filter if self.data_vendor_filter != "전체" else ""
        ym = self.data_month_filter if self.data_month_filter != "전체" else ""
        school = self.data_school_filter if self.data_school_filter != "전체" else ""
        self.data_collection_rows = get_filtered_collections(
            table=tbl, vendor=vendor, year_month=ym, school=school,
        )

    # ── 데이터 업로드 (P1 복원) ──

    def set_data_upload_vendor(self, v: str):
        self.data_upload_vendor = v

    async def handle_data_upload(self, files: list[rx.UploadFile]):
        """CSV/Excel 업로드 → real_collection 일괄 등록.

        필수 컬럼: 업체(또는 vendor), 학교명(또는 school_name),
                   수거일(또는 collect_date), 중량(또는 weight)
        선택 컬럼: 품목(item_type), 단가(unit_price), 기사(driver), 메모(memo)
        """
        if not files:
            self.data_msg = "파일을 선택하세요."
            self.data_ok = False
            return
        f = files[0]
        fname = (f.filename or "").lower()
        try:
            raw = await f.read()
        except Exception as e:
            logger.warning(f"파일 읽기 실패: {e}")
            self.data_msg = "파일 읽기 실패"
            self.data_ok = False
            return

        rows: list[dict] = []
        # ── CSV ──
        if fname.endswith(".csv"):
            import csv as _csv, io as _io
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("euc-kr", errors="replace")
            reader = _csv.DictReader(_io.StringIO(text))
            for r in reader:
                rows.append(dict(r))
        # ── Excel ──
        elif fname.endswith((".xlsx", ".xls")):
            try:
                import openpyxl as _xl, io as _io
                wb = _xl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    d = {}
                    for i, h in enumerate(headers):
                        val = row[i] if i < len(row) else ""
                        d[h] = str(val) if val is not None else ""
                    rows.append(d)
                wb.close()
            except Exception as e:
                logger.warning(f"Excel 파싱 실패: {e}")
                self.data_msg = "Excel 파일 형식 오류"
                self.data_ok = False
                return
        else:
            self.data_msg = "CSV 또는 Excel(.xlsx) 파일만 지원됩니다."
            self.data_ok = False
            return

        if not rows:
            self.data_msg = "파일에 데이터가 없습니다."
            self.data_ok = False
            return

        # ── 컬럼 매핑 (한글/영문) ──
        col_map = {
            "vendor": "vendor", "업체": "vendor", "거래처업체": "vendor",
            "school_name": "school_name", "학교": "school_name", "학교명": "school_name", "거래처": "school_name",
            "collect_date": "collect_date", "수거일": "collect_date", "날짜": "collect_date",
            "item_type": "item_type", "품목": "item_type", "구분": "item_type",
            "weight": "weight", "중량": "weight", "음식물(kg)": "weight", "kg": "weight",
            "unit_price": "unit_price", "단가": "unit_price", "단가(원)": "unit_price",
            "driver": "driver", "기사": "driver",
            "memo": "memo", "메모": "memo", "비고": "memo",
        }

        def _normalize(row: dict) -> dict:
            out = {}
            for k, v in row.items():
                key = str(k or "").strip().lower()
                tgt = col_map.get(key) or col_map.get(str(k or "").strip())
                if tgt:
                    out[tgt] = v
            return out

        normalized: list[dict] = []
        total = len(rows)
        for idx, r in enumerate(rows):
            n = _normalize(r)
            # vendor 강제 지정 (옵션)
            if self.data_upload_vendor and not n.get("vendor"):
                n["vendor"] = self.data_upload_vendor
            normalized.append(n)
            self.data_upload_progress = int((idx + 1) / max(total, 1) * 50)
            yield

        # ── 일괄 INSERT ──
        succ, fail = bulk_insert_collections(normalized, uploader="hq_admin")
        self.data_upload_success = succ
        self.data_upload_fail = fail
        self.data_upload_progress = 100
        self.data_msg = f"업로드 완료 — 성공 {succ}건 / 실패·중복 {fail}건"
        self.data_ok = succ > 0
        self.load_collection_data()

    # ── 행 편집 (P1 복원) ──

    def open_edit_row(self, row_id: str, weight: str, unit_price: str,
                       item_type: str = "", memo: str = ""):
        self.edit_row_id = str(row_id)
        self.edit_row_weight = str(weight or "0")
        self.edit_row_unit_price = str(unit_price or "0")
        self.edit_row_item_type = str(item_type or "")
        self.edit_row_memo = str(memo or "")
        self.edit_row_open = True

    def close_edit_row(self):
        self.edit_row_open = False
        self.edit_row_id = ""

    def set_edit_row_weight(self, v: str):
        self.edit_row_weight = v

    def set_edit_row_unit_price(self, v: str):
        self.edit_row_unit_price = v

    def set_edit_row_item_type(self, v: str):
        self.edit_row_item_type = v

    def set_edit_row_memo(self, v: str):
        self.edit_row_memo = v

    def save_edit_row(self):
        if not self.edit_row_id:
            self.data_msg = "수정할 행이 선택되지 않았습니다."
            self.data_ok = False
            return
        try:
            w = float(self.edit_row_weight or 0)
            up = float(self.edit_row_unit_price or 0)
        except (ValueError, TypeError):
            self.data_msg = "중량/단가는 숫자여야 합니다."
            self.data_ok = False
            return
        v = self.data_vendor_filter if self.data_vendor_filter != "전체" else ""
        ok = update_collection_row(
            row_id=int(self.edit_row_id),
            weight=w, unit_price=up,
            item_type=self.edit_row_item_type,
            memo=self.edit_row_memo,
            vendor=v,
        )
        if ok:
            self.data_msg = f"ID {self.edit_row_id} 수정 완료"
            self.data_ok = True
            self.edit_row_open = False
            self.load_collection_data()
        else:
            self.data_msg = "수정 실패"
            self.data_ok = False

    def delete_collection_row(self, row_id: str):
        v = self.data_vendor_filter if self.data_vendor_filter != "전체" else ""
        ok = delete_collection(int(row_id), vendor=v)
        if ok:
            self.data_msg = f"ID {row_id} 삭제 완료"
            self.data_ok = True
            self.load_collection_data()
        else:
            self.data_msg = "삭제 실패"
            self.data_ok = False

    # ── 처리확인 ──

    def set_proc_vendor_filter(self, v: str):
        self.proc_vendor_filter = v
        self.load_processing()

    def set_proc_driver_filter(self, d: str):
        self.proc_driver_filter = d
        self.load_processing()

    def set_proc_status_filter(self, s: str):
        self.proc_status_filter = s
        self.load_processing()

    def load_processing(self):
        """처리확인 데이터 로드"""
        vendor = self.proc_vendor_filter if self.proc_vendor_filter != "전체" else ""
        driver = self.proc_driver_filter if self.proc_driver_filter != "전체" else ""
        status = self.proc_status_filter if self.proc_status_filter != "전체" else ""
        rows = get_hq_processing_confirms(vendor=vendor, driver=driver, status=status)
        self.proc_rows = rows
        self.proc_total_count = len(rows)
        self.proc_pending_count = sum(1 for r in rows if r.get("status") == "submitted")
        self.proc_confirmed_count = sum(1 for r in rows if r.get("status") == "confirmed")
        try:
            self.proc_total_weight = round(
                sum(float(r.get("total_weight", 0)) for r in rows), 1
            )
        except (ValueError, TypeError):
            self.proc_total_weight = 0.0

        # ── 오늘 수거량 vs 처리량 비교 (P1 복원 — 섹션 5) ──
        if vendor:
            today = datetime.now().strftime("%Y-%m-%d")
            ym = datetime.now().strftime("%Y-%m")
            try:
                colls = get_filtered_collections(
                    table="real_collection",
                    vendor=vendor,
                    year_month=ym,
                )
            except Exception:
                colls = []
            try:
                today_coll = sum(
                    float(c.get("weight", 0) or 0)
                    for c in colls
                    if str(c.get("collect_date", "")) == today
                )
            except (ValueError, TypeError):
                today_coll = 0.0
            try:
                today_proc = sum(
                    float(r.get("total_weight", 0) or 0)
                    for r in rows
                    if str(r.get("confirm_date", "")) == today
                )
            except (ValueError, TypeError):
                today_proc = 0.0
            self.proc_today_coll_weight = round(today_coll, 1)
            self.proc_today_proc_weight = round(today_proc, 1)
            self.proc_today_diff = round(today_proc - today_coll, 1)
            self.proc_show_today_compare = True
        else:
            self.proc_today_coll_weight = 0.0
            self.proc_today_proc_weight = 0.0
            self.proc_today_diff = 0.0
            self.proc_show_today_compare = False

    def confirm_proc_item(self, row_id: str):
        """처리확인 건 확인"""
        v = self.proc_vendor_filter if self.proc_vendor_filter != "전체" else ""
        ok = confirm_processing_item(int(row_id), vendor=v)
        if ok:
            self.data_msg = f"처리확인 ID {row_id} 확인 완료"
            self.data_ok = True
        else:
            self.data_msg = "확인 처리 실패"
            self.data_ok = False
        self.load_processing()

    def reject_proc_item(self, row_id: str):
        """처리확인 건 반려"""
        v = self.proc_vendor_filter if self.proc_vendor_filter != "전체" else ""
        ok = reject_processing_item(int(row_id), vendor=v)
        if ok:
            self.data_msg = f"처리확인 ID {row_id} 반려 완료"
            self.data_ok = True
        else:
            self.data_msg = "반려 처리 실패"
            self.data_ok = False
        self.load_processing()

    # ══════════════════════════════
    #  이벤트 — 외주업체관리 (섹션C)
    # ══════════════════════════════

    def load_vendor_tab(self):
        """외주업체 탭 진입 시 로드"""
        self.vendor_msg = ""
        self.vendor_list = get_all_vendor_info()
        self.school_master_list = get_school_master_all()

    def set_vendor_sub_tab(self, tab: str):
        self.vendor_sub_tab = tab
        self.vendor_msg = ""
        if tab == "외주업체목록":
            self.vendor_list = get_all_vendor_info()
            # 탭 전환 시 드릴다운 초기화
            self.selected_vendor_for_clients = ""
            self.selected_vendor_clients = []
        elif tab == "학교별칭":
            self.school_master_list = get_school_master_all()
        elif tab == "안전평가":
            self.load_safety_data()

    # #7 담당거래처 드릴다운 핸들러
    def show_vendor_clients(self, vendor: str):
        """외주업체목록에서 업체 클릭 시 — 해당 업체의 담당거래처만 로드"""
        if not vendor:
            return
        self.selected_vendor_for_clients = vendor
        try:
            self.selected_vendor_clients = get_customers_by_vendor(vendor) or []
        except Exception as e:
            logger.warning(f"show_vendor_clients 실패 ({vendor}): {e}")
            self.selected_vendor_clients = []

    def hide_vendor_clients(self):
        """드릴다운 닫기"""
        self.selected_vendor_for_clients = ""
        self.selected_vendor_clients = []

    # ── 업체 등록/수정 폼 세터 ──

    def set_vf_vendor(self, v: str):
        self.vf_vendor = v

    def set_vf_biz_name(self, v: str):
        self.vf_biz_name = v

    def set_vf_rep(self, v: str):
        self.vf_rep = v

    def set_vf_biz_no(self, v: str):
        self.vf_biz_no = v

    def set_vf_address(self, v: str):
        self.vf_address = v

    def set_vf_contact(self, v: str):
        self.vf_contact = v

    def set_vf_email(self, v: str):
        self.vf_email = v

    def set_vf_vehicle_no(self, v: str):
        self.vf_vehicle_no = v

    def set_vf_account(self, v: str):
        self.vf_account = v

    def set_vf_schools_text(self, v: str):
        self.vf_schools_text = v

    def save_vendor(self):
        """업체 정보 저장 (등록/수정) — P2: 학교 배정 포함"""
        if not self.vf_vendor or not self.vf_biz_name:
            self.vendor_msg = "업체 ID와 상호명은 필수입니다."
            self.vendor_ok = False
            return
        ok = save_hq_vendor_info({
            "vendor": self.vf_vendor,
            "biz_name": self.vf_biz_name,
            "rep": self.vf_rep,
            "biz_no": self.vf_biz_no,
            "address": self.vf_address,
            "contact": self.vf_contact,
            "email": self.vf_email,
            "vehicle_no": self.vf_vehicle_no,
            "account": self.vf_account,
        })
        if ok:
            # ── P2: 학교 배정 처리 ──
            assigned_count = 0
            try:
                if self.vf_schools_text.strip():
                    new_schools = [
                        s.strip() for s in self.vf_schools_text.split(",") if s.strip()
                    ]
                    # 기존 배정 학교 조회
                    existing = get_school_master_all()
                    current = [
                        s.get("school_name") for s in existing
                        if s.get("vendor") == self.vf_vendor
                    ]
                    # 새로 추가
                    to_add = [s for s in new_schools if s not in current]
                    for school in to_add:
                        db_upsert(
                            "school_master",
                            {"school_name": school, "vendor": self.vf_vendor},
                        )
                    # 제거 (vendor 해제)
                    to_remove = [s for s in current if s not in new_schools]
                    for school in to_remove:
                        db_upsert(
                            "school_master",
                            {"school_name": school, "vendor": ""},
                        )
                    assigned_count = len(new_schools)
            except Exception as e:
                logger.warning(f"학교 배정 처리 실패: {e}")

            extra = f" (학교 {assigned_count}곳 배정)" if assigned_count else ""
            self.vendor_msg = f"'{self.vf_biz_name}' 저장 완료{extra}"
            self.vendor_ok = True
            self.vendor_list = get_all_vendor_info()
            self.school_master_list = get_school_master_all()
            # 폼 초기화
            self.vf_vendor = ""
            self.vf_biz_name = ""
            self.vf_rep = ""
            self.vf_biz_no = ""
            self.vf_address = ""
            self.vf_contact = ""
            self.vf_email = ""
            self.vf_vehicle_no = ""
            self.vf_account = ""
            self.vf_schools_text = ""
        else:
            self.vendor_msg = "저장 실패"
            self.vendor_ok = False

    def load_vendor_for_edit(self, vendor_id: str):
        """업체 정보 편집을 위해 로드 — P2: 학교 배정 포함"""
        v = next((x for x in self.vendor_list if x.get("vendor") == vendor_id), None)
        if v:
            self.vf_vendor = v.get("vendor", "")
            self.vf_biz_name = v.get("biz_name", "")
            self.vf_rep = v.get("rep", "")
            self.vf_biz_no = v.get("biz_no", "")
            self.vf_address = v.get("address", "")
            self.vf_contact = v.get("contact", "")
            self.vf_email = v.get("email", "")
            self.vf_vehicle_no = v.get("vehicle_no", "")
            self.vf_account = v.get("account", "")
            # P2: 배정 학교 로드
            schools = get_school_master_all()
            assigned = [
                s.get("school_name", "") for s in schools
                if s.get("vendor") == vendor_id and s.get("school_name")
            ]
            self.vf_schools_text = ", ".join(assigned)
            self.vendor_sub_tab = "업체등록"

    # ── 학교 별칭 ──

    def set_alias_school_sel(self, s: str):
        self.alias_school_sel = s
        # 현재 별칭 로드
        row = next((x for x in self.school_master_list if x.get("school_name") == s), None)
        self.alias_input = row.get("alias", "") if row else ""

    def set_alias_input(self, v: str):
        self.alias_input = v

    def save_alias(self):
        """학교 별칭 저장"""
        if not self.alias_school_sel:
            self.vendor_msg = "학교를 선택하세요."
            self.vendor_ok = False
            return
        cleaned = ",".join([a.strip() for a in self.alias_input.split(",") if a.strip()])
        ok = update_school_alias(self.alias_school_sel, cleaned)
        if ok:
            self.vendor_msg = f"'{self.alias_school_sel}' 별칭 저장 완료"
            self.vendor_ok = True
            self.school_master_list = get_school_master_all()
        else:
            self.vendor_msg = "별칭 저장 실패"
            self.vendor_ok = False

    # ── 안전관리 평가 ──

    def set_eval_vendor_sel(self, v: str):
        self.eval_vendor_sel = v

    def set_eval_year(self, y: str):
        self.eval_year = y

    def set_eval_month(self, m: str):
        self.eval_month = m

    def calculate_eval(self):
        """안전관리 평가 점수 계산"""
        if not self.eval_vendor_sel:
            self.vendor_msg = "업체를 선택하세요."
            self.vendor_ok = False
            return
        ym = f"{self.eval_year}-{self.eval_month.zfill(2)}"
        result = hq_calculate_safety_score(self.eval_vendor_sel, ym)
        self.eval_result = {
            "vendor": str(result.get("vendor", "")),
            "year_month": str(result.get("year_month", "")),
            "violation_score": str(result.get("violation_score", 0)),
            "checklist_score": str(result.get("checklist_score", 0)),
            "daily_check_score": str(result.get("daily_check_score", 0)),
            "education_score": str(result.get("education_score", 0)),
            "total_score": str(result.get("total_score", 0)),
            "grade": str(result.get("grade", "")),
        }
        self.vendor_msg = f"{self.eval_vendor_sel} 평가 완료: {result['grade']}등급 ({result['total_score']}점)"
        self.vendor_ok = True
        self.safety_scores_list = hq_get_safety_scores()

    def load_safety_data(self):
        """안전 평가 데이터 로드"""
        self.safety_scores_list = hq_get_safety_scores()
        self.violations_list = hq_get_violations()

    # ── 위반 기록 ──

    def set_viol_vendor(self, v: str):
        self.viol_vendor = v

    def set_viol_driver(self, v: str):
        self.viol_driver = v

    def set_viol_date(self, v: str):
        self.viol_date = v

    def set_viol_type(self, v: str):
        self.viol_type = v

    def set_viol_location(self, v: str):
        self.viol_location = v

    def set_viol_fine(self, v: str):
        self.viol_fine = v

    def set_viol_memo(self, v: str):
        self.viol_memo = v

    def set_viol_vendor_filter(self, v: str):
        self.viol_vendor_filter = v
        vendor = v if v != "전체" else ""
        self.violations_list = hq_get_violations(vendor=vendor)

    def save_violation(self):
        """위반 기록 저장"""
        if not self.viol_vendor or not self.viol_driver or not self.viol_date:
            self.vendor_msg = "업체, 기사, 위반일은 필수입니다."
            self.vendor_ok = False
            return
        try:
            fine = int(self.viol_fine)
        except (ValueError, TypeError):
            fine = 0
        ok = hq_add_violation(
            vendor=self.viol_vendor, driver=self.viol_driver,
            violation_date=self.viol_date, violation_type=self.viol_type,
            location=self.viol_location, fine_amount=fine,
            memo=self.viol_memo,
        )
        if ok:
            self.vendor_msg = f"위반 기록 저장 완료 ({self.viol_type} / {self.viol_date})"
            self.vendor_ok = True
            self.violations_list = hq_get_violations()
            self.viol_driver = ""
            self.viol_date = ""
            self.viol_location = ""
            self.viol_fine = "0"
            self.viol_memo = ""
        else:
            self.vendor_msg = "저장 실패"
            self.vendor_ok = False

    # ══════════════════════════════
    #  이벤트 — 수거일정 (섹션D)
    # ══════════════════════════════

    def load_schedule_tab(self):
        """수거일정 탭 진입 시 로드"""
        self.sched_msg = ""
        if not self.vendor_list:
            self.vendor_list = get_all_vendor_info()
        self.load_schedules()

    def set_sched_sub_tab(self, tab: str):
        self.sched_sub_tab = tab
        self.sched_msg = ""
        if tab == "일정조회":
            self.load_schedules()
        elif tab == "오늘현황":
            self.load_today_status()
        elif tab == "급식승인":
            self.load_meal_approval()
        elif tab == "NEIS연동":
            self.neis_meal_dates = []
            self.neis_result_msg = ""

    def set_sched_vendor_filter(self, v: str):
        self.sched_vendor_filter = v
        self.load_schedules()

    def set_sched_month_filter(self, m: str):
        self.sched_month_filter = m
        self.load_schedules()

    def load_schedules(self):
        vendor = self.sched_vendor_filter if self.sched_vendor_filter != "전체" else ""
        ym = self.sched_month_filter if self.sched_month_filter != "전체" else ""
        self.sched_rows = get_hq_schedules(vendor=vendor, year_month=ym)

    # 일정 등록 폼 세터
    def set_sf_vendor(self, v: str):
        self.sf_vendor = v

    def set_sf_month_key(self, v: str):
        self.sf_month_key = v

    def set_sf_weekdays(self, v: str):
        self.sf_weekdays = v

    def set_sf_schools(self, v: str):
        self.sf_schools = v

    def set_sf_items(self, v: str):
        self.sf_items = v

    def set_sf_driver(self, v: str):
        self.sf_driver = v

    def save_sched(self):
        """일정 저장 (중복 방지 포함)"""
        import json as _json
        if not self.sf_vendor or not self.sf_month_key:
            self.sched_msg = "업체와 월(또는 날짜)은 필수입니다."
            self.sched_ok = False
            return
        weekdays = [w.strip() for w in self.sf_weekdays.split(",") if w.strip()]
        schools = [s.strip() for s in self.sf_schools.split(",") if s.strip()]
        items = [i.strip() for i in self.sf_items.split(",") if i.strip()]
        # 중복 체크 (P1)
        schools_json = _json.dumps(schools)
        if check_schedule_duplicate(self.sf_vendor, self.sf_month_key, schools_json):
            self.sched_msg = (
                f"이미 등록된 일정입니다 — {self.sf_vendor} / "
                f"{self.sf_month_key} / {', '.join(schools)}"
            )
            self.sched_ok = False
            return
        ok = hq_save_schedule({
            "vendor": self.sf_vendor,
            "month": self.sf_month_key,
            "weekdays": _json.dumps(weekdays),
            "schools": _json.dumps(schools),
            "items": _json.dumps(items),
            "driver": self.sf_driver,
            "registered_by": "admin",
        })
        if ok:
            self.sched_msg = f"일정 저장 완료 ({self.sf_month_key})"
            self.sched_ok = True
            self.load_schedules()
        else:
            self.sched_msg = "저장 실패"
            self.sched_ok = False

    def delete_sched(self, sched_id: str):
        """일정 삭제"""
        v = self.sched_vendor_filter if self.sched_vendor_filter != "전체" else ""
        ok = hq_delete_schedule(sched_id, vendor=v)
        if ok:
            self.sched_msg = f"일정 ID {sched_id} 삭제 완료"
            self.sched_ok = True
        else:
            self.sched_msg = "삭제 실패"
            self.sched_ok = False
        self.load_schedules()

    # ══════════════════════════════
    #  오늘현황 (P1 복원)
    # ══════════════════════════════
    def set_today_vendor_filter(self, v: str):
        self.today_vendor_filter = v
        self.load_today_status()

    def load_today_status(self):
        """오늘 수거 현황 로드 (real_collection 오늘자)"""
        from datetime import datetime as _dt
        today_str = _dt.now().strftime("%Y-%m-%d")
        ym = today_str[:7]
        vendor = (
            self.today_vendor_filter
            if self.today_vendor_filter and self.today_vendor_filter != "전체"
            else ""
        )
        try:
            rows = get_filtered_collections(
                table="real_collection", vendor=vendor, year_month=ym,
            )
        except Exception as e:
            logger.warning(f"load_today_status 실패: {e}")
            rows = []
        # 오늘자만 필터
        today_rows = [r for r in rows if str(r.get("collect_date", ""))[:10] == today_str]
        total_w = 0.0
        for r in today_rows:
            try:
                total_w += float(r.get("weight", 0) or 0)
            except (ValueError, TypeError):
                pass
        self.today_collection_rows = today_rows
        self.today_total_weight = f"{total_w:,.1f}"
        self.today_total_count = len(today_rows)

    # ══════════════════════════════
    #  급식승인 워크플로우 (P1 복원)
    # ══════════════════════════════
    def set_meal_approval_vendor(self, v: str):
        self.meal_approval_vendor = v
        self.load_meal_approval()

    def set_meal_approval_month(self, m: str):
        self.meal_approval_month = m
        self.load_meal_approval()

    def set_meal_approval_status(self, s: str):
        self.meal_approval_status = s
        self.load_meal_approval()

    def set_meal_approval_driver(self, v: str):
        self.meal_approval_driver = v

    def set_meal_approval_offset(self, v: str):
        self.meal_approval_offset = v

    def set_meal_cancel_note(self, v: str):
        self.meal_cancel_note = v

    def toggle_meal_id(self, mid: str):
        sid = str(mid)
        if sid in self.meal_selected_ids:
            self.meal_selected_ids = [x for x in self.meal_selected_ids if x != sid]
        else:
            self.meal_selected_ids = self.meal_selected_ids + [sid]

    def select_all_meals(self):
        self.meal_selected_ids = [str(r.get("id", "")) for r in self.meal_draft_rows]

    def clear_meal_selection(self):
        self.meal_selected_ids = []

    def load_meal_approval(self):
        """급식승인 대기 목록 로드"""
        vendor = (
            self.meal_approval_vendor
            if self.meal_approval_vendor and self.meal_approval_vendor != "전체"
            else ""
        )
        ym = self.meal_approval_month or ""
        # 현재 보고 있는 status 의 행만 메인 리스트에 표시
        self.meal_draft_rows = get_meal_schedules(
            vendor=vendor, status=self.meal_approval_status, year_month=ym,
        )
        # 카운트 — 같은 vendor/월 기준으로 status 별 집계
        self.meal_pending_count = len(
            get_meal_schedules(vendor=vendor, status="draft", year_month=ym)
        )
        self.meal_approved_count = len(
            get_meal_schedules(vendor=vendor, status="approved", year_month=ym)
        )
        self.meal_cancelled_count = len(
            get_meal_schedules(vendor=vendor, status="cancelled", year_month=ym)
        )
        self.meal_selected_ids = []

    def approve_selected_meals(self):
        """선택한 급식일정 일괄 승인 → schedules 자동 반영"""
        if not self.meal_selected_ids:
            self.sched_msg = "승인할 항목을 선택하세요."
            self.sched_ok = False
            return
        try:
            offset = int(self.meal_approval_offset)
        except (ValueError, TypeError):
            offset = 1
        ids = [int(x) for x in self.meal_selected_ids if str(x).isdigit()]
        succ, fail = approve_meal_schedules(
            ids=ids,
            approved_by="admin",
            driver=self.meal_approval_driver or "",
            collect_offset=offset,
        )
        self.sched_msg = f"급식일정 승인 완료 — 성공 {succ}건 / 실패 {fail}건"
        self.sched_ok = succ > 0
        self.load_meal_approval()

    def cancel_selected_meals(self):
        """선택한 급식일정 일괄 반려"""
        if not self.meal_selected_ids:
            self.sched_msg = "반려할 항목을 선택하세요."
            self.sched_ok = False
            return
        ids = [int(x) for x in self.meal_selected_ids if str(x).isdigit()]
        succ, fail = cancel_meal_schedules(ids=ids, note=self.meal_cancel_note or "")
        self.sched_msg = f"급식일정 반려 완료 — 성공 {succ}건 / 실패 {fail}건"
        self.sched_ok = succ > 0
        self.load_meal_approval()

    # ── NEIS 급식연동 ──

    def set_neis_vendor(self, v: str):
        self.neis_vendor = v
        self.neis_school_list = get_neis_schools_by_vendor(v)
        self.neis_meal_dates = []
        self.neis_result_msg = ""

    def set_neis_month(self, m: str):
        self.neis_month = m

    def set_neis_school_sel(self, s: str):
        self.neis_school_sel = s

    def set_neis_collect_offset(self, v: str):
        self.neis_collect_offset = v

    def set_neis_item_type(self, v: str):
        self.neis_item_type = v

    def set_neis_driver(self, v: str):
        self.neis_driver = v

    def fetch_neis_meals(self):
        """NEIS API로 급식일 조회"""
        from zeroda_reflex.utils.neis_api import fetch_meal_dates
        if not self.neis_school_sel or not self.neis_month:
            self.sched_msg = "학교와 조회월을 선택하세요."
            self.sched_ok = False
            return
        school_info = next(
            (s for s in self.neis_school_list if s["name"] == self.neis_school_sel), None
        )
        if not school_info:
            self.sched_msg = "NEIS 학교코드가 없습니다."
            self.sched_ok = False
            return
        try:
            year = int(self.neis_month[:4])
            month = int(self.neis_month[5:7])
        except (ValueError, IndexError):
            self.sched_msg = "월 형식이 올바르지 않습니다."
            self.sched_ok = False
            return
        result = fetch_meal_dates(
            school_info["neis_edu_code"],
            school_info["neis_school_code"],
            year, month,
        )
        if result["success"]:
            self.neis_meal_dates = result.get("meal_dates", [])
            self.neis_result_msg = result["message"]
            self.sched_msg = result["message"]
            self.sched_ok = True
        else:
            self.neis_meal_dates = []
            self.neis_result_msg = result["message"]
            self.sched_msg = result["message"]
            self.sched_ok = False

    def create_neis_schedules(self):
        """NEIS 급식일 기반 수거일정 일괄 생성"""
        from datetime import timedelta
        if not self.neis_meal_dates:
            self.sched_msg = "먼저 NEIS 급식일을 조회하세요."
            self.sched_ok = False
            return
        try:
            offset = int(self.neis_collect_offset)
        except (ValueError, TypeError):
            offset = 0
        success = 0
        fail = 0
        for md in self.neis_meal_dates:
            try:
                dt = datetime.strptime(md, "%Y-%m-%d")
                collect_dt = dt + timedelta(days=offset)
                collect_date = collect_dt.strftime("%Y-%m-%d")
                ok = save_neis_meal_schedule(
                    school_name=self.neis_school_sel,
                    vendor=self.neis_vendor,
                    meal_date=md,
                    collect_date=collect_date,
                    item_type=self.neis_item_type,
                    driver=self.neis_driver,
                )
                if ok:
                    success += 1
                else:
                    fail += 1
            except Exception as e:
                logger.warning("NEIS 일정 저장 실패: %s", e)
                fail += 1
        self.sched_msg = f"수거일정 생성 완료: 성공 {success}건, 실패 {fail}건"
        self.sched_ok = success > 0
        self.load_schedules()

    # ══════════════════════════════
    #  이벤트 — 정산관리 (섹션E)
    # ══════════════════════════════

    def set_settle_year(self, y: str):
        self.settle_year = y
        self.load_settlement()

    def set_settle_month(self, m: str):
        self.settle_month = m
        self.load_settlement()

    def set_settle_vendor(self, v: str):
        self.settle_vendor = v
        self.load_settlement()

    def set_settle_cust_type(self, v: str):
        self.settle_cust_type = v
        self.load_settlement()

    def load_settlement(self):
        try:
            y = int(self.settle_year)
            m = int(self.settle_month)
        except (ValueError, TypeError):
            return
        vendor = self.settle_vendor if self.settle_vendor != "전체" else ""

        # 1) 원본 수거 데이터 로드
        rows = get_settlement_data(y, m, vendor)

        # 2) 단가 보정 + cust_type/tax_type 부여
        rows = _correct_settlement_prices(rows, vendor)

        # 3) 거래처 유형 필터
        if self.settle_cust_type and self.settle_cust_type != "전체":
            rows = [r for r in rows if r.get("cust_type") == self.settle_cust_type]

        # 4) settlement_list (학교 단위, PDF/이메일용)
        school_map: dict[str, dict] = {}
        for r in rows:
            sn = r.get("school_name", "")
            if not sn:
                continue
            try:
                w = float(r.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            try:
                amt = float(r.get("amount", 0) or 0)
            except (ValueError, TypeError):
                amt = 0.0
            entry = school_map.setdefault(sn, {
                "name": sn,
                "cust_type": r.get("cust_type", ""),
                "fixed_fee": r.get("fixed_fee", "0"),
                "total_weight": 0.0,
                "amount": 0.0,
            })
            entry["total_weight"] += w
            entry["amount"] += amt
        self.settlement_list = [
            {
                "name": v["name"],
                "cust_type": v["cust_type"],
                "fixed_fee": v["fixed_fee"],
                "total_weight": str(round(v["total_weight"], 1)),
                "amount": str(int(round(v["amount"]))),
            }
            for v in school_map.values()
        ]

        # 5) 세금 분류별 합계 계산
        total_weight = 0.0
        supply_amount = 0.0   # 공급가액 (vat_10 + tax_free)
        vat_amount = 0.0
        fixed_fee_amount = 0.0
        fixed_fee_vat_amount = 0.0
        tax_free_amount = 0.0

        # fixed_fee 계열은 학교 단위 1건만 카운트해야 하므로 set 으로 중복 제거
        seen_fixed: set[str] = set()
        for r in rows:
            try:
                w = float(r.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            try:
                amt = float(r.get("amount", 0) or 0)
            except (ValueError, TypeError):
                amt = 0.0
            try:
                ff = float(r.get("fixed_fee", 0) or 0)
            except (ValueError, TypeError):
                ff = 0.0
            total_weight += w
            ttype = r.get("tax_type", "vat_10")
            sn = r.get("school_name", "")
            if ttype == "tax_free":
                tax_free_amount += amt
            elif ttype == "vat_10":
                supply_amount += amt
                vat_amount += amt * 0.1
            elif ttype == "fixed_fee":
                key = f"ff::{sn}"
                if key not in seen_fixed and ff > 0:
                    fixed_fee_amount += ff
                    seen_fixed.add(key)
            elif ttype == "fixed_fee_vat":
                key = f"ffv::{sn}"
                if key not in seen_fixed and ff > 0:
                    fixed_fee_vat_amount += ff
                    seen_fixed.add(key)

        # 합산 (UI KPI 표시용)
        gross_supply = supply_amount + tax_free_amount + fixed_fee_amount + fixed_fee_vat_amount
        gross_vat = vat_amount + (fixed_fee_vat_amount * 0.1)
        grand_total = gross_supply + gross_vat

        self.settle_rows = rows
        self.settle_summary = {
            "total_weight": str(round(total_weight, 1)),
            "total_amount": str(int(round(gross_supply))),
            "supply_amount": str(int(round(supply_amount))),
            "tax_free_amount": str(int(round(tax_free_amount))),
            "fixed_fee_amount": str(int(round(fixed_fee_amount))),
            "fixed_fee_vat_amount": str(int(round(fixed_fee_vat_amount))),
            "vat": str(int(round(gross_vat))),
            "grand_total": str(int(round(grand_total))),
            "count": str(len(rows)),
        }

        # KPI 라벨 (필터된 단일 유형일 때 안내 문구)
        ct = self.settle_cust_type
        if ct in ("학교", "기타1(면세사업장)"):
            self.settle_tax_label = "공급가액 (면세)"
        elif ct == "기타":
            self.settle_tax_label = "월 고정비"
        elif ct == "기타2(부가세포함)":
            self.settle_tax_label = "월 고정비 + VAT"
        elif ct in ("기업", "관공서", "일반업장"):
            self.settle_tax_label = "공급가액 (+ VAT 10%)"
        else:
            self.settle_tax_label = "공급가액"

        # P3: 수신자(거래처) 정보 자동 로드 + 이메일 템플릿 생성
        self._load_receiver_info(vendor)
        self._build_email_template()

    def _load_receiver_info(self, vendor: str):
        """P3: 거래처(수신자) 정보 자동 로드"""
        self.settle_rcv_rep = ""
        self.settle_rcv_biz_no = ""
        self.settle_rcv_phone = ""
        self.settle_rcv_address = ""
        self.settle_rcv_biz_type = ""
        self.settle_rcv_biz_item = ""
        if not vendor:
            return
        try:
            customers = get_customers_by_vendor(vendor) or []
        except Exception:
            customers = []
        if not customers:
            return
        # 거래처 유형 필터가 적용된 경우, 첫번째 일치 항목 사용
        target = None
        if self.settle_cust_type and self.settle_cust_type != "전체":
            target = next(
                (c for c in customers if c.get("cust_type") == self.settle_cust_type),
                None,
            )
        if not target:
            target = customers[0]
        self.settle_rcv_rep = str(target.get("ceo", "") or "")
        self.settle_rcv_biz_no = str(target.get("biz_no", "") or "")
        self.settle_rcv_phone = str(target.get("phone", "") or "")
        self.settle_rcv_address = str(target.get("address", "") or "")
        self.settle_rcv_biz_type = str(target.get("biz_type", "") or "")
        self.settle_rcv_biz_item = str(target.get("biz_item", "") or "")
        # 이메일 자동 채움 (비어있을 때만)
        if not self.email_to:
            self.email_to = str(target.get("email", "") or "")

    def _build_email_template(self):
        """P3: 이메일 제목/본문 자동 생성"""
        try:
            from zeroda_reflex.utils.database import get_vendor_info
            vinfo = get_vendor_info(self.settle_vendor) or {}
        except Exception:
            vinfo = {}
        biz_name = str(vinfo.get("biz_name") or self.settle_vendor or "")
        contact = str(vinfo.get("contact", "") or "")

        self.settle_email_subject = (
            f"[{biz_name}] {self.settle_year}년 {self.settle_month}월 거래명세서"
        )

        overdue_body = ""
        try:
            od_amt = float(self.overdue_amount or 0)
        except (ValueError, TypeError):
            od_amt = 0.0
        if od_amt > 0:
            memo_line = f"비고: {self.overdue_memo}\n" if self.overdue_memo else ""
            overdue_body = (
                "\n※ 미납 안내\n"
                f"미납금액: {int(od_amt):,}원\n"
                f"미납개월: {self.overdue_months or '확인 필요'}\n"
                f"{memo_line}"
                "조속한 납부 부탁드립니다.\n"
            )

        self.settle_email_body = (
            "담당자님께,\n\n"
            f"안녕하세요. {biz_name} 입니다.\n\n"
            f"{self.settle_year}년 {self.settle_month}월 거래명세서를 첨부하여 발송드립니다.\n"
            "확인 후 문의사항이 있으시면 연락 주시기 바랍니다.\n"
            f"{overdue_body}\n"
            "감사합니다.\n"
            f"{biz_name} 드림\n"
            f"연락처: {contact}"
        )

    def set_settle_email_subject(self, v: str):
        self.settle_email_subject = v

    def set_settle_email_body(self, v: str):
        self.settle_email_body = v

    def rebuild_email_template(self):
        """P3: 이메일 템플릿 재생성"""
        self._build_email_template()

    # ── 미수금 setter ──
    def set_overdue_amount(self, v: str):
        self.overdue_amount = v or "0"

    def set_overdue_months(self, v: str):
        self.overdue_months = v

    def set_overdue_memo(self, v: str):
        self.overdue_memo = v

    @rx.var
    def has_overdue(self) -> bool:
        try:
            return float(self.overdue_amount or 0) > 0
        except (ValueError, TypeError):
            return False

    @rx.var
    def overdue_warning_text(self) -> str:
        try:
            amt = int(float(self.overdue_amount or 0))
        except (ValueError, TypeError):
            amt = 0
        if amt <= 0:
            return ""
        return f"⚠️ 미납금액 {amt:,}원이 이메일/문자 본문에 자동 포함됩니다."

    # ══════════════════════════════
    #  이벤트 — 탄소감축 (섹션E)
    # ══════════════════════════════

    def set_carbon_year(self, y: str):
        self.carbon_year = y
        self.load_carbon()

    def set_carbon_month(self, m: str):
        self.carbon_month = m
        self.load_carbon()

    def load_carbon(self):
        try:
            y = int(self.carbon_year)
        except (ValueError, TypeError):
            return
        m = 0
        if self.carbon_month != "전체":
            try:
                m = int(self.carbon_month)
            except (ValueError, TypeError):
                pass
        data = get_carbon_data(y, m)
        ranking = data.pop("school_ranking", [])
        # 차트용 숫자 필드 추가
        for r in ranking:
            r["weight_num"] = float(r.get("total_weight", 0) or 0)
            r["carbon_num"] = float(r.get("carbon", 0) or 0)
        self.carbon_school_ranking = ranking
        self.carbon_data = data

        # P3: 품목별 기여도 (음식물 0.47, 재활용 0.21, 일반 0.09)
        def _f(v):
            try:
                return float(str(v).replace(",", "") or 0)
            except (ValueError, TypeError):
                return 0.0
        food = _f(data.get("food_kg", 0))
        recycle = _f(data.get("recycle_kg", 0))
        general = _f(data.get("general_kg", 0))
        self.carbon_by_item = [
            {
                "item_type": "음식물",
                "weight": f"{food:,.1f}",
                "carbon_kg": f"{food * 0.47:,.1f}",
            },
            {
                "item_type": "재활용",
                "weight": f"{recycle:,.1f}",
                "carbon_kg": f"{recycle * 0.21:,.1f}",
            },
            {
                "item_type": "일반",
                "weight": f"{general:,.1f}",
                "carbon_kg": f"{general * 0.09:,.1f}",
            },
        ]

        # P3: 월별 추이 (연간 조회 시에만)
        if self.carbon_month == "전체":
            trend = []
            for mm in range(1, 13):
                try:
                    md = get_carbon_data(y, mm) or {}
                except Exception:
                    md = {}
                cr = _f(md.get("carbon_reduced", 0))
                if cr > 0:
                    trend.append({
                        "month": f"{mm}월",
                        "carbon_num": cr,
                        "carbon_kg": f"{cr:,.1f}",
                    })
            self.carbon_monthly_trend = trend
        else:
            self.carbon_monthly_trend = []

    def download_carbon_csv(self):
        """탄소감축 데이터 CSV 다운로드"""
        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["학교명", "수거량(kg)", "CO₂감축(kg)"])
        for r in self.carbon_school_ranking:
            writer.writerow([
                r.get("school_name", ""),
                r.get("total_weight", ""),
                r.get("carbon", ""),
            ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return rx.download(
            data=csv_bytes,
            filename=f"탄소감축_{self.carbon_year}_{self.carbon_month}.csv",
        )

    # ══════════════════════════════
    #  이벤트 — 안전관리 (섹션F)
    # ══════════════════════════════

    def set_safety_vendor_filter(self, v: str):
        self.safety_vendor_filter = v
        self.load_safety()

    def set_safety_sub_tab(self, t: str):
        self.safety_sub_tab = t
        self.load_safety()

    def set_daily_check_month(self, m: str):
        self.daily_check_month = m
        if self.safety_sub_tab == "일일점검":
            self._load_daily_checks()

    def load_safety(self):
        """서브탭별 안전관리 데이터 라우팅 (P1 복원)"""
        sub = self.safety_sub_tab
        if sub == "안전교육":
            self._load_safety_education()
        elif sub == "점검결과":
            self._load_safety_checklist()
        elif sub == "사고보고":
            self._load_accident_reports()
        elif sub == "일일점검":
            self._load_daily_checks()
        elif sub == "기사모니터링":
            self._load_driver_monitor()
        else:
            # 기본: 안전교육 + 사고
            self._load_safety_education()
            self._load_accident_reports()

    # ── 서브탭별 로더 ──
    def _load_safety_education(self):
        vendor = self.safety_vendor_filter if self.safety_vendor_filter != "전체" else ""
        self.safety_edu_rows = get_hq_safety_education(vendor)

    def _load_accident_reports(self):
        vendor = self.safety_vendor_filter if self.safety_vendor_filter != "전체" else ""
        self.safety_accident_rows = get_hq_accident_reports(vendor)

    def _load_safety_checklist(self):
        vendor = self.safety_vendor_filter if self.safety_vendor_filter != "전체" else ""
        rows = get_hq_safety_checklist(vendor)
        self.safety_checklist_rows = rows
        # P2: 불합격(>0) 점검 건수 카운트
        fail = 0
        for r in rows:
            try:
                if int(r.get("total_fail", 0) or 0) > 0:
                    fail += 1
            except (ValueError, TypeError):
                pass
        self.safety_checklist_fail_count = fail

    def _load_daily_checks(self):
        vendor = self.safety_vendor_filter if self.safety_vendor_filter != "전체" else ""
        ym = self.daily_check_month or datetime.now().strftime("%Y-%m")
        rows = get_hq_daily_checks(vendor, ym)
        self.daily_check_rows = rows
        ok = 0
        fail = 0
        for r in rows:
            try:
                ok += int(r.get("total_ok", 0) or 0)
            except (ValueError, TypeError):
                pass
            try:
                fail += int(r.get("total_fail", 0) or 0)
            except (ValueError, TypeError):
                pass
        self.daily_check_ok_count = ok
        self.daily_check_fail_count = fail
        total = ok + fail
        self.daily_check_ok_rate = (
            f"{round(ok / total * 100, 1)}%" if total > 0 else "0%"
        )

    def _load_driver_monitor(self):
        """기사 활동 모니터링 — 오늘 수거 입력 시각 기준 상태 분류"""
        vendor_filter = self.safety_vendor_filter if self.safety_vendor_filter != "전체" else ""
        try:
            users = get_all_users() or []
        except Exception:
            users = []
        drivers = [u for u in users if u.get("role") == "driver"]
        if vendor_filter:
            drivers = [d for d in drivers if d.get("vendor") == vendor_filter]

        # 오늘 수거 데이터 (vendor 무관 전체 한 번에)
        today = datetime.now().strftime("%Y-%m-%d")
        ym = datetime.now().strftime("%Y-%m")
        try:
            collections = get_filtered_collections(
                table="real_collection",
                vendor=vendor_filter,
                year_month=ym,
            )
        except Exception:
            collections = []
        today_data = [c for c in collections if str(c.get("collect_date", "")) == today]

        now = datetime.now()
        result = []
        ALERT_MIN = 30
        WARN_MIN = 60
        EMERG_MIN = 120

        for drv in drivers:
            driver_name = drv.get("name") or drv.get("user_id") or ""
            vendor_n = drv.get("vendor", "")
            mine = [c for c in today_data if c.get("driver") == driver_name]
            if not mine:
                result.append({
                    "status":      "⚫ 금일미입력",
                    "driver":      driver_name,
                    "vendor":      vendor_n,
                    "last_school": "-",
                    "last_time":   "-",
                    "elapsed":     "-",
                })
                continue
            try:
                last = max(
                    mine,
                    key=lambda x: str(x.get("submitted_at", x.get("collect_time", ""))),
                )
            except Exception:
                last = mine[-1]
            last_time_str = str(last.get("submitted_at", last.get("collect_time", "")))
            elapsed_min = 9999.0
            try:
                last_dt = datetime.strptime(last_time_str[:16], "%Y-%m-%d %H:%M")
                elapsed_min = (now - last_dt).total_seconds() / 60
            except Exception:
                pass
            if elapsed_min <= ALERT_MIN:
                status = "🟢 정상"
            elif elapsed_min <= WARN_MIN:
                status = "🟡 주의"
            elif elapsed_min <= EMERG_MIN:
                status = "🟠 경고"
            else:
                status = "🔴 긴급확인"
            result.append({
                "status":      status,
                "driver":      driver_name,
                "vendor":      vendor_n,
                "last_school": str(last.get("school_name", "-")),
                "last_time":   last_time_str[:16] if last_time_str else "-",
                "elapsed":     f"{int(elapsed_min)}분" if elapsed_min < 9999 else "-",
            })

        self.driver_monitor_rows = result
        self.monitor_normal_count = sum(1 for r in result if "정상" in r["status"])
        self.monitor_caution_count = sum(1 for r in result if "주의" in r["status"])
        self.monitor_warning_count = sum(
            1 for r in result if "경고" in r["status"] and "긴급" not in r["status"]
        )
        self.monitor_emergency_count = sum(1 for r in result if "긴급" in r["status"])

    # ── 사고 상태 변경 ──
    def set_accident_status_id(self, v: str):
        self.accident_status_id = v

    def set_accident_new_status(self, v: str):
        self.accident_new_status = v

    def update_accident(self):
        """사고 상태 변경 핸들러"""
        try:
            aid = int(self.accident_status_id)
        except (ValueError, TypeError):
            self.safety_msg = "사고 ID를 숫자로 입력하세요."
            self.safety_ok = False
            return
        ok = update_accident_status(aid, self.accident_new_status)
        if ok:
            self.safety_msg = f"✅ 사고 ID {aid} → {self.accident_new_status}"
            self.safety_ok = True
            self._load_accident_reports()
        else:
            self.safety_msg = "❌ 상태 변경 실패"
            self.safety_ok = False

    # ══════════════════════════════
    #  이벤트 — 폐기물분석 (섹션F)
    # ══════════════════════════════

    def set_analytics_year(self, y: str):
        self.analytics_year = y
        self.load_analytics()

    def set_analytics_month(self, m: str):
        self.analytics_month = m
        self.load_analytics()

    def load_analytics(self):
        """폐기물 발생 분석 데이터 로드"""
        try:
            y = int(self.analytics_year)
        except (ValueError, TypeError):
            return
        m = 0
        if self.analytics_month != "전체":
            try:
                m = int(self.analytics_month)
            except (ValueError, TypeError):
                pass
        result = get_waste_analytics(y, m)
        # 차트용 숫자 필드 추가 (recharts는 숫자 데이터 필요)
        by_item = result.get("by_item", [])
        for r in by_item:
            r["weight_num"] = float(r.get("weight", 0) or 0)
            r["count_num"] = int(r.get("count", 0) or 0)
        self.analytics_by_item = by_item

        by_school = result.get("by_school", [])
        for r in by_school:
            r["weight_num"] = float(r.get("weight", 0) or 0)
        self.analytics_by_school = by_school

        by_vendor = result.get("by_vendor", [])
        for r in by_vendor:
            r["weight_num"] = float(r.get("weight", 0) or 0)
        self.analytics_by_vendor = by_vendor

        by_month = result.get("by_month", [])
        for r in by_month:
            r["weight_num"] = float(r.get("weight", 0) or 0)
            r["count_num"] = int(r.get("count", 0) or 0)
        self.analytics_by_month = by_month

        self.analytics_data = {
            "total_weight": str(result.get("total_weight", 0)),
            "total_count": str(result.get("total_count", 0)),
            "avg_weight": str(result.get("avg_weight", 0)),
        }

        # P3: 전월 대비 변화율 (MoM)
        self.analytics_mom_change = "0%"
        if self.analytics_month != "전체":
            try:
                cur_m = int(self.analytics_month)
                if cur_m > 1:
                    prev = get_waste_analytics(y, cur_m - 1) or {}
                    pw = float(prev.get("total_weight", 0) or 0)
                    cw = float(result.get("total_weight", 0) or 0)
                    if pw > 0:
                        chg = ((cw - pw) / pw) * 100
                        self.analytics_mom_change = f"{chg:+.1f}%"
            except Exception:
                pass

        # P3: 요일별/계절별/기사별 분석 (수거 데이터 직접 집계)
        self._calc_extended_analytics(y, m)

    def _calc_extended_analytics(self, year: int, month: int):
        """P3: 요일별/계절별/기사별 통계 계산"""
        from datetime import datetime as dt
        WEEKDAY_NAMES = ["월", "화", "수", "목", "금", "토", "일"]
        SEASON_MAP = {
            1: "겨울", 2: "겨울", 3: "봄", 4: "봄", 5: "봄", 6: "여름",
            7: "여름", 8: "여름", 9: "가을", 10: "가을", 11: "가을", 12: "겨울",
        }
        try:
            if month > 0:
                ym = f"{year}-{str(month).zfill(2)}"
                rows = get_filtered_collections(year_month=ym) or []
            else:
                rows = []
                for mm in range(1, 13):
                    ym = f"{year}-{str(mm).zfill(2)}"
                    try:
                        rows.extend(get_filtered_collections(year_month=ym) or [])
                    except Exception:
                        continue
        except Exception:
            rows = []

        wd_map = {d: [] for d in WEEKDAY_NAMES}
        season_map: dict = {"봄": [], "여름": [], "가을": [], "겨울": []}
        driver_map: dict = {}

        for c in rows:
            try:
                w = float(c.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            d_str = str(c.get("collect_date", "") or "")
            try:
                d_obj = dt.strptime(d_str[:10], "%Y-%m-%d")
                wd_map[WEEKDAY_NAMES[d_obj.weekday()]].append(w)
                season_map[SEASON_MAP[d_obj.month]].append(w)
            except Exception:
                pass
            drv = str(c.get("driver", "") or "")
            sch = str(c.get("school_name", "") or "")
            if drv:
                if drv not in driver_map:
                    driver_map[drv] = {"weights": [], "schools": set()}
                driver_map[drv]["weights"].append(w)
                if sch:
                    driver_map[drv]["schools"].add(sch)

        wd_result = []
        for wd in WEEKDAY_NAMES:
            vals = wd_map[wd]
            if vals:
                wd_result.append({
                    "weekday": wd,
                    "avg_kg": f"{sum(vals)/len(vals):,.1f}",
                    "total_kg": f"{sum(vals):,.1f}",
                    "avg_num": round(sum(vals)/len(vals), 1),
                    "count": str(len(vals)),
                })
        self.analytics_by_weekday = wd_result

        season_result = []
        for sn in ["봄", "여름", "가을", "겨울"]:
            vals = season_map[sn]
            if vals:
                season_result.append({
                    "season": sn,
                    "total_kg": f"{sum(vals):,.1f}",
                    "avg_daily_kg": f"{sum(vals)/len(vals):,.1f}",
                    "total_num": round(sum(vals), 1),
                    "count": str(len(vals)),
                })
        self.analytics_by_season = season_result

        driver_result = []
        for drv, info in driver_map.items():
            vals = info["weights"]
            driver_result.append({
                "driver": drv,
                "total_kg": f"{sum(vals):,.1f}",
                "total_num": round(sum(vals), 1),
                "count": str(len(vals)),
                "schools": str(len(info["schools"])),
            })
        driver_result.sort(key=lambda x: x["total_num"], reverse=True)
        self.analytics_by_driver = driver_result

    # ══════════════════════════════
    #  P2: 거래처관리 (섹션 신규)
    # ══════════════════════════════

    def set_cust_vendor_filter(self, v: str):
        self.cust_vendor_filter = v
        self.load_customers()

    def set_cust_type_filter(self, v: str):
        self.cust_type_filter = v
        self.load_customers()

    def set_cf_name(self, v: str):
        self.cf_name = v

    def set_cf_cust_type(self, v: str):
        self.cf_cust_type = v

    def set_cf_price_food(self, v: str):
        self.cf_price_food = v

    def set_cf_price_recycle(self, v: str):
        self.cf_price_recycle = v

    def set_cf_price_general(self, v: str):
        self.cf_price_general = v

    def set_cf_address(self, v: str):
        self.cf_address = v

    def set_cf_phone(self, v: str):
        self.cf_phone = v

    def set_cf_biz_no(self, v: str):
        self.cf_biz_no = v

    def set_cf_ceo(self, v: str):
        self.cf_ceo = v

    def set_cf_fixed_fee(self, v: str):
        self.cf_fixed_fee = v

    def set_cf_email(self, v: str):
        self.cf_email = v

    def load_customers(self):
        """거래처 목록 로드"""
        vendor = self.cust_vendor_filter
        if not vendor:
            self.cust_rows = []
            return
        cust_type = self.cust_type_filter if self.cust_type_filter != "전체" else None
        try:
            self.cust_rows = get_customers_by_vendor(vendor, cust_type)
        except Exception as e:
            self.cust_rows = []
            logger.error(f"[load_customers] {e}", exc_info=True)
            self.cust_msg = "❌ 거래처 조회에 실패했습니다."
            self.cust_ok = False

    def _safe_int(self, v: str) -> int:
        try:
            return int(float(v or 0))
        except (ValueError, TypeError):
            return 0

    def save_customer(self):
        """거래처 저장/수정"""
        if not self.cust_vendor_filter or not self.cf_name.strip():
            self.cust_msg = "❌ 업체와 거래처명은 필수입니다."
            self.cust_ok = False
            return
        data = {
            "vendor": self.cust_vendor_filter,
            "name": self.cf_name.strip(),
            "cust_type": self.cf_cust_type,
            "price_food": self._safe_int(self.cf_price_food),
            "price_recycle": self._safe_int(self.cf_price_recycle),
            "price_general": self._safe_int(self.cf_price_general),
            "address": self.cf_address,
            "phone": self.cf_phone,
            "biz_no": self.cf_biz_no,
            "ceo": self.cf_ceo,
            "fixed_fee": self._safe_int(self.cf_fixed_fee),
            "email": self.cf_email,
        }
        try:
            ok = save_customer(data)
        except Exception as e:
            logger.error(f"[save_customer] {e}", exc_info=True)
            self.cust_msg = "❌ 저장에 실패했습니다."
            self.cust_ok = False
            return
        if ok:
            self.cust_msg = f"✅ '{self.cf_name}' 거래처 저장 완료"
            self.cust_ok = True
            self.load_customers()
            # 폼 초기화
            self.cf_name = ""
            self.cf_cust_type = "학교"
            self.cf_price_food = "0"
            self.cf_price_recycle = "0"
            self.cf_price_general = "0"
            self.cf_address = ""
            self.cf_phone = ""
            self.cf_biz_no = ""
            self.cf_ceo = ""
            self.cf_fixed_fee = "0"
            self.cf_email = ""
        else:
            self.cust_msg = "❌ 저장 실패"
            self.cust_ok = False

    def load_customer_for_edit(self, name: str):
        """거래처 편집 로드"""
        target = next((c for c in self.cust_rows if c.get("name") == name), None)
        if not target:
            return
        self.cf_name = str(target.get("name", ""))
        self.cf_cust_type = str(target.get("cust_type", "학교") or "학교")
        self.cf_price_food = str(target.get("price_food", "0") or "0")
        self.cf_price_recycle = str(target.get("price_recycle", "0") or "0")
        self.cf_price_general = str(target.get("price_general", "0") or "0")
        self.cf_address = str(target.get("address", "") or "")
        self.cf_phone = str(target.get("phone", "") or "")
        self.cf_biz_no = str(target.get("biz_no", "") or "")
        self.cf_ceo = str(target.get("ceo", "") or "")
        self.cf_fixed_fee = str(target.get("fixed_fee", "0") or "0")
        self.cf_email = str(target.get("email", "") or "")

    def delete_customer_row(self, name: str):
        """거래처 삭제"""
        if not self.cust_vendor_filter or not name:
            return
        try:
            ok = delete_customer(self.cust_vendor_filter, name)
        except Exception as e:
            logger.error(f"[delete_customer] {e}", exc_info=True)
            self.cust_msg = "❌ 삭제에 실패했습니다."
            self.cust_ok = False
            return
        if ok:
            self.cust_msg = f"✅ '{name}' 삭제 완료"
            self.cust_ok = True
            self.load_customers()
        else:
            self.cust_msg = "❌ 삭제 실패"
            self.cust_ok = False

    # ══════════════════════════════
    #  P3: 현장사진
    # ══════════════════════════════

    def set_photo_vendor_filter(self, v: str):
        self.photo_vendor_filter = v

    def set_photo_type_filter(self, v: str):
        self.photo_type_filter = v

    def set_photo_date_from(self, v: str):
        self.photo_date_from = v

    def set_photo_date_to(self, v: str):
        self.photo_date_to = v

    def load_photos(self):
        """현장사진 목록 로드 (필터 적용, 최대 200건)"""
        try:
            ptype = ""
            if self.photo_type_filter and self.photo_type_filter != "전체":
                ptype = self.photo_type_filter
            rows = get_photo_records_all(
                vendor=self.photo_vendor_filter or "",
                photo_type=ptype,
                date_from=self.photo_date_from or "",
                date_to=self.photo_date_to or "",
                limit=200,
            )
            # 한글 라벨 추가
            for r in rows:
                pt = str(r.get("photo_type", "") or "")
                r["photo_type_kr"] = PHOTO_TYPE_MAP.get(pt, pt)
            self.photo_rows = rows
            self.photo_msg = f"✅ {len(rows)}건 조회"
        except Exception as e:
            self.photo_rows = []
            logger.error(f"[load_photo_records] {e}", exc_info=True)
            self.photo_msg = "❌ 조회에 실패했습니다."

    # ══════════════════════════════
    #  P2: 폐기물분석 — 서브탭 + 이상치 + 기상분석
    # ══════════════════════════════

    def set_analytics_sub_tab(self, tab: str):
        self.analytics_sub_tab = tab

    def set_anomaly_threshold(self, v: str):
        self.anomaly_threshold = v

    def detect_anomalies(self):
        """Z-Score 기반 이상치 탐지"""
        import statistics
        try:
            year = self.analytics_year
            month = self.analytics_month
            ym = year if month == "전체" else f"{year}-{month.zfill(2)}"
            all_colls = get_filtered_collections(year_month=ym)
        except Exception as e:
            logger.error(f"[load_anomaly_data] {e}", exc_info=True)
            self.anomaly_msg = "❌ 데이터 조회에 실패했습니다."
            self.anomaly_rows = []
            self.anomaly_count = 0
            return

        # 일별 집계
        daily: dict = {}
        for c in all_colls:
            date = str(c.get("collect_date", "") or "")
            if not date:
                continue
            try:
                w = float(c.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            daily[date] = daily.get(date, 0.0) + w

        if len(daily) < 3:
            self.anomaly_rows = []
            self.anomaly_count = 0
            self.anomaly_mean = "0"
            self.anomaly_std = "0"
            self.anomaly_msg = "⚠️ 최소 3일 이상의 데이터가 필요합니다."
            return

        values = list(daily.values())
        mean_val = statistics.mean(values)
        std_val = statistics.stdev(values) if len(values) > 1 else 1.0
        if std_val == 0:
            std_val = 1.0

        self.anomaly_mean = f"{mean_val:,.1f}"
        self.anomaly_std = f"{std_val:,.1f}"

        try:
            threshold = float(self.anomaly_threshold)
        except (ValueError, TypeError):
            threshold = 2.0

        result = []
        for date, total_kg in sorted(daily.items()):
            z = (total_kg - mean_val) / std_val
            if abs(z) > threshold:
                result.append({
                    "collect_date": date,
                    "total_kg": f"{total_kg:,.1f}",
                    "z_score": f"{z:+.2f}",
                    "status": "⚠️ 이상",
                })
        self.anomaly_rows = result
        self.anomaly_count = len(result)
        self.anomaly_msg = f"✅ {len(daily)}일 분석 — 이상치 {len(result)}건 탐지"

    def set_weather_start_date(self, v: str):
        self.weather_start_date = v

    def set_weather_end_date(self, v: str):
        self.weather_end_date = v

    def run_weather_analysis(self):
        """기상 상관분석 실행"""
        from zeroda_reflex.utils.weather_service import fetch_daily_weather
        import statistics

        if not self.weather_start_date or not self.weather_end_date:
            self.weather_msg = "❌ 시작일과 종료일을 입력하세요."
            self.weather_ok = False
            return

        try:
            wresp = fetch_daily_weather(self.weather_start_date, self.weather_end_date)
        except Exception as e:
            logger.error(f"[load_weather_analysis] 기상 API: {e}", exc_info=True)
            self.weather_msg = "❌ 기상 데이터 조회에 실패했습니다."
            self.weather_ok = False
            return

        if not wresp.get("success"):
            self.weather_msg = f"❌ {wresp.get('message', '기상 데이터 조회 실패')}"
            self.weather_ok = False
            return

        weather_data = wresp.get("data", [])
        if not weather_data:
            self.weather_msg = "❌ 기상 데이터가 비어 있습니다."
            self.weather_ok = False
            return

        # 수거 데이터 일별 집계 (연도 기준 조회 후 범위 필터)
        try:
            year = self.weather_start_date[:4]
            all_colls = get_filtered_collections(year_month=year)
        except Exception as e:
            logger.error(f"[load_weather_analysis] 수거 데이터: {e}", exc_info=True)
            self.weather_msg = "❌ 수거 데이터 조회에 실패했습니다."
            self.weather_ok = False
            return

        daily_kg: dict = {}
        for c in all_colls:
            date = str(c.get("collect_date", "") or "")
            if not date or not (self.weather_start_date <= date <= self.weather_end_date):
                continue
            try:
                w = float(c.get("weight", 0) or 0)
            except (ValueError, TypeError):
                w = 0.0
            daily_kg[date] = daily_kg.get(date, 0.0) + w

        # 병합
        merged = []
        for w in weather_data:
            date_key = str(w.get("date", ""))
            if date_key in daily_kg:
                try:
                    merged.append({
                        "date_key": date_key,
                        "total_kg": float(daily_kg[date_key]),
                        "temp_avg": float(w.get("temp_avg", 0) or 0),
                        "rain": float(w.get("rain", 0) or 0),
                        "humidity": float(w.get("humidity", 0) or 0),
                        "wind": float(w.get("wind", 0) or 0),
                    })
                except (ValueError, TypeError):
                    continue

        if len(merged) < 5:
            self.weather_msg = f"⚠️ 데이터가 부족합니다 ({len(merged)}일 / 최소 5일 필요)."
            self.weather_ok = False
            return

        # 피어슨 상관계수
        def pearson(x_list, y_list):
            n = len(x_list)
            if n < 3:
                return 0.0
            mx = statistics.mean(x_list)
            my = statistics.mean(y_list)
            try:
                sx = statistics.stdev(x_list)
                sy = statistics.stdev(y_list)
            except statistics.StatisticsError:
                return 0.0
            if sx == 0 or sy == 0:
                return 0.0
            cov = sum((x - mx) * (y - my) for x, y in zip(x_list, y_list)) / (n - 1)
            return cov / (sx * sy)

        kgs = [m["total_kg"] for m in merged]
        self.weather_corr_temp = f"{pearson(kgs, [m['temp_avg'] for m in merged]):+.3f}"
        self.weather_corr_rain = f"{pearson(kgs, [m['rain'] for m in merged]):+.3f}"
        self.weather_corr_humidity = f"{pearson(kgs, [m['humidity'] for m in merged]):+.3f}"
        self.weather_corr_wind = f"{pearson(kgs, [m['wind'] for m in merged]):+.3f}"

        # 우천 vs 맑은날
        rainy = [m["total_kg"] for m in merged if m["rain"] > 0.5]
        clear = [m["total_kg"] for m in merged if m["rain"] <= 0.5]
        rainy_avg = statistics.mean(rainy) if rainy else 0.0
        clear_avg = statistics.mean(clear) if clear else 0.0
        diff_pct = ((rainy_avg - clear_avg) / clear_avg * 100) if clear_avg > 0 else 0.0

        self.weather_rainy_avg = f"{rainy_avg:,.1f}"
        self.weather_clear_avg = f"{clear_avg:,.1f}"
        self.weather_diff_pct = f"{diff_pct:+.1f}"

        # 온도 구간별 평균
        bins = [
            (-100.0, 0.0, "영하"),
            (0.0, 10.0, "0~10°C"),
            (10.0, 20.0, "10~20°C"),
            (20.0, 30.0, "20~30°C"),
            (30.0, 100.0, "30°C+"),
        ]
        temp_bins = []
        for low, high, label in bins:
            subset = [m["total_kg"] for m in merged if low <= m["temp_avg"] < high]
            if subset:
                temp_bins.append({
                    "temp_range": label,
                    "avg_kg": f"{statistics.mean(subset):,.1f}",
                    "count": str(len(subset)),
                })
        self.weather_temp_bins = temp_bins
        self.weather_msg = f"✅ {len(merged)}일 데이터 분석 완료"
        self.weather_ok = True

    # ══════════════════════════════
    #  PDF 다운로드 핸들러
    # ══════════════════════════════

    def download_statement_pdf(self):
        """정산 거래명세서 PDF 다운로드 (본사관리자)"""
        from zeroda_reflex.utils.pdf_export import build_statement_pdf
        from zeroda_reflex.utils.database import (
            get_vendor_info, get_settlement_data, get_customer_details,
        )
        try:
            y = int(self.settle_year)
            m = int(self.settle_month)
        except (ValueError, TypeError):
            return None
        vendor = self.settle_vendor if self.settle_vendor else ""
        if not vendor or not self.settlement_list:
            return None
        vendor_info = get_vendor_info(vendor)
        rows = get_settlement_data(y, m, vendor)
        if not rows:
            return None
        first_school = self.settlement_list[0].get("name", "") if self.settlement_list else ""
        cust_list = get_customer_details(vendor)
        cust_info = next((c for c in cust_list if c["name"] == first_school), {})
        biz_info = {
            "biz_no": str(cust_info.get("biz_no", "")),
            "representative": str(cust_info.get("representative", "")),
            "address": str(cust_info.get("address", "")),
        }
        cust_type = str(cust_info.get("cust_type", ""))
        fixed_fee = float(cust_info.get("fixed_fee", 0) or 0)
        pdf_bytes = build_statement_pdf(
            vendor, first_school, y, m,
            rows, biz_info, vendor_info,
            cust_type, fixed_fee,
        )
        if pdf_bytes:
            return rx.download(
                data=pdf_bytes,
                filename=f"거래명세서_{vendor}_{y}-{str(m).zfill(2)}.pdf",
            )
        return None

    # ══════════════════════════════
    #  이메일 발송 핸들러 (Phase 6)
    # ══════════════════════════════

    def set_email_to(self, v: str):
        self.email_to = v
        self.email_msg = ""

    async def send_statement_email(self):
        """거래명세서 PDF를 이메일로 발송 (본사관리자)"""
        from zeroda_reflex.utils.pdf_export import build_statement_pdf
        from zeroda_reflex.utils.email_service import send_email_with_pdf
        from zeroda_reflex.utils.database import get_vendor_info, get_customer_details
        if not self.email_to or "@" not in self.email_to:
            self.email_msg = "유효한 이메일 주소를 입력하세요."
            self.email_ok = False
            return
        self.email_sending = True
        self.email_msg = ""
        yield
        try:
            y = int(self.settle_year)
            m = int(self.settle_month)
        except (ValueError, TypeError):
            self.email_msg = "연/월을 선택하세요."
            self.email_ok = False
            self.email_sending = False
            return
        vendor = self.settle_vendor
        if not vendor or vendor == "전체" or not self.settlement_list:
            self.email_msg = "업체를 선택하세요."
            self.email_ok = False
            self.email_sending = False
            return
        vendor_info = get_vendor_info(vendor)
        rows = get_settlement_data(y, m, vendor)
        first_school = self.settlement_list[0].get("name", "") if self.settlement_list else ""
        cust_list = get_customer_details(vendor)
        cust_info = next((c for c in cust_list if c["name"] == first_school), {})
        biz_info = {
            "biz_no": str(cust_info.get("biz_no", "")),
            "representative": str(cust_info.get("representative", "")),
            "address": str(cust_info.get("address", "")),
        }
        cust_type = str(cust_info.get("cust_type", ""))
        fixed_fee = float(cust_info.get("fixed_fee", 0) or 0)
        pdf_bytes = build_statement_pdf(
            vendor, first_school, y, m, rows, biz_info, vendor_info, cust_type, fixed_fee,
        )
        if not pdf_bytes:
            self.email_msg = "PDF 생성 실패"
            self.email_ok = False
            self.email_sending = False
            return
        filename = f"거래명세서_{vendor}_{y}-{str(m).zfill(2)}.pdf"
        subject = f"[ZERODA] {y}년 {m}월 거래명세서 — {vendor}"
        body = (
            f"안녕하세요.\n\n"
            f"{vendor}의 {y}년 {m}월 거래명세서를 첨부하여 보내드립니다.\n\n"
            f"확인 부탁드립니다.\n감사합니다.\n\n"
        )
        # 미수금 안내 자동 포함 (P1 복원)
        try:
            _od_amt = float(self.overdue_amount or 0)
        except (ValueError, TypeError):
            _od_amt = 0.0
        if _od_amt > 0:
            body += "\n※ 미납 안내\n"
            body += f"미납금액: {int(_od_amt):,}원\n"
            if self.overdue_months:
                body += f"미납개월: {self.overdue_months}\n"
            if self.overdue_memo:
                body += f"비고: {self.overdue_memo}\n"
            body += "조속한 납부 부탁드립니다.\n\n"
        body += "— ZERODA 폐기물데이터플랫폼"
        ok, msg = send_email_with_pdf(self.email_to, subject, body, pdf_bytes, filename)
        self.email_ok = ok
        self.email_msg = msg
        self.email_sending = False

    # ══════════════════════════════
    #  SMS 발송 핸들러 (Phase 8)
    # ══════════════════════════════

    def set_sms_to(self, v: str):
        self.sms_to = v
        self.sms_msg = ""

    async def send_statement_sms(self):
        """거래명세서 요약을 SMS로 발송 (본사관리자)"""
        from zeroda_reflex.utils.sms_service import (
            send_statement_sms as _send_sms,
            build_summary_sms_text,
        )
        from zeroda_reflex.utils.database import (
            get_vendor_info, get_settlement_data,
        )

        if not self.sms_to or len(self.sms_to.replace("-", "")) < 10:
            self.sms_msg = "유효한 전화번호를 입력하세요."
            self.sms_ok = False
            return

        self.sms_sending = True
        self.sms_msg = ""
        yield

        try:
            y = int(self.settle_year)
            m = int(self.settle_month)
        except (ValueError, TypeError):
            self.sms_msg = "연/월을 선택하세요."
            self.sms_ok = False
            self.sms_sending = False
            return

        vendor = self.settle_vendor
        if vendor == "전체" or not self.settle_rows:
            self.sms_msg = "업체를 선택하고 정산 데이터를 확인하세요."
            self.sms_ok = False
            self.sms_sending = False
            return

        # 학교 단위 합계 (settlement_list 우선, 없으면 첫 row)
        if self.settlement_list:
            first = self.settlement_list[0]
            school = first.get("name", "")
            total_weight = float(first.get("total_weight", 0) or 0)
            total_amount = float(first.get("amount", 0) or 0)
            cust_type = str(first.get("cust_type", ""))
        else:
            first = self.settle_rows[0] if self.settle_rows else {}
            school = first.get("school_name", first.get("name", ""))
            total_weight = float(first.get("weight", 0) or 0)
            total_amount = float(first.get("amount", 0) or 0)
            cust_type = str(first.get("cust_type", ""))

        vendor_info = get_vendor_info(vendor) or {}
        contact = vendor_info.get("contact", "")

        try:
            od_amt = float(self.overdue_amount or 0)
        except (ValueError, TypeError):
            od_amt = 0.0

        text = build_summary_sms_text(
            vendor_name=vendor,
            school=school,
            year=y,
            month=m,
            total_weight=total_weight,
            total_amount=total_amount,
            contact=contact,
            overdue_amount=od_amt,
            overdue_months=self.overdue_months,
            cust_type=cust_type,
        )

        ok, msg = _send_sms(
            to_phone=self.sms_to,
            message=text,
            vendor_name=vendor,
            vendor_contact=contact,
        )
        self.sms_ok = ok
        self.sms_msg = msg
        self.sms_sending = False

    async def send_detail_sms(self):
        """상세 SMS (LMS, 일별 수거 내역 포함) 발송 — P1 복원"""
        from zeroda_reflex.utils.sms_service import (
            send_statement_sms as _send_sms,
            build_detail_sms_text,
        )
        from zeroda_reflex.utils.database import get_vendor_info, get_settlement_data

        if not self.sms_to or len(self.sms_to.replace("-", "")) < 10:
            self.sms_msg = "유효한 전화번호를 입력하세요."
            self.sms_ok = False
            return

        self.sms_sending = True
        self.sms_msg = ""
        yield

        try:
            y = int(self.settle_year)
            m = int(self.settle_month)
        except (ValueError, TypeError):
            self.sms_msg = "연/월을 선택하세요."
            self.sms_ok = False
            self.sms_sending = False
            return

        vendor = self.settle_vendor
        if vendor == "전체" or not self.settle_rows:
            self.sms_msg = "업체를 선택하고 정산 데이터를 확인하세요."
            self.sms_ok = False
            self.sms_sending = False
            return

        if self.settlement_list:
            first = self.settlement_list[0]
            school = first.get("name", "")
            total_weight = float(first.get("total_weight", 0) or 0)
            total_amount = float(first.get("amount", 0) or 0)
            cust_type = str(first.get("cust_type", ""))
        else:
            first = self.settle_rows[0]
            school = first.get("school_name", "")
            total_weight = sum(float(r.get("weight", 0) or 0) for r in self.settle_rows)
            total_amount = sum(float(r.get("amount", 0) or 0) for r in self.settle_rows)
            cust_type = str(first.get("cust_type", ""))

        # 일별 내역은 첫 학교 기준
        rows = [
            {
                "collect_date": r.get("collect_date", ""),
                "item_type": r.get("item_type", ""),
                "weight": r.get("weight", 0),
            }
            for r in self.settle_rows
            if r.get("school_name") == school
        ]

        vendor_info = get_vendor_info(vendor) or {}
        contact = vendor_info.get("contact", "")

        try:
            od_amt = float(self.overdue_amount or 0)
        except (ValueError, TypeError):
            od_amt = 0.0

        text = build_detail_sms_text(
            vendor_name=vendor,
            school=school,
            year=y,
            month=m,
            rows=rows,
            total_weight=total_weight,
            total_amount=total_amount,
            contact=contact,
            overdue_amount=od_amt,
            overdue_months=self.overdue_months,
            cust_type=cust_type,
        )

        ok, msg = _send_sms(
            to_phone=self.sms_to,
            message=text,
            vendor_name=vendor,
            vendor_contact=contact,
        )
        self.sms_ok = ok
        self.sms_msg = msg
        self.sms_sending = False

    # ══════════════════════════════
    #  Excel 다운로드 핸들러
    # ══════════════════════════════

    def download_collection_excel(self):
        """수거데이터 Excel 다운로드 (본사관리자 대시보드)"""
        # data_collection_rows에서 현재 필터링된 수거 데이터 추출
        from zeroda_reflex.utils.excel_export import export_collection_data
        data = self.data_collection_rows
        if not data:
            return None
        xlsx = export_collection_data(data, self.selected_year, self.selected_month)
        if xlsx:
            return rx.download(
                data=xlsx,
                filename=f"수거데이터_{self.selected_year}-{self.selected_month.zfill(2)}.xlsx"
            )
        return None

    def download_settlement_excel(self):
        """정산내역 Excel 다운로드 (본사관리자 정산관리)"""
        # settle_rows와 settle_summary에서 정산 데이터 추출
        from zeroda_reflex.utils.excel_export import export_settlement
        data = self.settle_rows
        summary = self.settle_summary if self.settle_summary else {}
        if not data:
            return None
        xlsx = export_settlement(
            data,
            summary,
            self.selected_year,
            self.selected_month
        )
        if xlsx:
            return rx.download(
                data=xlsx,
                filename=f"정산내역_{self.selected_year}-{self.selected_month.zfill(2)}.xlsx"
            )
        return None

    def download_carbon_excel(self):
        """탄소감축 데이터 Excel 다운로드 (본사관리자 탄소감축)"""
        # carbon_data와 carbon_school_ranking에서 탄소 감축 데이터 추출
        from zeroda_reflex.utils.excel_export import export_carbon_data
        if not self.carbon_data:
            return None
        xlsx = export_carbon_data(
            self.carbon_data,
            self.carbon_school_ranking,
            self.selected_year
        )
        if xlsx:
            return rx.download(
                data=xlsx,
                filename=f"탄소감축_{self.selected_year}.xlsx"
            )
        return None

    # ══════════════════════════════
    #  직인 관리 (본사관리자 — 모든 업체)
    # ══════════════════════════════

    def set_stamp_vendor_select(self, v: str):
        self.stamp_vendor_select = v

    def load_current_stamp(self):
        """선택한 업체의 현재 직인 경로 조회."""
        from zeroda_reflex.utils.database import get_vendor_info
        if not self.stamp_vendor_select:
            self.stamp_current_path = ""
            return
        info = get_vendor_info(self.stamp_vendor_select)
        self.stamp_current_path = info.get("stamp_path", "")

    async def handle_stamp_upload(self, files: list):
        """본사관리자 직인 업로드 (모든 업체 가능)."""
        import os
        import uuid
        from zeroda_reflex.utils.database import set_vendor_stamp, get_vendor_info

        STAMP_DIR = "/opt/zeroda-platform/storage/stamps"
        MAX_SIZE = 2 * 1024 * 1024
        ALLOWED_EXT = {".png", ".jpg", ".jpeg"}

        vendor = (self.stamp_vendor_select or "").strip()
        if not vendor:
            self.stamp_upload_status = "❌ 업체를 먼저 선택하세요."
            return
        if not files:
            self.stamp_upload_status = "❌ 파일이 없습니다."
            return

        self.stamp_upload_loading = True
        self.stamp_upload_status = ""
        yield

        try:
            f = files[0]
            raw = await f.read()
            if len(raw) > MAX_SIZE:
                self.stamp_upload_status = "❌ 파일 크기 2MB 초과."
                self.stamp_upload_loading = False
                return

            ext = os.path.splitext(f.filename or "")[1].lower()
            if ext not in ALLOWED_EXT:
                self.stamp_upload_status = "❌ PNG/JPG 만 가능합니다."
                self.stamp_upload_loading = False
                return

            os.makedirs(STAMP_DIR, exist_ok=True)
            safe_slug = "".join(c for c in vendor if c.isalnum() or c in "-_")[:20] or "vendor"
            fname = f"{safe_slug}_{uuid.uuid4().hex[:8]}{ext}"
            target = os.path.join(STAMP_DIR, fname)

            with open(target, "wb") as w:
                w.write(raw)

            try:
                from PIL import Image
                img = Image.open(target)
                img.verify()
            except Exception:
                os.remove(target)
                self.stamp_upload_status = "❌ 유효한 이미지가 아닙니다."
                self.stamp_upload_loading = False
                return

            updated_by = self.user_name or self.user_id or "admin"
            ok = set_vendor_stamp(vendor, target, updated_by)
            if ok:
                self.stamp_upload_status = f"✅ {vendor} 직인 등록 완료"
                self.stamp_current_path = target
            else:
                self.stamp_upload_status = "❌ DB 저장 실패"
        except Exception as e:
            logger.error(f"[upload_stamp_image] {e}", exc_info=True)
            self.stamp_upload_status = "❌ 업로드 중 오류가 발생했습니다."
        finally:
            self.stamp_upload_loading = False

    # ============================================================
    # 웨이크워드 P2-3 — 사용 통계 조회
    # ============================================================
    wake_stats_rows: list[dict] = []
    wake_stats_period: str = "7d"

    async def load_wake_stats(self):
        """기간별 사용자 호출 통계."""
        from ..utils.database import get_db
        days = {"1d": 1, "7d": 7, "30d": 30}.get(self.wake_stats_period, 7)
        conn = get_db()
        try:
            cur = conn.execute(
                "SELECT username, "
                "  SUM(CASE WHEN event_type='wake_fired'    THEN 1 ELSE 0 END) AS fired, "
                "  SUM(CASE WHEN event_type='voice_success' THEN 1 ELSE 0 END) AS success, "
                "  SUM(CASE WHEN event_type='voice_failed'  THEN 1 ELSE 0 END) AS failed, "
                "  SUM(CASE WHEN event_type='cancel'        THEN 1 ELSE 0 END) AS cancel "
                "FROM wake_stats "
                "WHERE occurred_at >= datetime('now', ?) "
                "GROUP BY username ORDER BY fired DESC",
                (f"-{days} days",),
            )
            rows = cur.fetchall()
            self.wake_stats_rows = [dict(r) for r in rows]
        except Exception as e:
            logger.warning("load_wake_stats 실패: %s", e)
            self.wake_stats_rows = []
        finally:
            conn.close()