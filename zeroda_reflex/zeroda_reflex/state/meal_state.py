# zeroda_reflex/state/meal_state.py
# 급식담당자 상태 관리 — 6메뉴
import reflex as rx
import logging
from datetime import datetime
from zeroda_reflex.state.auth_state import AuthState
from zeroda_reflex.utils.filters import safe_int

logger = logging.getLogger(__name__)
from zeroda_reflex.utils.database import (
    meal_get_menus, meal_save_menu, meal_delete_menu,
    meal_analyze_waste, meal_get_analysis_summary,
    meal_get_menu_ranking, meal_get_weekday_pattern,
    meal_get_cost_savings, meal_get_settlement,
    meal_get_esg, school_filter_collections,
    get_school_standard, WASTE_GRADE_TABLE,
    meal_get_menus_by_month, meal_get_collected_dates,
    meal_get_school_student_count,
    meal_get_monthly_trend, meal_get_seasonal_compare,
)


# ── 메뉴 목록 ──
MEAL_TABS = ["식단등록", "스마트잔반분석", "AI잔반분석", "수거현황", "정산확인", "ESG보고서"]


class MealState(AuthState):
    """급식담당자 전체 상태"""

    # ══════════════════════════════
    #  공통
    # ══════════════════════════════
    active_tab: str = "식단등록"
    selected_year: str = str(datetime.now().year)
    selected_month: str = str(datetime.now().month).zfill(2)
    site_name: str = ""  # 담당 급식소 (= user_schools의 첫번째)
    msg: str = ""
    msg_ok: bool = False

    # ══════════════════════════════
    #  탭1: 식단등록
    # ══════════════════════════════
    menu_rows: list[dict] = []
    # 입력 폼
    mf_date: str = ""
    mf_menu: str = ""
    mf_calories: str = "0"
    mf_servings: str = "0"

    # ══════════════════════════════
    #  탭2: 스마트잔반분석
    # ══════════════════════════════
    analysis_rows: list[dict] = []
    analysis_summary: dict = {}
    best_menus: list[dict] = []
    worst_menus: list[dict] = []
    weekday_pattern: list[dict] = []

    # ══════════════════════════════
    #  탭3: AI잔반분석
    # ══════════════════════════════
    cost_data: dict = {}
    # AI 분석 결과 상태
    ai_api_key: str = ""              # 사용자 입력 API 키
    ai_comprehensive_result: str = "" # 종합분석 결과 (마크다운)
    ai_recommend_result: list[dict] = []  # 추천식단 결과
    ai_loading: bool = False          # 분석 진행중 플래그
    ai_error: str = ""                # 에러 메시지

    # ══════════════════════════════
    #  탭4: 수거현황
    # ══════════════════════════════
    collection_rows: list[dict] = []

    # ══════════════════════════════
    #  탭5: 정산확인
    # ══════════════════════════════
    settle_data: dict = {}
    settle_items: list[dict] = []

    # ══════════════════════════════
    #  탭6: ESG보고서
    # ══════════════════════════════
    esg_data: dict = {}

    # ══════════════════════════════
    #  수정4: 학교급식법 공식기준
    # ══════════════════════════════
    school_standard: dict = {}
    waste_grade_table: list[dict] = []

    # ══════════════════════════════
    #  수정1: 달력형 식단 보기
    # ══════════════════════════════
    calendar_month: str = ""
    # 평탄화 달력: 각 셀 dict에 row_end="1" 이면 주(週) 마지막 셀
    calendar_cells: list[dict] = []

    # ══════════════════════════════
    #  수정3: 잔반 트렌드
    # ══════════════════════════════
    trend_subtab: str = "weekday"
    monthly_trend: list[dict] = []
    seasonal_compare: list[dict] = []

    # ══════════════════════════════
    #  Computed vars
    # ══════════════════════════════

    @rx.var
    def has_msg(self) -> bool:
        return len(self.msg) > 0

    @rx.var
    def year_month(self) -> str:
        return f"{self.selected_year}-{self.selected_month.zfill(2)}"

    @rx.var
    def has_menus(self) -> bool:
        return len(self.menu_rows) > 0

    @rx.var
    def menu_count(self) -> int:
        return len(self.menu_rows)

    @rx.var
    def has_analysis(self) -> bool:
        return len(self.analysis_rows) > 0

    @rx.var
    def has_best(self) -> bool:
        return len(self.best_menus) > 0

    @rx.var
    def has_worst(self) -> bool:
        return len(self.worst_menus) > 0

    @rx.var
    def has_weekday(self) -> bool:
        return len(self.weekday_pattern) > 0

    @rx.var
    def has_cost(self) -> bool:
        return bool(self.cost_data)

    @rx.var
    def has_ai_result(self) -> bool:
        return len(self.ai_comprehensive_result) > 0

    @rx.var
    def has_ai_recommend(self) -> bool:
        return len(self.ai_recommend_result) > 0

    @rx.var
    def has_ai_error(self) -> bool:
        return len(self.ai_error) > 0

    @rx.var
    def has_api_key(self) -> bool:
        """API 키 존재 여부 (사용자 입력 또는 환경변수)"""
        if self.ai_api_key:
            return True
        import os
        return bool(os.environ.get("ANTHROPIC_API_KEY", ""))

    @rx.var
    def has_collection(self) -> bool:
        return len(self.collection_rows) > 0

    @rx.var
    def has_settle(self) -> bool:
        return bool(self.settle_data) and safe_int(self.settle_data.get("row_count", 0)) > 0

    @rx.var
    def has_settle_items(self) -> bool:
        return len(self.settle_items) > 0

    @rx.var
    def has_esg(self) -> bool:
        return bool(self.esg_data) and safe_int(self.esg_data.get("count", 0)) > 0

    @rx.var
    def standard_compliance(self) -> str:
        """1인당 평균 잔반량 / 245g × 100 — 기준 대비 비율(%)"""
        avg_pp = float(self.analysis_summary.get("avg_waste_pp", "0") or "0")
        if avg_pp <= 0:
            return "0"
        return str(round(avg_pp / 245 * 100, 1))

    @rx.var
    def has_standard(self) -> bool:
        return bool(self.school_standard)

    @rx.var
    def has_calendar(self) -> bool:
        return len(self.calendar_cells) > 0

    @rx.var
    def has_monthly_trend(self) -> bool:
        return len(self.monthly_trend) > 0

    @rx.var
    def has_seasonal(self) -> bool:
        return len(self.seasonal_compare) > 0

    # ══════════════════════════════
    #  초기화
    # ══════════════════════════════

    def on_meal_load(self):
        if not self.is_authenticated:
            return rx.redirect("/")
        if self.user_role not in ("meal_manager", "school_nutrition"):
            return rx.redirect("/")
        # site_name = 첫 번째 학교
        schools = [s.strip() for s in self.user_schools.split(",") if s.strip()]
        if schools:
            self.site_name = schools[0]
        self.load_school_standard()
        self.load_tab_data()

    # ══════════════════════════════
    #  탭 전환 / 필터
    # ══════════════════════════════

    def set_active_tab(self, tab: str):
        self.active_tab = tab
        self.msg = ""
        self.load_tab_data()

    def set_selected_year(self, y: str):
        self.selected_year = y
        self.load_tab_data()

    def set_selected_month(self, m: str):
        self.selected_month = m.zfill(2)
        self.load_tab_data()

    def load_tab_data(self):
        if not self.site_name:
            return
        tab = self.active_tab
        if tab == "식단등록":
            self.load_menus()
            self.load_meal_calendar()
        elif tab == "스마트잔반분석":
            self.load_analysis()
        elif tab == "AI잔반분석":
            self.load_ai_analysis()
        elif tab == "수거현황":
            self.load_collections()
        elif tab == "정산확인":
            self.load_settlement()
        elif tab == "ESG보고서":
            self.load_esg()

    # ══════════════════════════════
    #  탭1: 식단등록
    # ══════════════════════════════

    def load_menus(self):
        ym = f"{self.selected_year}-{self.selected_month.zfill(2)}"
        self.menu_rows = meal_get_menus(self.site_name, ym)

    def set_mf_date(self, v: str):
        self.mf_date = v

    def set_mf_menu(self, v: str):
        self.mf_menu = v

    def set_mf_calories(self, v: str):
        self.mf_calories = v

    def set_mf_servings(self, v: str):
        self.mf_servings = v

    def save_menu(self):
        if not self.mf_date or not self.mf_menu:
            self.msg = "날짜와 메뉴를 입력하세요."
            self.msg_ok = False
            return
        import json
        menu_list = [m.strip() for m in self.mf_menu.split(",") if m.strip()]
        menu_json = json.dumps(menu_list, ensure_ascii=False)
        try:
            cal = int(self.mf_calories)
        except (ValueError, TypeError):
            cal = 0
        try:
            srv = int(self.mf_servings)
        except (ValueError, TypeError):
            srv = 0

        # ── 칼로리 범위 검증 (0~5000kcal) ──
        if cal < 0 or cal > 5000:
            self.msg = "칼로리는 0~5000 범위여야 합니다."
            self.msg_ok = False
            return

        # ── 인원수 범위 검증 (0~10000명) ──
        if srv < 0 or srv > 10000:
            self.msg = "인원수는 0~10,000 범위여야 합니다."
            self.msg_ok = False
            return

        ok = meal_save_menu(
            self.site_name, self.mf_date, "중식",
            menu_json, cal, srv,
        )
        if ok:
            self.msg = f"{self.mf_date} 식단 저장 완료"
            self.msg_ok = True
            self.mf_date = ""
            self.mf_menu = ""
            self.mf_calories = "0"
            self.mf_servings = "0"
            self.load_menus()
        else:
            self.msg = "저장 실패"
            self.msg_ok = False

    def delete_menu(self, meal_date: str):
        ok = meal_delete_menu(self.site_name, meal_date)
        if ok:
            self.msg = f"{meal_date} 식단 삭제"
            self.msg_ok = True
            self.load_menus()
        else:
            self.msg = "삭제 실패"
            self.msg_ok = False

    # ──────────────────────────────
    #  식단 CSV/Excel 일괄 업로드
    # ──────────────────────────────
    upload_progress: int = 0  # 업로드 진행률 (0~100)

    async def handle_meal_upload(self, files: list[rx.UploadFile]):
        """CSV 또는 Excel 파일로 식단 일괄 등록.
        필수 컬럼: 날짜, 메뉴  |  선택 컬럼: 칼로리, 배식인원
        """
        import json as _json
        if not files:
            self.msg = "파일을 선택하세요."
            self.msg_ok = False
            return

        file = files[0]
        fname = file.filename.lower() if file.filename else ""

        # ── 파일 읽기 ──
        try:
            raw = await file.read()
        except Exception as e:
            logger.warning(f"파일 읽기 실패: {e}")
            self.msg = "파일 읽기 실패"
            self.msg_ok = False
            return

        rows: list[dict] = []

        # ── CSV 파싱 ──
        if fname.endswith(".csv"):
            import csv, io
            try:
                text = raw.decode("utf-8-sig")  # BOM 처리
            except UnicodeDecodeError:
                text = raw.decode("euc-kr", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            for r in reader:
                rows.append(dict(r))

        # ── Excel 파싱 (.xlsx / .xls) ──
        elif fname.endswith((".xlsx", ".xls")):
            try:
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    d = {}
                    for i, h in enumerate(headers):
                        val = row[i] if i < len(row) else ""
                        d[h] = str(val) if val is not None else ""
                    rows.append(d)
                wb.close()
            except Exception as e:
                logger.warning(f"Excel 파싱 실패: {e}")
                self.msg = "Excel 파일 형식 오류"
                self.msg_ok = False
                return
        else:
            self.msg = "CSV 또는 Excel(.xlsx) 파일만 지원됩니다."
            self.msg_ok = False
            return

        if not rows:
            self.msg = "파일에 데이터가 없습니다."
            self.msg_ok = False
            return

        # ── 컬럼명 매핑 (한글/영문 유연 매핑) ──
        col_map = {
            "날짜": "date", "date": "date", "meal_date": "date",
            "메뉴": "menu", "menu": "menu", "menu_items": "menu",
            "칼로리": "cal", "calories": "cal", "kcal": "cal",
            "배식인원": "srv", "servings": "srv", "인원": "srv", "인원수": "srv",
        }

        def _find_col(row_keys: list[str], target: str) -> str:
            """row의 key 중 target에 매핑되는 컬럼명 찾기"""
            for k in row_keys:
                if col_map.get(k.strip().lower()) == target:
                    return k
            return ""

        sample_keys = list(rows[0].keys())
        date_col = _find_col(sample_keys, "date")
        menu_col = _find_col(sample_keys, "menu")
        cal_col = _find_col(sample_keys, "cal")
        srv_col = _find_col(sample_keys, "srv")

        if not date_col or not menu_col:
            self.msg = "필수 컬럼 누락: '날짜'와 '메뉴' 컬럼이 필요합니다."
            self.msg_ok = False
            return

        # ── 행별 저장 ──
        ok_count = 0
        fail_count = 0
        total = len(rows)

        for idx, r in enumerate(rows):
            date_val = str(r.get(date_col, "")).strip()
            menu_val = str(r.get(menu_col, "")).strip()

            # 빈 행 스킵
            if not date_val or not menu_val:
                fail_count += 1
                continue

            # 날짜 형식 정리 (YYYY-MM-DD)
            date_val = date_val[:10]

            # 칼로리/인원 파싱
            cal = safe_int(r.get(cal_col, "0")) if cal_col else 0
            srv = safe_int(r.get(srv_col, "0")) if srv_col else 0

            # 범위 검증
            cal = max(0, min(cal, 5000))
            srv = max(0, min(srv, 10000))

            # 메뉴를 JSON 배열로 변환
            menu_list = [m.strip() for m in menu_val.split(",") if m.strip()]
            menu_json = _json.dumps(menu_list, ensure_ascii=False)

            ok = meal_save_menu(
                self.site_name, date_val, "중식",
                menu_json, cal, srv,
            )
            if ok:
                ok_count += 1
            else:
                fail_count += 1

            # 진행률 업데이트
            self.upload_progress = int((idx + 1) / total * 100)
            yield  # Reflex에서 중간 상태 업데이트

        # ── 결과 메시지 ──
        self.upload_progress = 0
        if fail_count == 0:
            self.msg = f"총 {ok_count}건 식단 일괄 등록 완료"
            self.msg_ok = True
        else:
            self.msg = f"완료: {ok_count}건 성공, {fail_count}건 실패"
            self.msg_ok = ok_count > 0

        # 목록 새로고침
        self.load_menus()

    # ══════════════════════════════
    #  수정2: NEIS 엑셀 업로드 파싱
    # ══════════════════════════════

    NUTRITION_KEYS = ["에너지(kcal)", "탄수화물(g)", "단백질(g)", "지방(g)",
                       "비타민A(μg RE)", "티아민(mg)", "리보플라빈(mg)", "비타민C(mg)", "칼슘(mg)"]

    @staticmethod
    def _parse_menu_items(raw: str) -> list:
        """NEIS 요리명 파싱: <br/>/\n 분리 + 알러지 코드 제거"""
        import re as _re
        if not raw:
            return []
        items = _re.split(r"<br\s*/?>\s*|\n", str(raw))
        result = []
        for item in items:
            item = item.strip()
            if not item:
                continue
            # 알러지 코드 제거 (예: "김치찌개 (5.9.)" → "김치찌개")
            item = _re.sub(r"\s*[\(\[]\s*[\d\s\.]+[\)\]]", "", item).strip()
            if item:
                result.append(item)
        return result

    @staticmethod
    def _parse_nutrition(raw: str) -> dict:
        """NEIS 영양소 문자열 → 9항목 dict"""
        import re as _re
        result: dict = {}
        if not raw:
            return result
        # "에너지(kcal) : 650" 형태 또는 숫자만 있는 경우 처리
        parts = _re.split(r"[,\n]", str(raw))
        for part in parts:
            part = part.strip()
            if ":" in part:
                k, _, v = part.partition(":")
                result[k.strip()] = v.strip()
        return result

    @staticmethod
    def _parse_calories(raw: str) -> int:
        """문자열에서 숫자(칼로리) 추출"""
        import re as _re
        m = _re.search(r"[\d]+\.?[\d]*", str(raw or ""))
        if m:
            try:
                return int(float(m.group()))
            except (ValueError, TypeError):
                return 0
        return 0

    async def handle_neis_upload(self, files: list[rx.UploadFile]):
        """NEIS 엑셀 파일 업로드 파싱 — 급식일자/요리명 컬럼 자동 감지"""
        import json as _json
        import re as _re
        if not files:
            self.msg = "파일을 선택하세요."
            self.msg_ok = False
            return
        file = files[0]
        fname = file.filename.lower() if file.filename else ""
        if not fname.endswith((".xlsx", ".xls")):
            self.msg = "NEIS 엑셀 파일(.xlsx)만 지원합니다."
            self.msg_ok = False
            return
        try:
            raw = await file.read()
        except Exception as e:
            logger.warning(f"NEIS 파일 읽기 실패: {e}")
            self.msg = "파일 읽기 실패"
            self.msg_ok = False
            return

        try:
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            # 헤더 행 찾기 (급식일자, 요리명 컬럼 검색)
            header_row_idx = None
            headers: list = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                cells = [str(c or "").strip() for c in row]
                if any("급식일자" in c or "날짜" in c for c in cells) and \
                   any("요리명" in c or "식단" in c or "메뉴" in c for c in cells):
                    header_row_idx = row_idx
                    headers = cells
                    break
            if header_row_idx is None:
                self.msg = "NEIS 형식 헤더를 찾지 못했습니다. '급식일자', '요리명' 컬럼이 필요합니다."
                self.msg_ok = False
                wb.close()
                return

            # 컬럼 인덱스 매핑
            def _col_idx(keywords: list) -> int:
                for kw in keywords:
                    for i, h in enumerate(headers):
                        if kw in h:
                            return i
                return -1

            date_col = _col_idx(["급식일자", "날짜"])
            menu_col = _col_idx(["요리명", "식단", "메뉴"])
            cal_col = _col_idx(["에너지", "열량", "칼로리", "kcal"])
            nutrition_cols: dict = {}
            for nk in self.NUTRITION_KEYS:
                idx = _col_idx([nk.split("(")[0]])
                if idx >= 0:
                    nutrition_cols[nk] = idx

            if date_col < 0 or menu_col < 0:
                self.msg = "급식일자 또는 요리명 컬럼을 찾지 못했습니다."
                self.msg_ok = False
                wb.close()
                return

            # 기본 학생수 조회
            default_srv = meal_get_school_student_count(self.site_name)

            ok_count = 0
            fail_count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                cells = list(row)
                date_raw = str(cells[date_col] if date_col < len(cells) else "").strip()
                menu_raw = str(cells[menu_col] if menu_col < len(cells) else "").strip()
                if not date_raw or not menu_raw or date_raw in ("None", ""):
                    continue
                # 날짜 정규화 (YYYYMMDD / YYYY-MM-DD / YYYY.MM.DD)
                date_clean = _re.sub(r"[.\-/]", "", date_raw)
                if len(date_clean) == 8 and date_clean.isdigit():
                    date_str = f"{date_clean[:4]}-{date_clean[4:6]}-{date_clean[6:8]}"
                else:
                    # 이미 YYYY-MM-DD 형태이거나 파싱 불가
                    date_str = date_raw[:10]
                # 메뉴 파싱
                menu_list = self._parse_menu_items(menu_raw)
                if not menu_list:
                    fail_count += 1
                    continue
                menu_json = _json.dumps(menu_list, ensure_ascii=False)
                # 칼로리
                cal_raw = str(cells[cal_col] if cal_col >= 0 and cal_col < len(cells) else "").strip()
                cal = self._parse_calories(cal_raw)
                cal = max(0, min(cal, 5000))
                # 영양정보
                nutrition: dict = {}
                for nk, ni in nutrition_cols.items():
                    if ni < len(cells):
                        nutrition[nk] = str(cells[ni] or "")
                nutrition_json = _json.dumps(nutrition, ensure_ascii=False)
                # 배식인원 (기본값 적용)
                srv = default_srv if default_srv > 0 else 0
                ok = meal_save_menu(
                    self.site_name, date_str, "중식",
                    menu_json, cal, srv, nutrition_json,
                )
                if ok:
                    ok_count += 1
                else:
                    fail_count += 1
            wb.close()
        except Exception as e:
            logger.warning(f"NEIS 파싱 오류: {e}")
            self.msg = f"NEIS 파싱 오류: {e}"
            self.msg_ok = False
            return

        if fail_count == 0:
            self.msg = f"NEIS 식단 {ok_count}건 등록 완료"
            self.msg_ok = True
        else:
            self.msg = f"완료: {ok_count}건 성공, {fail_count}건 실패"
            self.msg_ok = ok_count > 0
        self.load_menus()
        self.load_meal_calendar()

    # ══════════════════════════════
    #  탭2: 스마트잔반분석
    # ══════════════════════════════

    def load_analysis(self):
        ym = f"{self.selected_year}-{self.selected_month.zfill(2)}"
        self.analysis_rows = meal_analyze_waste(self.site_name, ym)
        # 차트용 숫자 필드 추가 (Phase 9)
        for r in self.analysis_rows:
            r["waste_num"] = float(r.get("waste_kg", 0) or 0)
            r["pp_num"] = float(r.get("waste_per_person", 0) or 0)
        self.analysis_summary = meal_get_analysis_summary(self.analysis_rows)
        best, worst = meal_get_menu_ranking(self.analysis_rows)
        self.best_menus = best
        self.worst_menus = worst
        wd = meal_get_weekday_pattern(self.analysis_rows)
        for r in wd:
            r["avg_num"] = float(r.get("avg_kg", 0) or 0)
            r["pp_num"] = float(r.get("avg_pp", 0) or 0)
        self.weekday_pattern = wd
        # 수정3: 트렌드 로드
        self.load_trends()

    # ══════════════════════════════
    #  수정4: 학교급식법 공식기준 로드
    # ══════════════════════════════

    def load_school_standard(self):
        from zeroda_reflex.utils.database import _detect_school_level
        school_type = _detect_school_level(self.site_name)
        self.school_standard = get_school_standard(school_type)
        self.waste_grade_table = list(WASTE_GRADE_TABLE)

    # ══════════════════════════════
    #  수정1: 달력형 식단 보기
    # ══════════════════════════════

    def load_meal_calendar(self):
        import calendar as _cal
        ym = f"{self.selected_year}-{self.selected_month.zfill(2)}"
        self.calendar_month = ym
        try:
            y = int(self.selected_year)
            m = int(self.selected_month)
        except (ValueError, TypeError):
            return
        # 식단 날짜 맵
        menu_list = meal_get_menus_by_month(self.site_name, ym)
        menu_map = {item["date"]: item["summary"] for item in menu_list}
        # 수거 날짜 집합
        collected = set(meal_get_collected_dates(self.site_name, ym))
        # 달력 셀 평탄화 (각 셀에 col 0~6, row_end="1"이면 행 마지막)
        first_wd, num_days = _cal.monthrange(y, m)
        cells: list = []
        col = 0
        # 앞쪽 빈 셀
        for _ in range(first_wd):
            cells.append({"day": "", "date": "", "has_menu": "0",
                          "has_collect": "0", "menu_summary": "",
                          "row_end": "1" if col == 6 else "0"})
            col += 1
        for day in range(1, num_days + 1):
            date_str = f"{y}-{str(m).zfill(2)}-{str(day).zfill(2)}"
            is_row_end = "1" if col == 6 else "0"
            cells.append({
                "day": str(day),
                "date": date_str,
                "has_menu": "1" if date_str in menu_map else "0",
                "has_collect": "1" if date_str in collected else "0",
                "menu_summary": menu_map.get(date_str, ""),
                "row_end": is_row_end,
            })
            col = (col + 1) % 7
        # 뒤쪽 빈 셀 패딩
        while col != 0:
            is_row_end = "1" if col == 6 else "0"
            cells.append({"day": "", "date": "", "has_menu": "0",
                          "has_collect": "0", "menu_summary": "", "row_end": is_row_end})
            col = (col + 1) % 7
        self.calendar_cells = cells

    # ══════════════════════════════
    #  수정3: 잔반 트렌드 로드
    # ══════════════════════════════

    def set_trend_subtab(self, tab: str):
        self.trend_subtab = tab

    def load_trends(self):
        if not self.site_name:
            return
        try:
            y = int(self.selected_year)
        except (ValueError, TypeError):
            return
        self.monthly_trend = meal_get_monthly_trend(self.site_name, y)
        self.seasonal_compare = meal_get_seasonal_compare(self.site_name, y)

    # ══════════════════════════════
    #  탭3: AI잔반분석
    # ══════════════════════════════

    def load_ai_analysis(self):
        # 분석 데이터 + 비용
        if not self.analysis_rows:
            ym = f"{self.selected_year}-{self.selected_month.zfill(2)}"
            self.analysis_rows = meal_analyze_waste(self.site_name, ym)
            self.analysis_summary = meal_get_analysis_summary(self.analysis_rows)
            best, worst = meal_get_menu_ranking(self.analysis_rows)
            self.best_menus = best
            self.worst_menus = worst
            self.weekday_pattern = meal_get_weekday_pattern(self.analysis_rows)
        self.cost_data = meal_get_cost_savings(self.site_name, self.analysis_rows)

    def set_ai_api_key(self, key: str):
        """사용자 입력 API 키 저장"""
        self.ai_api_key = key.strip()
        self.ai_error = ""

    async def run_ai_comprehensive(self):
        """AI 종합 잔반분석 실행 (비동기)"""
        from zeroda_reflex.utils.ai_service import (
            build_comprehensive_prompt, call_claude_api,
        )
        if not self.analysis_rows:
            self.ai_error = "분석할 데이터가 없습니다. 먼저 잔반 데이터를 로드하세요."
            return

        self.ai_loading = True
        self.ai_error = ""
        self.ai_comprehensive_result = ""
        yield  # UI 업데이트 (로딩 표시)

        prompt = build_comprehensive_prompt(
            self.site_name, self.analysis_rows, self.cost_data,
        )
        result = call_claude_api(prompt, self.ai_api_key)

        self.ai_loading = False
        if result.startswith("[ERROR]"):
            self.ai_error = result.replace("[ERROR] ", "")
        else:
            self.ai_comprehensive_result = result

    async def run_ai_recommend(self):
        """AI 추천식단 생성 (비동기)"""
        from zeroda_reflex.utils.ai_service import (
            build_recommend_prompt, call_claude_api, parse_recommend_json,
        )
        if not self.analysis_rows:
            self.ai_error = "분석할 데이터가 없습니다."
            return

        self.ai_loading = True
        self.ai_error = ""
        self.ai_recommend_result = []
        yield  # UI 업데이트

        prompt = build_recommend_prompt(self.site_name, self.analysis_rows)
        result = call_claude_api(prompt, self.ai_api_key)

        self.ai_loading = False
        if result.startswith("[ERROR]"):
            self.ai_error = result.replace("[ERROR] ", "")
        else:
            parsed = parse_recommend_json(result)
            if parsed:
                # dict의 모든 값을 str로 변환 (Reflex 직렬화)
                self.ai_recommend_result = [
                    {
                        "day": str(r.get("day", "")),
                        "menu": ", ".join(r["menu"]) if isinstance(r.get("menu"), list) else str(r.get("menu", "")),
                        "expected_waste": str(r.get("expected_waste", "")),
                        "reason": str(r.get("reason", "")),
                    }
                    for r in parsed
                ]
            else:
                # JSON 파싱 실패 시 원문 텍스트를 종합분석 결과에 표시
                self.ai_comprehensive_result = result

    def download_ai_pdf(self):
        """AI 월말명세서 PDF 다운로드"""
        from zeroda_reflex.utils.pdf_export import build_ai_meal_statement_pdf
        try:
            y = int(self.selected_year)
            m = int(self.selected_month)
        except (ValueError, TypeError):
            return None
        if not self.analysis_rows:
            return None
        best, worst = meal_get_menu_ranking(self.analysis_rows)
        pdf_bytes = build_ai_meal_statement_pdf(
            self.site_name, y, m, self.analysis_rows,
            menu_ranking={"best": best, "worst": worst},
            ai_comment=self.ai_comprehensive_result[:2000] if self.ai_comprehensive_result else "",
            cost_savings=self.cost_data,
            weekday_pattern=self.weekday_pattern,
        )
        if pdf_bytes:
            return rx.download(
                data=pdf_bytes,
                filename=f"AI월말명세서_{self.site_name}_{y}-{str(m).zfill(2)}.pdf",
            )
        return None

    # ══════════════════════════════
    #  탭4: 수거현황
    # ══════════════════════════════

    def load_collections(self):
        try:
            y = int(self.selected_year)
            m = int(self.selected_month)
        except (ValueError, TypeError):
            return
        self.collection_rows = school_filter_collections(self.site_name, y, m)

    # ══════════════════════════════
    #  탭5: 정산확인
    # ══════════════════════════════

    def load_settlement(self):
        try:
            y = int(self.selected_year)
            m = int(self.selected_month)
        except (ValueError, TypeError):
            return
        result = meal_get_settlement(self.site_name, y, m)
        self.settle_items = result.pop("items", [])
        self.settle_data = result

    # ══════════════════════════════
    #  탭6: ESG보고서
    # ══════════════════════════════

    def load_esg(self):
        try:
            y = int(self.selected_year)
        except (ValueError, TypeError):
            return
        self.esg_data = meal_get_esg(self.site_name, y)

    # ══════════════════════════════
    #  PDF 다운로드 핸들러
    # ══════════════════════════════

    def download_smart_pdf(self):
        """스마트월말명세서 PDF 다운로드"""
        from zeroda_reflex.utils.pdf_export import build_meal_statement_pdf
        try:
            y = int(self.selected_year)
            m = int(self.selected_month)
        except (ValueError, TypeError):
            return None
        if not self.analysis_rows:
            return None
        best, worst = meal_get_menu_ranking(self.analysis_rows)
        ranking = {"best": best, "worst": worst}
        pdf_bytes = build_meal_statement_pdf(
            self.site_name, y, m, self.analysis_rows, ranking,
        )
        if pdf_bytes:
            return rx.download(
                data=pdf_bytes,
                filename=f"스마트월말명세서_{self.site_name}_{y}-{str(m).zfill(2)}.pdf",
            )
        return None

    def download_esg_pdf(self):
        """급식 ESG 보고서 PDF 다운로드"""
        from zeroda_reflex.utils.pdf_export import build_school_esg_pdf
        from zeroda_reflex.utils.database import school_filter_collections
        try:
            y = int(self.selected_year)
        except (ValueError, TypeError):
            return None
        rows = school_filter_collections(self.site_name, y, 0)
        if not rows:
            return None
        month_label = f"{y}년 전체"
        pdf_bytes = build_school_esg_pdf(self.site_name, y, month_label, rows)
        if pdf_bytes:
            return rx.download(
                data=pdf_bytes,
                filename=f"ESG보고서_{self.site_name}_{y}.pdf",
            )
        return None

    # ══════════════════════════════
    #  Excel 다운로드 핸들러
    # ══════════════════════════════

    def download_meal_excel(self):
        """식단 및 잔반 분석 데이터 Excel 다운로드 (급식담당자)"""
        # menu_rows + analysis_rows + cost_data 통합 다운로드
        from zeroda_reflex.utils.excel_export import export_meal_data
        meal_data = {
            "menus": self.menu_rows,
            "analysis": self.analysis_rows,
            "best_menus": self.best_menus,
            "worst_menus": self.worst_menus,
            "weekday_pattern": self.weekday_pattern,
            "cost_data": self.cost_data,
        }
        if not any(meal_data.values()):
            return None
        xlsx = export_meal_data(meal_data, self.site_name, self.selected_year, self.selected_month)
        if xlsx:
            return rx.download(
                data=xlsx,
                filename=f"식단분석_{self.site_name}_{self.selected_year}-{self.selected_month.zfill(2)}.xlsx"
            )
        return None
