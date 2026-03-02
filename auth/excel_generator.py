# zeroda_platform/services/excel_generator.py
# ==========================================
# 엑셀 생성 (정산, 수거일보, 홈택스 양식)
# ==========================================

import io
from datetime import datetime
from config.settings import COMPANY_NAME, EXCEL_PASSWORD


def _wb():
    """openpyxl 워크북 생성"""
    try:
        import openpyxl
        return openpyxl.Workbook()
    except ImportError:
        return None


def _style_header(ws, row, cols, fill_color='1a73e8'):
    """헤더 행 스타일"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        fill = PatternFill('solid', fgColor=fill_color)
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    except Exception:
        pass


def _border_all(ws, min_row, min_col, max_row, max_col):
    """전체 테두리"""
    try:
        from openpyxl.styles import Border, Side
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=min_row, min_col=min_col,
                                max_row=max_row, max_col=max_col):
            for cell in row:
                cell.border = border
    except Exception:
        pass


# ──────────────────────────────────────────
# 월별 정산 엑셀
# ──────────────────────────────────────────

def generate_settlement_excel(vendor, year, month, settlement_data):
    """
    월별 정산 엑셀 생성
    반환: bytes
    """
    wb = _wb()
    if not wb:
        return None

    ws = wb.active
    ws.title = f"{month}월 정산"

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{year}년 {month}월 수거 정산서 ({vendor})"

    try:
        from openpyxl.styles import Font, Alignment
        ws['A1'].font      = Font(bold=True, size=14, name='맑은 고딕')
        ws['A1'].alignment = Alignment(horizontal='center')
    except Exception:
        pass

    # 헤더
    headers = ['학교명', '수거량(kg)', '단가(원/kg)', '공급가액', '부가세(10%)', '합계금액']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, len(headers))

    # 데이터
    total_kg  = 0
    total_amt = 0
    for row_i, r in enumerate(settlement_data, 3):
        kg  = float(r.get('수거량', 0) or 0)
        amt = int(r.get('금액', 0)    or 0)
        sup = int(amt / 1.1)
        vat = amt - sup
        total_kg  += kg
        total_amt += amt

        ws.cell(row=row_i, column=1, value=r.get('학교명', ''))
        ws.cell(row=row_i, column=2, value=round(kg, 1))
        ws.cell(row=row_i, column=3, value=int(r.get('단가', 0) or 0))
        ws.cell(row=row_i, column=4, value=sup)
        ws.cell(row=row_i, column=5, value=vat)
        ws.cell(row=row_i, column=6, value=amt)

    # 합계 행
    sum_row = len(settlement_data) + 3
    ws.cell(row=sum_row, column=1, value='합 계')
    ws.cell(row=sum_row, column=2, value=round(total_kg, 1))
    ws.cell(row=sum_row, column=4, value=int(total_amt / 1.1))
    ws.cell(row=sum_row, column=5, value=total_amt - int(total_amt / 1.1))
    ws.cell(row=sum_row, column=6, value=total_amt)

    try:
        from openpyxl.styles import PatternFill, Font
        fill = PatternFill('solid', fgColor='e8f4fd')
        bold = Font(bold=True, name='맑은 고딕')
        for col in range(1, 7):
            ws.cell(row=sum_row, column=col).fill = fill
            ws.cell(row=sum_row, column=col).font = bold
    except Exception:
        pass

    _border_all(ws, 2, 1, sum_row, 6)

    # 열 너비
    widths = [20, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ──────────────────────────────────────────
# 홈택스 전자계산서 양식
# ──────────────────────────────────────────

def generate_hometax_excel(vendor, year, month,
                           settlement_data, vendor_info=None):
    """
    홈택스 전자세금계산서 일괄등록 양식 생성
    반환: bytes
    """
    wb = _wb()
    if not wb:
        return None

    ws = wb.active
    ws.title = '전자세금계산서'

    # 홈택스 양식 헤더 (국세청 일괄등록 기준)
    headers = [
        '작성일자', '공급자사업자번호', '공급자상호',
        '공급받는자사업자번호', '공급받는자상호',
        '공급가액', '세액', '합계금액', '품목', '규격', '수량', '단가', '비고'
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers), fill_color='34a853')

    issue_date = f"{year}{month:02d}01"

    vendor_biz_no = ''
    vendor_name   = vendor
    if vendor_info:
        vendor_biz_no = vendor_info.get('biz_no', '')
        vendor_name   = vendor_info.get('biz_name', vendor)

    for row_i, r in enumerate(settlement_data, 2):
        kg  = float(r.get('수거량', 0) or 0)
        amt = int(r.get('금액', 0) or 0)
        sup = int(amt / 1.1)
        vat = amt - sup

        ws.cell(row=row_i, column=1,  value=issue_date)
        ws.cell(row=row_i, column=2,  value=vendor_biz_no)
        ws.cell(row=row_i, column=3,  value=vendor_name)
        ws.cell(row=row_i, column=4,  value=r.get('학교_사업자번호', ''))
        ws.cell(row=row_i, column=5,  value=r.get('학교명', ''))
        ws.cell(row=row_i, column=6,  value=sup)
        ws.cell(row=row_i, column=7,  value=vat)
        ws.cell(row=row_i, column=8,  value=amt)
        ws.cell(row=row_i, column=9,  value='음식물폐기물 수거')
        ws.cell(row=row_i, column=10, value='kg')
        ws.cell(row=row_i, column=11, value=round(kg, 1))
        ws.cell(row=row_i, column=12, value=int(r.get('단가', 0) or 0))
        ws.cell(row=row_i, column=13, value=f"{year}년 {month}월")

    _border_all(ws, 1, 1, len(settlement_data) + 1, len(headers))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ──────────────────────────────────────────
# 수거일보 엑셀
# ──────────────────────────────────────────

def generate_collection_excel(school_name, year, month, collection_data):
    """
    학교별 월간 수거일보 엑셀
    반환: bytes
    """
    wb = _wb()
    if not wb:
        return None

    ws = wb.active
    ws.title = f"{school_name}_{month}월"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"{year}년 {month}월 수거일보 - {school_name}"
    try:
        from openpyxl.styles import Font, Alignment
        ws['A1'].font      = Font(bold=True, size=13, name='맑은 고딕')
        ws['A1'].alignment = Alignment(horizontal='center')
    except Exception:
        pass

    headers = ['날짜', '수거량(kg)', '단가(원)', '금액(원)', '수거업체', '수거기사']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    _style_header(ws, 2, len(headers))

    total_kg = 0
    for row_i, r in enumerate(
            sorted(collection_data, key=lambda x: x.get('날짜', '')), 3):
        kg    = float(r.get('음식물(kg)', 0) or 0)
        price = float(r.get('단가(원)', 162) or 162)
        amt   = int(kg * price)
        total_kg += kg

        ws.cell(row=row_i, column=1, value=r.get('날짜', ''))
        ws.cell(row=row_i, column=2, value=round(kg, 1))
        ws.cell(row=row_i, column=3, value=int(price))
        ws.cell(row=row_i, column=4, value=amt)
        ws.cell(row=row_i, column=5, value=r.get('수거업체', ''))
        ws.cell(row=row_i, column=6, value=r.get('수거기사', ''))

    sum_row = len(collection_data) + 3
    ws.cell(row=sum_row, column=1, value='합 계')
    ws.cell(row=sum_row, column=2, value=round(total_kg, 1))

    _border_all(ws, 2, 1, sum_row, 6)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()