# zeroda_platform/services/excel_generator.py
# 수거 데이터 엑셀 생성 - 한글 완벽 지원
from io import BytesIO
from datetime import datetime


def generate_collection_excel(vendor: str, school_name: str,
                               year: int, month: int, rows: list) -> bytes:
    """수거 내역 엑셀 생성"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month}월"

    # ── 스타일 정의 ──────────────────────
    BLUE  = "1a73e8"
    GREEN = "34a853"
    GRAY  = "f5f5f5"

    header_font  = Font(name='맑은 고딕', bold=True, color="FFFFFF", size=11)
    normal_font  = Font(name='맑은 고딕', size=10)
    title_font   = Font(name='맑은 고딕', bold=True, size=14, color=BLUE)
    total_font   = Font(name='맑은 고딕', bold=True, size=11)

    blue_fill  = PatternFill("solid", fgColor=BLUE)
    green_fill = PatternFill("solid", fgColor=GREEN)
    gray_fill  = PatternFill("solid", fgColor=GRAY)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 제목 ────────────────────────────
    ws.merge_cells('A1:F1')
    ws['A1'] = f"거래명세서 - {school_name} ({year}년 {month}월)"
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f"공급자: {vendor}　　발행일: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(name='맑은 고딕', size=10, color="555555")
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 20

    # ── 헤더 ────────────────────────────
    headers = ['날짜', '학교명', '품목', '수거량(kg)', '단가(원)', '금액(원)']
    col_widths = [15, 20, 15, 14, 14, 16]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font      = header_font
        cell.fill      = blue_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    # ── 데이터 ──────────────────────────
    total_weight = 0.0
    total_amount = 0.0

    for row_idx, r in enumerate(rows, start=4):
        weight     = float(r.get('weight') or r.get('음식물(kg)') or 0)
        unit_price = float(r.get('unit_price') or r.get('단가(원)') or 0)
        amount     = weight * unit_price
        total_weight += weight
        total_amount += amount

        fill = gray_fill if row_idx % 2 == 0 else PatternFill()

        data = [
            r.get('collect_date', r.get('날짜', '')),
            r.get('school_name',  r.get('학교명', school_name)),
            r.get('item_type',    r.get('재활용방법', '')),
            weight,
            unit_price,
            amount,
        ]
        aligns = [center, left, center, right, right, right]

        for col_idx, (val, aln) in enumerate(zip(data, aligns), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = normal_font
            cell.alignment = aln
            cell.border    = border
            if fill.fgColor and fill.fgColor.rgb != '00000000':
                cell.fill = fill
            if col_idx in (4, 5, 6) and isinstance(val, float):
                cell.number_format = '#,##0.00' if col_idx == 4 else '#,##0'

        ws.row_dimensions[row_idx].height = 18

    # ── 합계 행 ──────────────────────────
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=2, value='합  계').font = total_font
    ws.cell(row=total_row, column=2).alignment = center
    ws.cell(row=total_row, column=4, value=total_weight).font = total_font
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=4).alignment = right
    ws.cell(row=total_row, column=6, value=total_amount).font = total_font
    ws.cell(row=total_row, column=6).number_format = '#,##0'
    ws.cell(row=total_row, column=6).alignment = right

    total_fill = PatternFill("solid", fgColor="FFF3CD")
    for col_idx in range(1, 7):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill   = total_fill
        cell.border = border
    ws.row_dimensions[total_row].height = 22

    # ── VAT 요약 ─────────────────────────
    vat   = total_amount * 0.1
    total_with_vat = total_amount + vat
    summary_row = total_row + 2

    summary = [
        ('공급가액', total_amount, BLUE),
        ('부가세(10%)', vat, '888888'),
        ('합계금액', total_with_vat, 'ea4335'),
    ]
    for i, (label, val, color) in enumerate(summary):
        r1 = summary_row + i
        ws.cell(row=r1, column=4, value=label).font = Font(name='맑은 고딕', bold=True, size=10, color=color)
        ws.cell(row=r1, column=4).alignment = right
        cell = ws.cell(row=r1, column=5, value=val)
        cell.font = Font(name='맑은 고딕', bold=True, size=10, color=color)
        cell.number_format = '#,##0'
        cell.alignment = right
        ws.merge_cells(start_row=r1, start_column=5, end_row=r1, end_column=6)

    # 저장
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_monthly_summary_excel(vendor: str, year: int, month: int, rows: list) -> bytes:
    """월별 학교별 정산 요약 엑셀"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month}월_정산요약"

    BLUE = "1a73e8"
    header_font = Font(name='맑은 고딕', bold=True, color="FFFFFF", size=11)
    normal_font = Font(name='맑은 고딕', size=10)
    title_font  = Font(name='맑은 고딕', bold=True, size=14, color=BLUE)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    ws.merge_cells('A1:E1')
    ws['A1'] = f"{year}년 {month}월 정산 요약 - {vendor}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['학교명', '수거횟수', '총수거량(kg)', '총공급가(원)', '부가세(원)']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = center
        cell.border = border
    ws.column_dimensions['A'].width = 22
    for col in 'BCDE':
        ws.column_dimensions[col].width = 16

    # 학교별 집계
    summary = defaultdict(lambda: {'count': 0, 'weight': 0.0, 'amount': 0.0})
    for r in rows:
        sn = r.get('school_name', r.get('학교명', ''))
        weight = float(r.get('weight') or r.get('음식물(kg)') or 0)
        price  = float(r.get('unit_price') or r.get('단가(원)') or 0)
        summary[sn]['count']  += 1
        summary[sn]['weight'] += weight
        summary[sn]['amount'] += weight * price

    for row_idx, (school, data) in enumerate(sorted(summary.items()), start=3):
        vat = data['amount'] * 0.1
        vals = [school, data['count'], data['weight'], data['amount'], vat]
        alns = [center, center, right, right, right]
        fmts = [None, None, '#,##0.0', '#,##0', '#,##0']
        fill = PatternFill("solid", fgColor="f5f5f5") if row_idx % 2 == 0 else PatternFill()

        for col_idx, (val, aln, fmt) in enumerate(zip(vals, alns, fmts), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = aln
            cell.border = border
            if fmt:
                cell.number_format = fmt
            if fill.fgColor and fill.fgColor.rgb != '00000000':
                cell.fill = fill

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
