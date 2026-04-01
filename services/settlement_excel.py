# services/settlement_excel.py
# 월말정산 엑셀 생성 + HTML 미리보기 — 수입내역(구분별 부가세) + 지출내역
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 부가세 규칙 ──
VAT_RULE = {'학교': False, '공공기관': False, '기업': True, '일반업장': True}

# ── 품목 → 단가 컬럼 매핑 ──
ITEM_PRICE_MAP = [
    ('음식물', 'price_food'),
    ('재활용', 'price_recycle'),
    ('일반폐기물', 'price_general'),
]

# ── 스타일 ──
_FONT_TITLE = Font(name='맑은 고딕', bold=True, size=14)
_FONT_H = Font(name='맑은 고딕', bold=True, size=11)
_FONT_N = Font(name='맑은 고딕', size=10)
_FONT_W = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
_FONT_BLUE = Font(name='맑은 고딕', size=10, color='0000FF')
_FONT_RED = Font(name='맑은 고딕', size=10, color='FF0000')
_FONT_RED_B = Font(name='맑은 고딕', bold=True, size=11, color='FF0000')
_FONT_BIG_B = Font(name='맑은 고딕', bold=True, size=12)

_FILL_HDR = PatternFill('solid', fgColor='4472C4')
_FILL_TOTAL = PatternFill('solid', fgColor='FFF2CC')
_CUST_FILL = {
    '학교':     PatternFill('solid', fgColor='DAEEF3'),
    '공공기관':  PatternFill('solid', fgColor='E2EFDA'),
    '기업':     PatternFill('solid', fgColor='FCE4D6'),
    '일반업장':  PatternFill('solid', fgColor='F2DCDB'),
}

_THIN = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin')
)
_C = Alignment(horizontal='center', vertical='center')
_R = Alignment(horizontal='right', vertical='center')


def _apply_border_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = _THIN


def generate_monthly_settlement_excel(
    vendor, year, month, customers_dict, collection_rows, expenses=None
):
    """
    월말정산 엑셀 생성

    Parameters
    ----------
    vendor : str
    year : int
    month : int
    customers_dict : dict
        load_customers_from_db(vendor) 결과 {name: {구분, price_food, ...}}
    collection_rows : list[dict]
        해당 월의 real_collection 전체 데이터
    expenses : list[dict] or None
        지출 항목 [{item, amount, date, memo}, ...]

    Returns
    -------
    bytes  (xlsx)
    """
    wb = Workbook()
    month_str = str(month).zfill(2)

    # ═══════════════════════════════════════
    # Sheet 1: 수입내역
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "수입내역"

    ws1.merge_cells('A1:I1')
    ws1['A1'] = f'( {month} )月  월말정산서 — 수입내역'
    ws1['A1'].font = _FONT_TITLE
    ws1['A1'].alignment = _C

    # 헤더
    headers = ['No', '구분', '거래처명', '품목', '누적수거량(kg)',
               '단가(원)', '공급가(원)', '부가세(원)', '정산금액(원)']
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=ci, value=h)
        c.font = _FONT_W
        c.fill = _FILL_HDR
        c.alignment = _C
        c.border = _THIN

    # ── 거래처별 누적 수거량 집계 (school_name + item_type) ──
    agg = {}
    for r in collection_rows:
        sn = r.get('school_name', '')
        it_raw = r.get('item_type', '음식물')
        w = float(r.get('weight', 0) or 0)
        key = (sn, it_raw)
        agg[key] = agg.get(key, 0) + w

    # ── 거래처 목록 기반으로 행 생성 ──
    row = 4
    first_data_row = row
    no = 0
    cust_groups = []  # [(start, end, cust_type)]

    # 구분별 정렬: 기업 → 일반업장 → 공공기관 → 학교
    type_order = {'기업': 0, '일반업장': 1, '공공기관': 2, '학교': 3}
    sorted_custs = sorted(
        customers_dict.items(),
        key=lambda x: (type_order.get(x[1].get('구분', '학교'), 9), x[0])
    )

    for cust_name, cust_info in sorted_custs:
        ctype = cust_info.get('구분', '학교')
        vat_apply = VAT_RULE.get(ctype, False)
        fill = _CUST_FILL.get(ctype)

        # 해당 거래처의 품목 목록 (단가 > 0인 것)
        items = []
        for label, price_key in ITEM_PRICE_MAP:
            price = float(cust_info.get(price_key, 0) or 0)
            if price > 0:
                # 아이템 타입 매칭 키들
                match_keys = []
                if price_key == 'price_food':
                    match_keys = ['food_waste', '음식물', '음식물쓰레기']
                elif price_key == 'price_recycle':
                    match_keys = ['recycle', '재활용']
                elif price_key == 'price_general':
                    match_keys = ['general', '사업장', '사업장폐기물', '일반폐기물']

                # 해당 거래처+품목의 누적 수거량
                total_kg = 0
                for mk in match_keys:
                    total_kg += agg.get((cust_name, mk), 0)

                items.append((label, price, total_kg))

        # 단가 등록된 품목이 없으면 음식물 기본행 1개
        if not items:
            items = [('음식물', 0, 0)]

        start = row
        for item_label, price, weight in items:
            no += 1
            ws1.cell(row=row, column=1, value=no).font = _FONT_N
            ws1.cell(row=row, column=1).alignment = _C
            ws1.cell(row=row, column=2, value=ctype).font = _FONT_N
            ws1.cell(row=row, column=2).alignment = _C
            if fill:
                ws1.cell(row=row, column=2).fill = fill
            ws1.cell(row=row, column=3, value=cust_name).font = _FONT_N
            ws1.cell(row=row, column=4, value=item_label).font = _FONT_N
            ws1.cell(row=row, column=4).alignment = _C
            ws1.cell(row=row, column=5, value=weight).font = _FONT_N
            ws1.cell(row=row, column=5).number_format = '#,##0'
            ws1.cell(row=row, column=6, value=price).font = _FONT_BLUE
            ws1.cell(row=row, column=6).number_format = '#,##0'
            ws1.cell(row=row, column=7, value=f'=E{row}*F{row}').font = _FONT_N
            ws1.cell(row=row, column=7).number_format = '#,##0'
            if vat_apply:
                ws1.cell(row=row, column=8, value=f'=G{row}*0.1').font = _FONT_N
            else:
                ws1.cell(row=row, column=8, value=0).font = _FONT_N
            ws1.cell(row=row, column=8).number_format = '#,##0'
            ws1.cell(row=row, column=9, value=f'=G{row}+H{row}').font = _FONT_N
            ws1.cell(row=row, column=9).number_format = '#,##0'

            for col in range(1, 10):
                ws1.cell(row=row, column=col).border = _THIN
                if col >= 5:
                    ws1.cell(row=row, column=col).alignment = _R
            row += 1

        end = row - 1
        cust_groups.append((start, end, ctype))

        # 거래처 내 여러 품목이면 구분+거래처명 셀 병합
        if end > start:
            ws1.merge_cells(f'B{start}:B{end}')
            ws1.merge_cells(f'C{start}:C{end}')
            ws1.cell(row=start, column=2).alignment = _C
            ws1.cell(row=start, column=3).alignment = Alignment(
                horizontal='left', vertical='center'
            )

    last_data_row = row - 1

    # ── 구분별 소계 ──
    subtotal_rows = {}
    for ctype_label in ['기업', '일반업장', '공공기관', '학교']:
        rows_of_type = [
            r for s, e, ct in cust_groups if ct == ctype_label
            for r in range(s, e + 1)
        ]
        if not rows_of_type:
            continue

        ws1.merge_cells(f'A{row}:D{row}')
        ws1.cell(row=row, column=1, value=f'{ctype_label} 소계').font = _FONT_H
        ws1.cell(row=row, column=1).alignment = _C
        fill = _CUST_FILL.get(ctype_label, _FILL_TOTAL)
        for col in range(1, 10):
            ws1.cell(row=row, column=col).fill = fill
            ws1.cell(row=row, column=col).border = _THIN

        sum_e = '+'.join([f'E{r}' for r in rows_of_type])
        ws1.cell(row=row, column=5, value=f'={sum_e}').font = _FONT_H
        ws1.cell(row=row, column=5).number_format = '#,##0'
        sum_g = '+'.join([f'G{r}' for r in rows_of_type])
        ws1.cell(row=row, column=7, value=f'={sum_g}').font = _FONT_H
        ws1.cell(row=row, column=7).number_format = '#,##0'
        sum_h = '+'.join([f'H{r}' for r in rows_of_type])
        ws1.cell(row=row, column=8, value=f'={sum_h}').font = _FONT_H
        ws1.cell(row=row, column=8).number_format = '#,##0'
        sum_i = '+'.join([f'I{r}' for r in rows_of_type])
        ws1.cell(row=row, column=9, value=f'={sum_i}').font = _FONT_H
        ws1.cell(row=row, column=9).number_format = '#,##0'

        subtotal_rows[ctype_label] = row
        row += 1

    # ── 총합계 ──
    ws1.merge_cells(f'A{row}:D{row}')
    ws1.cell(row=row, column=1, value='총 합계').font = _FONT_H
    ws1.cell(row=row, column=1).alignment = _C
    for col in range(1, 10):
        ws1.cell(row=row, column=col).fill = _FILL_TOTAL
        ws1.cell(row=row, column=col).border = _THIN
    ws1.cell(row=row, column=5,
             value=f'=SUM(E{first_data_row}:E{last_data_row})').font = _FONT_H
    ws1.cell(row=row, column=5).number_format = '#,##0'
    ws1.cell(row=row, column=7,
             value=f'=SUM(G{first_data_row}:G{last_data_row})').font = _FONT_H
    ws1.cell(row=row, column=7).number_format = '#,##0'
    ws1.cell(row=row, column=8,
             value=f'=SUM(H{first_data_row}:H{last_data_row})').font = _FONT_H
    ws1.cell(row=row, column=8).number_format = '#,##0'
    ws1.cell(row=row, column=9,
             value=f'=SUM(I{first_data_row}:I{last_data_row})').font = _FONT_H
    ws1.cell(row=row, column=9).number_format = '#,##0'
    total_row_rev = row

    # 열 너비
    for i, w in enumerate([6, 10, 14, 12, 14, 12, 14, 12, 14], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # AutoFilter (구분 컬럼 필터)
    ws1.auto_filter.ref = f'A3:I{last_data_row}'

    # ═══════════════════════════════════════
    # Sheet 2: 지출내역
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("지출내역")

    ws2.merge_cells('A1:E1')
    ws2['A1'] = f'( {month} )月  월말정산서 — 지출내역'
    ws2['A1'].font = _FONT_TITLE
    ws2['A1'].alignment = _C

    exp_headers = ['No', '지출항목', '금액(원)', '결제일', '비고']
    for ci, h in enumerate(exp_headers, 1):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = _FONT_W
        c.fill = _FILL_HDR
        c.alignment = _C
        c.border = _THIN

    exp_row = 4
    if expenses:
        for i, e in enumerate(expenses, 1):
            ws2.cell(row=exp_row, column=1, value=i).font = _FONT_N
            ws2.cell(row=exp_row, column=1).alignment = _C
            ws2.cell(row=exp_row, column=2, value=e.get('item', '')).font = _FONT_N
            ws2.cell(row=exp_row, column=3,
                     value=e.get('amount', 0)).font = _FONT_RED
            ws2.cell(row=exp_row, column=3).number_format = '#,##0;[Red](#,##0);"-"'
            ws2.cell(row=exp_row, column=3).alignment = _R
            ws2.cell(row=exp_row, column=4,
                     value=e.get('date', '')).font = _FONT_N
            ws2.cell(row=exp_row, column=4).alignment = _C
            _apply_border_row(ws2, exp_row, 1, 5)
            exp_row += 1
    else:
        # 빈 템플릿 10행
        for i in range(1, 11):
            ws2.cell(row=exp_row, column=1, value=i).font = _FONT_N
            ws2.cell(row=exp_row, column=1).alignment = _C
            _apply_border_row(ws2, exp_row, 1, 5)
            exp_row += 1

    last_exp_row = exp_row - 1

    # 지출합계
    ws2.merge_cells(f'A{exp_row}:B{exp_row}')
    ws2.cell(row=exp_row, column=1, value='지출 합계').font = _FONT_H
    ws2.cell(row=exp_row, column=1).alignment = _C
    ws2.cell(row=exp_row, column=3,
             value=f'=SUM(C4:C{last_exp_row})').font = _FONT_H
    ws2.cell(row=exp_row, column=3).number_format = '#,##0;[Red](#,##0);"-"'
    for col in range(1, 6):
        ws2.cell(row=exp_row, column=col).fill = _FILL_TOTAL
        ws2.cell(row=exp_row, column=col).border = _THIN
    exp_total_row = exp_row

    # 순수익 (매출 - 지출)
    exp_row += 2
    ws2.merge_cells(f'A{exp_row}:B{exp_row}')
    ws2.cell(row=exp_row, column=1, value='순수익 (매출 - 지출)').font = _FONT_BIG_B
    ws2.cell(row=exp_row, column=1).alignment = _C
    ws2.cell(row=exp_row, column=3,
             value=f"=수입내역!I{total_row_rev}+C{exp_total_row}").font = _FONT_BIG_B
    ws2.cell(row=exp_row, column=3).number_format = '#,##0;[Red](#,##0);"-"'
    for col in range(1, 6):
        ws2.cell(row=exp_row, column=col).border = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('double'), bottom=Side('double')
        )

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12

    # ── 저장 → bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════
# HTML 미리보기 (엑셀과 동일한 레이아웃)
# ═══════════════════════════════════════════════════

_CTYPE_BG = {
    '학교':     '#DAEEF3',
    '공공기관':  '#E2EFDA',
    '기업':     '#FCE4D6',
    '일반업장':  '#F2DCDB',
}


def generate_settlement_html(month, customers_dict, collection_rows):
    """
    월말정산 수입내역을 엑셀과 동일한 HTML 테이블로 반환.

    Returns
    -------
    str  (HTML)
    """
    # ── 누적 수거량 집계 ──
    agg = {}
    for r in collection_rows:
        sn = r.get('school_name', '')
        it_raw = r.get('item_type', '음식물')
        w = float(r.get('weight', 0) or 0)
        agg[(sn, it_raw)] = agg.get((sn, it_raw), 0) + w

    # ── 거래처별 행 데이터 구성 ──
    type_order = {'기업': 0, '일반업장': 1, '공공기관': 2, '학교': 3}
    sorted_custs = sorted(
        customers_dict.items(),
        key=lambda x: (type_order.get(x[1].get('구분', '학교'), 9), x[0])
    )

    table_rows = []       # [(cust_name, ctype, item_label, kg, price, supply, vat, total, rowspan)]
    subtotals = {}        # {ctype: {kg, supply, vat, total}}
    grand = {'kg': 0, 'supply': 0, 'vat': 0, 'total': 0}

    for cust_name, cust_info in sorted_custs:
        ctype = cust_info.get('구분', '학교')
        vat_apply = VAT_RULE.get(ctype, False)

        items = []
        for label, price_key in ITEM_PRICE_MAP:
            price = float(cust_info.get(price_key, 0) or 0)
            if price > 0:
                match_keys = []
                if price_key == 'price_food':
                    match_keys = ['food_waste', '음식물', '음식물쓰레기']
                elif price_key == 'price_recycle':
                    match_keys = ['recycle', '재활용']
                elif price_key == 'price_general':
                    match_keys = ['general', '사업장', '사업장폐기물', '일반폐기물']
                total_kg = sum(agg.get((cust_name, mk), 0) for mk in match_keys)
                items.append((label, price, total_kg))

        if not items:
            items = [('음식물', 0, 0)]

        rowspan = len(items)
        first = True
        for item_label, price, kg in items:
            supply = round(kg * price)
            vat = round(supply * 0.1) if vat_apply else 0
            total = supply + vat

            if ctype not in subtotals:
                subtotals[ctype] = {'kg': 0, 'supply': 0, 'vat': 0, 'total': 0}
            subtotals[ctype]['kg'] += kg
            subtotals[ctype]['supply'] += supply
            subtotals[ctype]['vat'] += vat
            subtotals[ctype]['total'] += total
            grand['kg'] += kg
            grand['supply'] += supply
            grand['vat'] += vat
            grand['total'] += total

            table_rows.append({
                'name': cust_name if first else None,
                'ctype': ctype if first else None,
                'rowspan': rowspan if first else 0,
                'item': item_label,
                'kg': kg,
                'price': price,
                'supply': supply,
                'vat': vat,
                'total': total,
                'vat_apply': vat_apply,
            })
            first = False

    # ── HTML 생성 ──
    css = """
<style>
.stbl{border-collapse:collapse;width:100%;font-family:'맑은 고딕',sans-serif;font-size:13px}
.stbl th{background:#4472C4;color:#fff;padding:6px 8px;text-align:center;border:1px solid #ccc;font-weight:bold}
.stbl td{padding:5px 8px;border:1px solid #ddd;vertical-align:middle}
.stbl .r{text-align:right}.stbl .c{text-align:center}
.stbl .sub{font-weight:bold}.stbl .grand{font-weight:bold;background:#FFF2CC}
.stbl .vat-yes{color:#c00}.stbl .vat-no{color:#888}
.stbl .blue{color:#0000FF}
</style>
"""
    h = css
    h += f'<h4 style="text-align:center;margin-bottom:8px">( {month} )月 월말정산서 — 수입내역</h4>\n'
    h += '<table class="stbl"><thead><tr>'
    for hdr in ['No', '구분', '거래처명', '품목', '누적수거량(kg)',
                '단가(원)', '공급가(원)', '부가세(원)', '정산금액(원)']:
        h += f'<th>{hdr}</th>'
    h += '</tr></thead><tbody>\n'

    no = 0
    for row in table_rows:
        no += 1
        bg = _CTYPE_BG.get(row.get('ctype', ''), '')
        h += '<tr>'
        h += f'<td class="c">{no}</td>'
        if row['rowspan'] > 0:
            rs = f' rowspan="{row["rowspan"]}"' if row['rowspan'] > 1 else ''
            bg_style = f' style="background:{bg}"' if bg else ''
            h += f'<td class="c"{rs}{bg_style}>{row["ctype"]}</td>'
            h += f'<td{rs}>{row["name"]}</td>'
        h += f'<td class="c">{row["item"]}</td>'
        h += f'<td class="r">{row["kg"]:,.0f}</td>'
        h += f'<td class="r blue">{row["price"]:,.0f}</td>'
        h += f'<td class="r">{row["supply"]:,.0f}</td>'
        vat_cls = 'vat-yes' if row['vat_apply'] else 'vat-no'
        vat_label = f'{row["vat"]:,.0f}' if row['vat_apply'] else '면세'
        h += f'<td class="r {vat_cls}">{vat_label}</td>'
        h += f'<td class="r">{row["total"]:,.0f}</td>'
        h += '</tr>\n'

    # 구분별 소계
    for ctype_label in ['기업', '일반업장', '공공기관', '학교']:
        if ctype_label not in subtotals:
            continue
        s = subtotals[ctype_label]
        bg = _CTYPE_BG.get(ctype_label, '')
        h += f'<tr class="sub" style="background:{bg}">'
        h += f'<td colspan="4" class="c">{ctype_label} 소계</td>'
        h += f'<td class="r">{s["kg"]:,.0f}</td><td></td>'
        h += f'<td class="r">{s["supply"]:,.0f}</td>'
        h += f'<td class="r">{s["vat"]:,.0f}</td>'
        h += f'<td class="r">{s["total"]:,.0f}</td>'
        h += '</tr>\n'

    # 총합계
    h += '<tr class="grand">'
    h += '<td colspan="4" class="c">총 합계</td>'
    h += f'<td class="r">{grand["kg"]:,.0f}</td><td></td>'
    h += f'<td class="r">{grand["supply"]:,.0f}</td>'
    h += f'<td class="r">{grand["vat"]:,.0f}</td>'
    h += f'<td class="r">{grand["total"]:,.0f}</td>'
    h += '</tr>\n'

    h += '</tbody></table>'
    return h, grand


def generate_expense_html(month, expense_rows, revenue_total=0):
    """
    월말정산 지출내역을 엑셀과 동일한 HTML 테이블로 반환.

    Parameters
    ----------
    month : int
    expense_rows : list[dict]
        DB expenses 테이블 행 [{item, amount, pay_date, memo}, ...]
    revenue_total : float
        수입내역 총 정산금액 (순수익 계산용)

    Returns
    -------
    str  (HTML)
    """
    css = """
<style>
.etbl{border-collapse:collapse;width:100%;font-family:'맑은 고딕',sans-serif;font-size:13px}
.etbl th{background:#4472C4;color:#fff;padding:6px 8px;text-align:center;border:1px solid #ccc;font-weight:bold}
.etbl td{padding:5px 8px;border:1px solid #ddd;vertical-align:middle}
.etbl .r{text-align:right}.etbl .c{text-align:center}
.etbl .red{color:#c00}.etbl .grand{font-weight:bold;background:#FFF2CC}
.etbl .profit{font-weight:bold;font-size:14px;border-top:3px double #333;border-bottom:3px double #333}
</style>
"""
    h = css
    h += f'<h4 style="text-align:center;margin-bottom:8px">( {month} )月 월말정산서 — 지출내역</h4>\n'
    h += '<table class="etbl"><thead><tr>'
    for hdr in ['No', '지출항목', '금액(원)', '결제일', '비고']:
        h += f'<th>{hdr}</th>'
    h += '</tr></thead><tbody>\n'

    total_exp = 0
    for i, e in enumerate(expense_rows, 1):
        amt = float(e.get('amount', 0) or 0)
        total_exp += amt
        amt_str = f'{abs(amt):,.0f}'
        if amt < 0:
            amt_str = f'({abs(amt):,.0f})'
        h += '<tr>'
        h += f'<td class="c">{i}</td>'
        h += f'<td>{e.get("item", "")}</td>'
        h += f'<td class="r red">{amt_str}</td>'
        h += f'<td class="c">{e.get("pay_date", "")}</td>'
        h += f'<td>{e.get("memo", "")}</td>'
        h += '</tr>\n'

    if not expense_rows:
        h += '<tr><td colspan="5" class="c" style="color:#888;padding:12px">'
        h += '등록된 지출 내역이 없습니다.</td></tr>\n'

    # 지출합계
    exp_str = f'{abs(total_exp):,.0f}'
    if total_exp < 0:
        exp_str = f'({abs(total_exp):,.0f})'
    h += '<tr class="grand">'
    h += '<td colspan="2" class="c">지출 합계</td>'
    h += f'<td class="r red">{exp_str}</td>'
    h += '<td></td><td></td></tr>\n'

    # 순수익
    profit = revenue_total + total_exp
    profit_color = '#c00' if profit < 0 else '#006600'
    profit_str = f'{abs(profit):,.0f}'
    if profit < 0:
        profit_str = f'({abs(profit):,.0f})'
    h += '<tr class="profit">'
    h += '<td colspan="2" class="c">순수익 (매출 - 지출)</td>'
    h += f'<td class="r" style="color:{profit_color};font-size:14px">{profit_str}</td>'
    h += '<td></td><td></td></tr>\n'

    h += '</tbody></table>'
    return h
